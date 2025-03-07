import customtkinter as ctk
from datetime import datetime, timedelta
import customtkinter as ctk
from tkcalendar import DateEntry
from PIL import Image,ImageTk
from tkinter import ttk, messagebox
import mysql.connector
from mysql.connector import Error,connect
import win32print
import win32ui
from win32con import *
from datetime import datetime, date, time
import re
import os
import uuid
from datetime import datetime, timedelta
from math import ceil
import tkinter as tk
from tkinter import messagebox, filedialog
from tkcalendar import DateEntry
import pandas as pd
def create_connection():
    try:
        connection = mysql.connector.connect(
            host='localhost',
            database='pharmacy',
            user='root',
            password='123@123'
        )
        return connection
    except Error as e:
        print(f"Error connecting to MySQL Database: {e}")
        return None
def execute_query(query, params=None):
    connection = create_connection()
    if connection is None:
        return

    try:
        cursor = connection.cursor()
        if params:
            cursor.execute(query, params)
        else:
            cursor.execute(query)
        connection.commit()
        return cursor
    except Error as e:
        print(f"Error executing query: {e}")
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()
def fetch_data(query, params=None):
    connection = create_connection()
    if connection is None:
        return None

    try:
        cursor = connection.cursor(dictionary=True)
        if params:
            cursor.execute(query, params)
        else:
            cursor.execute(query)
        return cursor.fetchall()
    except Error as e:
        print(f"Error fetching data: {e}")
        return None
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()
def reset_grid_configuration(widget):
    # Reset column and row configurations
    for col in range(widget.grid_size()[0]):  # Reset all columns
        widget.grid_columnconfigure(col, weight=0, minsize=0)
    for row in range(widget.grid_size()[1]):  # Reset all rows
        widget.grid_rowconfigure(row, weight=0, minsize=0)

def finance():
    clear_widgets(root)
    
    # Create main container with gradient background
    container = ctk.CTkFrame(
        root,
        fg_color=("#EEF2FF", "#1a1a2e")  # Light purple/blue gradient
    )
    container.pack(expand=True, fill="both", padx=40, pady=30)
    
    # Decorative header section
    header_frame = ctk.CTkFrame(container, fg_color="transparent")
    header_frame.pack(fill="x", pady=(20, 40))
    
    # Main title with modern styling
    title_label = ctk.CTkLabel(
        header_frame, 
        text="Finance",
        font=("Helvetica", 32, "bold"),
        text_color=("#1e40af", "#93c5fd")  # Dark blue for light mode, light blue for dark mode
    )
    title_label.pack()
    
    # Subtitle with softer color
    subtitle_label = ctk.CTkLabel(
        header_frame,
        text="Select an option to manage your bill",
        font=("Helvetica", 16),
        text_color=("#6b7280", "#9ca3af")  # Gray that works in both modes
    )
    subtitle_label.pack(pady=(5, 0))
    
    # Create a card-like frame for buttons
    card_frame = ctk.CTkFrame(
        container,
        fg_color=("#ffffff", "#1e1e2d"),  # White for light mode, dark blue for dark mode
        corner_radius=20
    )
    card_frame.pack(padx=40, pady=20, ipady=30, fill="x")
    
    # Button configurations
    button_width = 300
    button_height = 50
    button_padding = 15
    
    # Add Medicine Button with icon-like prefix
    add_button = ctk.CTkButton(
        card_frame,
        text="➕ Add New Bill",
        width=button_width,
        height=button_height,
        fg_color=("#3b82f6", "#2563eb"),  # Blue that works in both modes
        text_color="white",
        font=("Helvetica", 18),
        corner_radius=15,
        hover_color=("#2563eb", "#1d4ed8"),  # Slightly darker blue on hover
        command=show_add_transaction
    )
    add_button.pack(pady=button_padding)
        # Edit Medicine Button
    edit_button = ctk.CTkButton(
        card_frame,
        text="📝 Edit Bill",
        width=button_width,
        height=button_height,
        fg_color=("#3b82f6", "#2563eb"),
        text_color="white",
        font=("Helvetica", 18),
        corner_radius=15,
        hover_color=("#2563eb", "#1d4ed8"),
        command=show_edit_transaction
    )
    edit_button.pack(pady=button_padding)
    # Edit Medicine Button
    edit_button = ctk.CTkButton(
        card_frame,
        text="📋 Bill Report",
        width=button_width,
        height=button_height,
        fg_color=("#3b82f6", "#2563eb"),
        text_color="white",
        font=("Helvetica", 18),
        corner_radius=15,
        hover_color=("#2563eb", "#1d4ed8"),
        command=show_transactions_report
    )
    edit_button.pack(pady=button_padding)
    
    # Back Button with refined styling
    back_button = ctk.CTkButton(
        container,
        text="← Back to main menu",
        width=button_width,
        height=40,
        fg_color="transparent",
        text_color=("#6b7280", "#9ca3af"),
        font=("Helvetica", 16),
        border_width=2,
        border_color=("#d1d5db", "#4b5563"),
        corner_radius=15,
        hover_color=("#f3f4f6", "#1f2937"),
        command=show_main_menu
    )
    back_button.pack(pady=(40, 20))
def show_return_management():
    clear_widgets(root)
    
    # Create main container with gradient background
    container = ctk.CTkFrame(
        root,
        fg_color=("#EEF2FF", "#1a1a2e")  # Light purple/blue gradient
    )
    container.pack(expand=True, fill="both", padx=40, pady=30)
    
    # Decorative header section
    header_frame = ctk.CTkFrame(container, fg_color="transparent")
    header_frame.pack(fill="x", pady=(20, 40))
    
    # Main title with modern styling
    title_label = ctk.CTkLabel(
        header_frame, 
        text="Return",
        font=("Helvetica", 32, "bold"),
        text_color=("#1e40af", "#93c5fd")  # Dark blue for light mode, light blue for dark mode
    )
    title_label.pack()
    
    # Subtitle with softer color
    subtitle_label = ctk.CTkLabel(
        header_frame,
        text="Select an option to manage refund",
        font=("Helvetica", 16),
        text_color=("#6b7280", "#9ca3af")  # Gray that works in both modes
    )
    subtitle_label.pack(pady=(5, 0))
    
    # Create a card-like frame for buttons
    card_frame = ctk.CTkFrame(
        container,
        fg_color=("#ffffff", "#1e1e2d"),  # White for light mode, dark blue for dark mode
        corner_radius=20
    )
    card_frame.pack(padx=40, pady=20, ipady=30, fill="x")
    
    # Button configurations
    button_width = 300
    button_height = 50
    button_padding = 15
    
    # Add Medicine Button with icon-like prefix
    add_button = ctk.CTkButton(
        card_frame,
        text="➕ Add New Refund",
        width=button_width,
        height=button_height,
        fg_color=("#3b82f6", "#2563eb"),  # Blue that works in both modes
        text_color="white",
        font=("Helvetica", 18),
        corner_radius=15,
        hover_color=("#2563eb", "#1d4ed8"),  # Slightly darker blue on hover
        command=show_add_return
    )
    add_button.pack(pady=button_padding)
    
    # Edit Medicine Button
    edit_button = ctk.CTkButton(
        card_frame,
        text="📋 Medicine Refund Report",
        width=button_width,
        height=button_height,
        fg_color=("#3b82f6", "#2563eb"),
        text_color="white",
        font=("Helvetica", 18),
        corner_radius=15,
        hover_color=("#2563eb", "#1d4ed8"),
        command=show_returns_report
    )
    edit_button.pack(pady=button_padding)
    
    # Back Button with refined styling
    back_button = ctk.CTkButton(
        container,
        text="← Back",
        width=button_width,
        height=40,
        fg_color="transparent",
        text_color=("#6b7280", "#9ca3af"),
        font=("Helvetica", 16),
        border_width=2,
        border_color=("#d1d5db", "#4b5563"),
        corner_radius=15,
        hover_color=("#f3f4f6", "#1f2937"),
        command=show_medicine_management
    )
    back_button.pack(pady=(40, 20))
def show_profit_management():
    clear_widgets(root)
    
    # Create main container with gradient background
    container = ctk.CTkFrame(
        root,
        fg_color=("#EEF2FF", "#1a1a2e")  # Light purple/blue gradient
    )
    container.pack(expand=True, fill="both", padx=40, pady=30)
    
    # Decorative header section
    header_frame = ctk.CTkFrame(container, fg_color="transparent")
    header_frame.pack(fill="x", pady=(20, 40))
    
    # Main title with modern styling
    title_label = ctk.CTkLabel(
        header_frame, 
        text="Profit Management",
        font=("Helvetica", 32, "bold"),
        text_color=("#1e40af", "#93c5fd")  # Dark blue for light mode, light blue for dark mode
    )
    title_label.pack()
    
    # Subtitle with softer color
    subtitle_label = ctk.CTkLabel(
        header_frame,
        text="Select an option to check your profit",
        font=("Helvetica", 16),
        text_color=("#6b7280", "#9ca3af")  # Gray that works in both modes
    )
    subtitle_label.pack(pady=(5, 0))
    
    # Create a card-like frame for buttons
    card_frame = ctk.CTkFrame(
        container,
        fg_color=("#ffffff", "#1e1e2d"),  # White for light mode, dark blue for dark mode
        corner_radius=20
    )
    card_frame.pack(padx=40, pady=20, ipady=30, fill="x")
    
    # Button configurations
    button_width = 300
    button_height = 50
    button_padding = 15
    
    # Add Medicine Button with icon-like prefix
    add_button = ctk.CTkButton(
        card_frame,
        text="📰 Profit Report",
        width=button_width,
        height=button_height,
        fg_color=("#3b82f6", "#2563eb"),  # Blue that works in both modes
        text_color="white",
        font=("Helvetica", 18),
        corner_radius=15,
        hover_color=("#2563eb", "#1d4ed8"),  # Slightly darker blue on hover
        command=show_profit_report
    )
    add_button.pack(pady=button_padding)
    
    # Edit Medicine Button
    edit_button = ctk.CTkButton(
        card_frame,
        text="📰 Detail Profit Report",
        width=button_width,
        height=button_height,
        fg_color=("#3b82f6", "#2563eb"),
        text_color="white",
        font=("Helvetica", 18),
        corner_radius=15,
        hover_color=("#2563eb", "#1d4ed8"),
        command=show_daily_medicine_profit_report
    )
    edit_button.pack(pady=button_padding)
    add_button = ctk.CTkButton(
        card_frame,
        text="📰 Actual Profit Report",
        width=button_width,
        height=button_height,
        fg_color=("#3b82f6", "#2563eb"),  # Blue that works in both modes
        text_color="white",
        font=("Helvetica", 18),
        corner_radius=15,
        hover_color=("#2563eb", "#1d4ed8"),  # Slightly darker blue on hover
        command=show_actual_profit_report
    )
    add_button.pack(pady=button_padding)
    # Back Button with refined styling
    back_button = ctk.CTkButton(
        container,
        text="← Back",
        width=button_width,
        height=40,
        fg_color="transparent",
        text_color=("#6b7280", "#9ca3af"),
        font=("Helvetica", 16),
        border_width=2,
        border_color=("#d1d5db", "#4b5563"),
        corner_radius=15,
        hover_color=("#f3f4f6", "#1f2937"),
        command=show_reporting
    )
    back_button.pack(pady=(40, 20))
def show_medicine_management():
    clear_widgets(root)
    
    # Create main container with gradient background
    container = ctk.CTkFrame(
        root,
        fg_color=("#EEF2FF", "#1a1a2e")  # Light purple/blue gradient
    )
    container.pack(expand=True, fill="both", padx=40, pady=30)
    
    # Decorative header section
    header_frame = ctk.CTkFrame(container, fg_color="transparent")
    header_frame.pack(fill="x", pady=(20, 40))
    
    # Main title with modern styling
    title_label = ctk.CTkLabel(
        header_frame, 
        text="Medicine Management",
        font=("Helvetica", 32, "bold"),
        text_color=("#1e40af", "#93c5fd")  # Dark blue for light mode, light blue for dark mode
    )
    title_label.pack()
    
    # Subtitle with softer color
    subtitle_label = ctk.CTkLabel(
        header_frame,
        text="Select an option to manage your medicines",
        font=("Helvetica", 16),
        text_color=("#6b7280", "#9ca3af")  # Gray that works in both modes
    )
    subtitle_label.pack(pady=(5, 0))
    
    # Create a card-like frame for buttons
    card_frame = ctk.CTkFrame(
        container,
        fg_color=("#ffffff", "#1e1e2d"),  # White for light mode, dark blue for dark mode
        corner_radius=20
    )
    card_frame.pack(padx=40, pady=20, ipady=30, fill="x")
    
    # Button configurations
    button_width = 300
    button_height = 50
    button_padding = 15
    
    # Add Medicine Button with icon-like prefix
    add_button = ctk.CTkButton(
        card_frame,
        text="➕ Add New Medicine",
        width=button_width,
        height=button_height,
        fg_color=("#3b82f6", "#2563eb"),  # Blue that works in both modes
        text_color="white",
        font=("Helvetica", 18),
        corner_radius=15,
        hover_color=("#2563eb", "#1d4ed8"),  # Slightly darker blue on hover
        command=show_add_medicine
    )
    add_button.pack(pady=button_padding)
    
    # Edit Medicine Button
    edit_button = ctk.CTkButton(
        card_frame,
        text="📝 Edit Medicine",
        width=button_width,
        height=button_height,
        fg_color=("#3b82f6", "#2563eb"),
        text_color="white",
        font=("Helvetica", 18),
        corner_radius=15,
        hover_color=("#2563eb", "#1d4ed8"),
        command=show_edit_medicine
    )
    edit_button.pack(pady=button_padding)
    add_button = ctk.CTkButton(
        card_frame,
        text="📋 Return Management",
        width=button_width,
        height=button_height,
        fg_color=("#3b82f6", "#2563eb"),  # Blue that works in both modes
        text_color="white",
        font=("Helvetica", 18),
        corner_radius=15,
        hover_color=("#2563eb", "#1d4ed8"),  # Slightly darker blue on hover
        command=show_return_management
    )
    add_button.pack(pady=button_padding)
    # Back Button with refined styling
    back_button = ctk.CTkButton(
        container,
        text="← Back to Main Menu",
        width=button_width,
        height=40,
        fg_color="transparent",
        text_color=("#6b7280", "#9ca3af"),
        font=("Helvetica", 16),
        border_width=2,
        border_color=("#d1d5db", "#4b5563"),
        corner_radius=15,
        hover_color=("#f3f4f6", "#1f2937"),
        command=show_main_menu
    )
    back_button.pack(pady=(40, 20))
    
    # Add smooth hover animations
def show_inventory_management():
    clear_widgets(root)
    
    # Create main container with gradient background
    container = ctk.CTkFrame(
        root,
        fg_color=("#EEF2FF", "#1a1a2e")  # Light purple/blue gradient
    )
    container.pack(expand=True, fill="both", padx=40, pady=30)
    
    # Decorative header section
    header_frame = ctk.CTkFrame(container, fg_color="transparent")
    header_frame.pack(fill="x", pady=(20, 40))
    
    # Main title with modern styling
    title_label = ctk.CTkLabel(
        header_frame, 
        text="Inventory Management",
        font=("Helvetica", 32, "bold"),
        text_color=("#1e40af", "#93c5fd")  # Dark blue for light mode, light blue for dark mode
    )
    title_label.pack()
    
    # Subtitle with softer color
    subtitle_label = ctk.CTkLabel(
        header_frame,
        text="Select an option to manage your medicine inventory",
        font=("Helvetica", 16),
        text_color=("#6b7280", "#9ca3af")  # Gray that works in both modes
    )
    subtitle_label.pack(pady=(5, 0))
    
    # Create a card-like frame for buttons
    card_frame = ctk.CTkFrame(
        container,
        fg_color=("#ffffff", "#1e1e2d"),  # White for light mode, dark blue for dark mode
        corner_radius=20
    )
    card_frame.pack(padx=40, pady=20, ipady=30, fill="x")
    
    # Button configurations
    button_width = 300
    button_height = 50
    button_padding = 15
    
    # Add Medicine Button with icon-like prefix
    add_button = ctk.CTkButton(
        card_frame,
        text="📦 Medicine Stock",
        width=button_width,
        height=button_height,
        fg_color=("#3b82f6", "#2563eb"),  # Blue that works in both modes
        text_color="white",
        font=("Helvetica", 18),
        corner_radius=15,
        hover_color=("#2563eb", "#1d4ed8"),  # Slightly darker blue on hover
        command=show_medicine_stock
    )
    add_button.pack(pady=button_padding)
    
    # Edit Medicine Button
    edit_button = ctk.CTkButton(
        card_frame,
        text="💊 Medicine Expiry",
        width=button_width,
        height=button_height,
        fg_color=("#3b82f6", "#2563eb"),
        text_color="white",
        font=("Helvetica", 18),
        corner_radius=15,
        hover_color=("#2563eb", "#1d4ed8"),
        command=show_medicine_expiry
    )
    edit_button.pack(pady=button_padding)
    
    # Back Button with refined styling
    back_button = ctk.CTkButton(
        container,
        text="← Back to Main Menu",
        width=button_width,
        height=40,
        fg_color="transparent",
        text_color=("#6b7280", "#9ca3af"),
        font=("Helvetica", 16),
        border_width=2,
        border_color=("#d1d5db", "#4b5563"),
        corner_radius=15,
        hover_color=("#f3f4f6", "#1f2937"),
        command=show_main_menu
    )
    back_button.pack(pady=(40, 20))
    
def show_supplier_management():
    clear_widgets(root)
    
    # Create main container with gradient background
    container = ctk.CTkFrame(
        root,
        fg_color=("#EEF2FF", "#1a1a2e")  # Light purple/blue gradient
    )
    container.pack(expand=True, fill="both", padx=40, pady=30)
    
    # Decorative header section
    header_frame = ctk.CTkFrame(container, fg_color="transparent")
    header_frame.pack(fill="x", pady=(20, 40))
    
    # Main title with modern styling
    title_label = ctk.CTkLabel(
        header_frame, 
        text="Supplier Management",
        font=("Helvetica", 32, "bold"),
        text_color=("#1e40af", "#93c5fd")  # Dark blue for light mode, light blue for dark mode
    )
    title_label.pack()
    
    # Subtitle with softer color
    subtitle_label = ctk.CTkLabel(
        header_frame,
        text="Select an option to manage your supplier",
        font=("Helvetica", 16),
        text_color=("#6b7280", "#9ca3af")  # Gray that works in both modes
    )
    subtitle_label.pack(pady=(5, 0))
    
    # Create a card-like frame for buttons
    card_frame = ctk.CTkFrame(
        container,
        fg_color=("#ffffff", "#1e1e2d"),  # White for light mode, dark blue for dark mode
        corner_radius=20
    )
    card_frame.pack(padx=40, pady=20, ipady=30, fill="x")
    
    # Button configurations
    button_width = 300
    button_height = 50
    button_padding = 15
    
    # Add Medicine Button with icon-like prefix
    add_button = ctk.CTkButton(
        card_frame,
        text="➕ Add New Supplier",
        width=button_width,
        height=button_height,
        fg_color=("#3b82f6", "#2563eb"),  # Blue that works in both modes
        text_color="white",
        font=("Helvetica", 18),
        corner_radius=15,
        hover_color=("#2563eb", "#1d4ed8"),  # Slightly darker blue on hover
        command=show_add_supplier
    )
    add_button.pack(pady=button_padding)
    
    # Edit Medicine Button
    edit_button = ctk.CTkButton(
        card_frame,
        text="📝 Edit Supplier",
        width=button_width,
        height=button_height,
        fg_color=("#3b82f6", "#2563eb"),
        text_color="white",
        font=("Helvetica", 18),
        corner_radius=15,
        hover_color=("#2563eb", "#1d4ed8"),
        command=show_edit_supplier
    )
    edit_button.pack(pady=button_padding)
    add_button = ctk.CTkButton(
        card_frame,
        text="🪟 View Supplier",
        width=button_width,
        height=button_height,
        fg_color=("#3b82f6", "#2563eb"),  # Blue that works in both modes
        text_color="white",
        font=("Helvetica", 18),
        corner_radius=15,
        hover_color=("#2563eb", "#1d4ed8"),  # Slightly darker blue on hover
        command=show_suppliers_report
    )
    add_button.pack(pady=button_padding)
    
    # Back Button with refined styling
    back_button = ctk.CTkButton(
        container,
        text="← Back to Main Menu",
        width=button_width,
        height=40,
        fg_color="transparent",
        text_color=("#6b7280", "#9ca3af"),
        font=("Helvetica", 16),
        border_width=2,
        border_color=("#d1d5db", "#4b5563"),
        corner_radius=15,
        hover_color=("#f3f4f6", "#1f2937"),
        command=show_main_menu
    )
    back_button.pack(pady=(40, 20))

def clear_widgets(container):
    for widget in container.winfo_children():
        widget.destroy()
import mysql.connector  # or any other database library you are using

# Define your connection globally
connection = mysql.connector.connect(
    host='localhost',
    user='root',
    password='123@123',
    database='pharmacy'
)
def show_add_medicine():
    clear_widgets(root)
    
    # Create a main frame to hold everything
    main_frame = ctk.CTkFrame(root)
    main_frame.pack(fill="both", expand=True, padx=20, pady=20)

    # Create a scrollable frame inside the main frame
    scrollable_frame = ctk.CTkScrollableFrame(main_frame)
    scrollable_frame.pack(fill="both", expand=True)

    # Create a grid layout
    scrollable_frame.grid_columnconfigure(0, weight=1)
    scrollable_frame.grid_columnconfigure(1, weight=1)

    # Back button
    back_icon_image = ctk.CTkImage(Image.open("back.png"), size=(30, 30))
    back_button = ctk.CTkButton(
        scrollable_frame,
        image=back_icon_image,
        text="",
        width=40,
        command=show_medicine_management,
        fg_color="transparent",
        hover_color="#E5E7E9"
    )
    back_button.grid(row=0, column=0, padx=5, pady=5, sticky="w")

    # Title
    title_label = ctk.CTkLabel(scrollable_frame, text="Add New Medicine", font=("Arial", 24, "bold"),text_color="midnight blue")
    title_label.grid(row=0, column=0, columnspan=2, padx=350, pady=(0, 0))

    # Function to create label and entry pairs
    def create_label_entry(row, text, placeholder="", default_value=None, readonly=False):
        label = ctk.CTkLabel(scrollable_frame, text=text, font=("Arial", 14), anchor="w")
        label.grid(row=row, column=0, padx=(350, 10), pady=5, sticky="w")
        
        entry = ctk.CTkEntry(scrollable_frame, placeholder_text=placeholder, width=250, height=35, font=("Arial", 14))
        entry.grid(row=row, column=1, padx=(10, 90), pady=10, sticky="")
        
        # Set default value if provided
        if default_value:
            entry.insert(0, default_value)
        
        if readonly:
            entry.configure(state="readonly")
        
        return entry

        

    # Create all labels and entries
    entry_medicine_id = create_label_entry(2, "Medicine ID:", "1001")
    entry_supplier_id = create_label_entry(3, "Supplier ID:", "1")
    entry_company_name = create_label_entry(4, "Company Name:", "GlaxoSmithKline")
    entry_name = create_label_entry(5, "Medicine Name:", "Panadol")
    entry_no_of_packs = create_label_entry(6, "No of Packs:", default_value="1")
    entry_units_per_pack = create_label_entry(7, "Units per Pack:", "100")
    entry_buying_price_pack = create_label_entry(8, "Buying Price of Pack:", "500")
    entry_selling_price_pack = create_label_entry(9, "Selling Price of Pack:", "600")
    entry_quantity = create_label_entry(10, "Total Quantity:", "1000", readonly=True)
    entry_batch = create_label_entry(11, "Batch Number (Optional):", "102007")
    entry_rack = create_label_entry(12, "Rack No (Optional):", "25A")
    
    # Date entries
    ctk.CTkLabel(scrollable_frame, text="Expiry Date:", font=("Arial", 14), anchor="w").grid(row=13, column=0, padx=(350, 10), pady=10, sticky="w")
    entry_expiry = DateEntry(scrollable_frame, date_pattern='y-mm-dd', background='white', foreground='black', borderwidth=2, width=12, font=("Arial", 14))
    entry_expiry.grid(row=13, column=1, padx=(10, 90), pady=10, sticky="")

    ctk.CTkLabel(scrollable_frame, text="Store Date:", font=("Arial", 14), anchor="w").grid(row=14, column=0, padx=(350, 10), pady=10, sticky="w")
    entry_store = DateEntry(scrollable_frame, date_pattern='y-mm-dd', background='white', foreground='black', borderwidth=2, width=12, font=("Arial", 14))
    entry_store.grid(row=14, column=1, padx=(10, 90), pady=10, sticky="")

    entry_buying_price = create_label_entry(15, "Buying Price (per unit):", "5", readonly=True)
    entry_selling_price = create_label_entry(16, "Selling Price (per unit):", "6", readonly=True)

    # Supplier name label
    supplier_name_label = ctk.CTkLabel(scrollable_frame, text="Supplier Name: ", font=("Arial", 14, "italic"), anchor="w")
    supplier_name_label.grid(row=17, column=0, columnspan=2, padx=20, pady=(20, 10), sticky="w")
    
    def check_supplier_name(event=None):
        supplier_id = entry_supplier_id.get()

        if not supplier_id.isdigit():
            supplier_name_label.configure(text="Supplier Name: Invalid ID")
            return

        try:
            query = "SELECT name FROM Suppliers WHERE supplier_id = %s"
            cursor = connection.cursor()
            cursor.execute(query, (supplier_id,))
            result = cursor.fetchone()
            cursor.close()

            if result:
                supplier_name_label.configure(text=f"Supplier Name: {result[0]}")
            else:
                supplier_name_label.configure(text="Supplier Name: Not found")
        except Error as e:
            print(f"Error fetching supplier name: {e}")
            supplier_name_label.configure(text="Error fetching supplier name")

    entry_supplier_id.bind("<KeyRelease>", check_supplier_name)

    def calculate_prices_and_quantity(event=None):
        try:
            no_of_packs = int(entry_no_of_packs.get())
            units_per_pack = int(entry_units_per_pack.get())
            buying_price_pack = float(entry_buying_price_pack.get())
            selling_price_pack = float(entry_selling_price_pack.get())

            total_quantity = no_of_packs * units_per_pack
            buying_price_per_unit = buying_price_pack / units_per_pack
            selling_price_per_unit = selling_price_pack / units_per_pack

            entry_quantity.configure(state="normal")
            entry_quantity.delete(0, tk.END)
            entry_quantity.insert(0, str(total_quantity))
            entry_quantity.configure(state="readonly")

            entry_buying_price.configure(state="normal")
            entry_buying_price.delete(0, tk.END)
            entry_buying_price.insert(0, f"{buying_price_per_unit:.2f}")
            entry_buying_price.configure(state="readonly")

            entry_selling_price.configure(state="normal")
            entry_selling_price.delete(0, tk.END)
            entry_selling_price.insert(0, f"{selling_price_per_unit:.2f}")
            entry_selling_price.configure(state="readonly")

        except ValueError:
            # Handle invalid input
            pass

    # Bind the calculation function to the relevant entry fields
    entry_no_of_packs.bind("<KeyRelease>", calculate_prices_and_quantity)
    entry_units_per_pack.bind("<KeyRelease>", calculate_prices_and_quantity)
    entry_buying_price_pack.bind("<KeyRelease>", calculate_prices_and_quantity)
    entry_selling_price_pack.bind("<KeyRelease>", calculate_prices_and_quantity)

    def check_duplicate(connection, medicine_id, name):
        """Check if the serial number or medicine name already exists in the database."""
        query = '''
        SELECT COUNT(*) FROM Medicines
        WHERE medicine_id = %s OR name = %s
        '''
        params = (medicine_id, name)
        cursor = connection.cursor()
        cursor.execute(query, params)
        count = cursor.fetchone()[0]
        cursor.close()
        return count > 0  # Returns True if a duplicate is found

    def check_supplier_exists(connection, supplier_id):
        """Check if the supplier ID exists in the Suppliers table."""
        query = "SELECT COUNT(*) FROM Suppliers WHERE supplier_id = %s"
        cursor = connection.cursor()
        cursor.execute(query, (supplier_id,))
        count = cursor.fetchone()[0]
        cursor.close()
        return count > 0  # Returns True if supplier exists

    def add_medicine():
        medicine_id = entry_medicine_id.get()
        supplier_id = entry_supplier_id.get()
        company_name = entry_company_name.get()
        name = entry_name.get()
        quantity = entry_quantity.get()
        batch = entry_batch.get()
        rack = entry_rack.get()
        expiry = entry_expiry.get()
        store = entry_store.get()
        buying_price = entry_buying_price.get()
        selling_price = entry_selling_price.get()
        store_time = datetime.now().strftime('%H:%M:%S')

        # Validate inputs
        if not all([medicine_id, supplier_id, company_name, name, quantity, expiry, store, buying_price, selling_price]):
            messagebox.showerror("Error", "All fields except Batch Number and Rack No are required.")
            return

        if not medicine_id.isdigit() or not supplier_id.isdigit():
            messagebox.showerror("Error", "Serial number and Supplier ID must be integers.")
            return

        if not quantity.isdigit():
            messagebox.showerror("Error", "Quantity must be an integer.")
            return

        try:
            expiry_date = datetime.strptime(expiry, '%Y-%m-%d')
            if expiry_date <= datetime.now():
                messagebox.showerror("Error", "Expiry date must be in the future.")
                return
        except ValueError:
            messagebox.showerror("Error", "Invalid expiry date format. Please use YYYY-MM-DD.")
            return

        try:
            buying_price_float = float(buying_price)
            selling_price_float = float(selling_price)
        except ValueError:
            messagebox.showerror("Error", "Buying and Selling price must be valid numbers.")
            return

        # Check if the supplier ID exists
        if not check_supplier_exists(connection, supplier_id):
            messagebox.showerror("Error", "Invalid Supplier ID.")
            return

        # Check for duplicates
        if check_duplicate(connection, medicine_id, name):
            messagebox.showerror("Error", "A medicine with this Medicine ID or Name already exists.")
            return

        try:
            query = '''
            INSERT INTO Medicines (medicine_id, supplier_id, company_name, name, quantity, batch, rack, expiry, store, buying_price, selling_price, store_time)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            '''
            params = (medicine_id, supplier_id, company_name, name, int(quantity), batch, rack, expiry, store, buying_price_float, selling_price_float, store_time)
            execute_query(query, params)
            
            query = '''
            INSERT INTO snapshot_medicines (medicine_id, supplier_id, company_name, name, quantity, batch, rack, expiry, store, buying_price, selling_price, store_time)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            '''
            params = (medicine_id, supplier_id, company_name, name, int(quantity), batch, rack, expiry, store, buying_price_float, selling_price_float, store_time)
            execute_query(query, params)

            messagebox.showinfo("Success", "Medicine added successfully.")
            show_medicine_management()
        except Error as e:
            print(f"Error during database insertion: {e}")
            messagebox.showerror("Error", f"Failed to add medicine: {e}")

    # Create the Add Medicine button
    add_button = ctk.CTkButton(scrollable_frame, text="Add Medicine", command=add_medicine)
    add_button.grid(row=17, column=0, columnspan=2, pady=20, padx=10, sticky="ew")


def show_edit_medicine():
    clear_widgets(root)
    db_config = {
        'user': 'root',
        'password': '123@123',
        'host': 'localhost',
        'database': 'pharmacy',
        'raise_on_warnings': True
    }

    def fetch_and_display_medicine():
        medicine_id = entry_medicine_id.get()
        query = """
        SELECT medicine_id, supplier_id, company_name, name, quantity, batch, rack, expiry, store, buying_price, selling_price 
        FROM medicines 
        WHERE medicine_id = %s
        """
        result = fetch_data(query, (medicine_id,))

        if result and len(result) > 0:
            medicine = result[0]
            for i, entry in enumerate(entries.values()):
                if isinstance(entry, DateEntry):
                    if medicine[i+1]:
                        entry.set_date(medicine[i+1].strftime('%Y-%m-%d'))
                else:
                    entry.delete(0, 'end')
                    entry.insert(0, medicine[i+1] if medicine[i+1] is not None else "")
        else:
            messagebox.showerror("Error", "Medicine not found.")

    # Create a responsive main frame
    main_frame = ctk.CTkFrame(root)
    main_frame.pack(fill="both", expand=True, padx=20, pady=20)

    # Top frame for search and back button
    top_frame = ctk.CTkFrame(main_frame)
    top_frame.pack(fill="x", pady=(0, 20))

    back_icon_image = ctk.CTkImage(Image.open("back.png"), size=(30, 30))
    back_button = ctk.CTkButton(
        top_frame,
        image=back_icon_image,
        text="",
        width=40,
        command=show_medicine_management,
        fg_color="transparent",
        hover_color="#E5E7E9"
    )
    back_button.pack(side="left", padx=(0, 10))

    ctk.CTkLabel(top_frame, text="Enter Medicine ID:").pack(side="left", padx=(0, 5))
    entry_medicine_id = ctk.CTkEntry(top_frame, width=120)
    entry_medicine_id.pack(side="left", padx=(0, 10))
    ctk.CTkButton(top_frame, text="Search", command=fetch_and_display_medicine, width=80).pack(side="left")

    # Create two columns with responsive weights
    columns_frame = ctk.CTkFrame(main_frame)
    columns_frame.pack(fill="both", expand=True)
    columns_frame.grid_columnconfigure(0, weight=1)
    columns_frame.grid_columnconfigure(1, weight=1)

    left_column = ctk.CTkFrame(columns_frame)
    left_column.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
    
    right_column = ctk.CTkFrame(columns_frame)
    right_column.grid(row=0, column=1, sticky="nsew", padx=(10, 0))

    # Left column - Medicine details
    ctk.CTkLabel(left_column, text="Medicine Details", font=("Arial", 16, "bold")).pack(pady=(0, 10))

    fields = [
        ("Supplier ID", ctk.CTkEntry),
        ("Company Name", ctk.CTkEntry),
        ("Medicine Name", ctk.CTkEntry),
        ("Quantity", ctk.CTkEntry),
        ("Batch Number (Optional)", ctk.CTkEntry),
        ("Rack No (Optional)", ctk.CTkEntry),
        ("Expiry Date", DateEntry),
        ("Store Date", DateEntry),
        ("Buying Price", ctk.CTkEntry),
        ("Selling Price", ctk.CTkEntry)
    ]

    entries = {}
    for label, widget in fields:
        frame = ctk.CTkFrame(left_column)
        frame.pack(fill="x", pady=5)
        ctk.CTkLabel(frame, text=label, width=150, anchor="w").pack(side="left", padx=(0, 10))
        if widget == DateEntry:
            entry = widget(frame, date_pattern='y-mm-dd', background='white', foreground='black', borderwidth=2, font=('Arial', 12))
        else:
            entry = widget(frame, width=200)
        entry.pack(side="left", expand=True, fill="x")
        entries[label.lower().replace(" ", "_")] = entry

    # Right column - Pack Details Calculator
    ctk.CTkLabel(right_column, text="Pack Details Calculator", font=("Arial", 16, "bold")).pack(pady=(0, 10))

    calc_fields = [
        ("Number of packs", ctk.CTkEntry),
        ("Quantity per pack", ctk.CTkEntry),
        ("Buying price per pack", ctk.CTkEntry),
        ("Selling price per pack", ctk.CTkEntry)
    ]

    calc_entries = {}
    for label, widget in calc_fields:
        frame = ctk.CTkFrame(right_column)
        frame.pack(fill="x", pady=5)
        ctk.CTkLabel(frame, text=label, width=150, anchor="w").pack(side="left", padx=(0, 10))
        entry = widget(frame, width=150)
        entry.pack(side="left", expand=True, fill="x")
        calc_entries[label.lower().replace(" ", "_")] = entry

    def calculate_pack_details():
        try:
            num_packs = int(calc_entries["number_of_packs"].get())
            quantity_per_pack = int(calc_entries["quantity_per_pack"].get())
            buying_price_per_pack = float(calc_entries["buying_price_per_pack"].get())
            selling_price_per_pack = float(calc_entries["selling_price_per_pack"].get())
            
            total_quantity = num_packs * quantity_per_pack
            total_buying_price = buying_price_per_pack / quantity_per_pack
            total_selling_price = selling_price_per_pack / quantity_per_pack
            profit = total_selling_price - total_buying_price

            result_labels = [
                f"Total quantity: {total_quantity}",
                f"Total buying price: RS{total_buying_price:.2f}",
                f"Total selling price: RS{total_selling_price:.2f}",
                f"Profit: RS{profit:.2f}"
            ]

            for i, text in enumerate(result_labels):
                result_frame.grid_slaves(row=i, column=0)[0].configure(text=text)
        except ValueError:
            messagebox.showerror("Error", "Invalid input. Please enter valid numbers.")

    def apply_calculated_values():
        try:
            num_packs = int(calc_entries["number_of_packs"].get())
            quantity_per_pack = int(calc_entries["quantity_per_pack"].get())
            buying_price_per_pack = float(calc_entries["buying_price_per_pack"].get())
            selling_price_per_pack = float(calc_entries["selling_price_per_pack"].get())

            new_quantity = num_packs * quantity_per_pack
            buying_price_per_unit = buying_price_per_pack / quantity_per_pack
            selling_price_per_unit = selling_price_per_pack / quantity_per_pack

            current_quantity = int(entries["quantity"].get() or 0)
            updated_quantity = current_quantity + new_quantity
            
            entries["quantity"].delete(0, 'end')
            entries["quantity"].insert(0, str(updated_quantity))
            
            entries["buying_price"].delete(0, 'end')
            entries["buying_price"].insert(0, str(buying_price_per_unit))
            
            entries["selling_price"].delete(0, 'end')
            entries["selling_price"].insert(0, str(selling_price_per_unit))

            messagebox.showinfo("Success", "Calculated values applied to the main form.")
        except ValueError:
            messagebox.showerror("Error", "Invalid input. Please enter valid numbers.")
        except ZeroDivisionError:
            messagebox.showerror("Error", "Quantity per pack cannot be zero.")

    ctk.CTkButton(right_column, text="Calculate", command=calculate_pack_details).pack(pady=10)

    result_frame = ctk.CTkFrame(right_column)
    result_frame.pack(fill="x", pady=10)

    for i in range(4):
        result_frame.grid_rowconfigure(i, weight=1)
        ctk.CTkLabel(result_frame, text="").grid(row=i, column=0, sticky="w", pady=2)

    ctk.CTkButton(right_column, text="Apply to Main Form", command=apply_calculated_values).pack(pady=10)

    def update_medicine():
        medicine_id = entry_medicine_id.get()
        fields_data = [entries[key].get() for key in entries]

        if not all([medicine_id] + fields_data[:4] + fields_data[6:]):
            messagebox.showerror("Error", "All fields except Batch Number and Rack No are required.")
            return

        try:
            query = '''
            UPDATE medicines 
            SET supplier_id = %s, company_name = %s, name = %s, quantity = %s, batch = %s, rack = %s, 
                expiry = %s, store = %s, buying_price = %s, selling_price = %s
            WHERE medicine_id = %s
            '''
            params = tuple(fields_data + [medicine_id])
            execute_query(query, params)

            store_time = datetime.now().time().strftime('%H:%M:%S')

            with mysql.connector.connect(**db_config) as conn:
                cursor = conn.cursor()

                cursor.execute('''
                    SELECT COALESCE(MAX(version), 0) FROM snapshot_medicines 
                    WHERE medicine_id = %s
                ''', (medicine_id,))
                max_version = cursor.fetchone()[0]
                
                new_version = max_version + 1

                query = '''
                INSERT INTO snapshot_medicines (medicine_id, supplier_id, company_name, name, quantity, batch, rack, expiry, store, buying_price, selling_price, store_time, version)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                '''
                params = (medicine_id,) + tuple(fields_data) + (store_time, new_version)

                execute_query(query, params)

            messagebox.showinfo("Success", "Medicine updated successfully.")
            show_medicine_management()
        except mysql.connector.Error as e:
            print(f"An error occurred: {e}")
            messagebox.showerror("Database Error", "An error occurred while updating the medicine.")

    ctk.CTkButton(main_frame, text="Update Medicine", command=update_medicine).pack(pady=20)

    # Make the UI responsive
    def on_resize(event):
        # Adjust column weights based on window width
        if event.width < 800:
            columns_frame.grid_columnconfigure(0, weight=1)
            columns_frame.grid_columnconfigure(1, weight=0)
            right_column.grid_remove()
            left_column.grid(row=0, column=0, sticky="nsew", padx=0)
        else:
            columns_frame.grid_columnconfigure(0, weight=1)
            columns_frame.grid_columnconfigure(1, weight=1)
            right_column.grid()
            left_column.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
            right_column.grid(row=0, column=1, sticky="nsew", padx=(10, 0))

    root.bind("<Configure>", on_resize)
def show_add_supplier():
    clear_widgets(root)
    back_icon_image = ctk.CTkImage(Image.open("back.png"), size=(40, 40))

    search_frame = ctk.CTkFrame(root, fg_color="transparent")
    search_frame.pack(side="left",padx=10,pady=10,fill="y")
    back_button = ctk.CTkButton(
        search_frame,
        image=back_icon_image,
        text="",
        width=40,
        command=show_supplier_management,
        fg_color="transparent",
        hover_color="#E5E7E9"
    )
    back_button.grid(row=0, column=0, padx=(0, 5), pady=0, sticky="nw")
    
    ctk.CTkLabel(root, text="Supplier ID").pack(pady=0, padx=20)
    entry_supplier_id = ctk.CTkEntry(root, placeholder_text="1")  # Placeholder for Supplier ID
    entry_supplier_id.pack(pady=0, padx=20)

    ctk.CTkLabel(root, text="Supplier Name").pack(pady=0, padx=20)  # Supplier Name Label
    entry_supplier_name = ctk.CTkEntry(root, placeholder_text="Muhammad Ali")  # Placeholder for Supplier Name
    entry_supplier_name.pack(pady=0, padx=20)

    ctk.CTkLabel(root, text="Contact Person").pack(pady=0, padx=20)
    entry_contact_person = ctk.CTkEntry(root, placeholder_text="John Doe")
    entry_contact_person.pack(pady=0, padx=20)
    
    ctk.CTkLabel(root, text="Phone (Optional)").pack(pady=0, padx=20)
    entry_phone = ctk.CTkEntry(root, placeholder_text="212-456-7890")
    entry_phone.pack(pady=0, padx=20)
    
    ctk.CTkLabel(root, text="Email (Optional)").pack(pady=0, padx=20)
    entry_email = ctk.CTkEntry(root, placeholder_text="ali@gmail.com")
    entry_email.pack(pady=0, padx=20)
    
    ctk.CTkLabel(root, text="Address (Optional)").pack(pady=0, padx=20)
    entry_address = ctk.CTkEntry(root, placeholder_text="Street, Suite 3")
    entry_address.pack(pady=0, padx=20)
    
    def check_supplier_exists(connection, supplier_id, supplier_name):
        """Check if the supplier ID or supplier name already exists in the Suppliers table."""
        query = '''
        SELECT COUNT(*) FROM suppliers
        WHERE supplier_id = %s OR name = %s
        '''
        cursor = connection.cursor()
        cursor.execute(query, (supplier_id, supplier_name))
        count = cursor.fetchone()[0]
        connection.commit()
        cursor.close()

        return count > 0  # Returns True if supplier_id or name exists
        

    def add_supplier():
        supplier_id = entry_supplier_id.get()  # Get supplier ID
        supplier_name = entry_supplier_name.get()  # Get Supplier Name
        contact_person = entry_contact_person.get()  # Get Contact Person
        phone = entry_phone.get()  # Get Phone
        email = entry_email.get()  # Get Email
        address = entry_address.get()  # Get Address

        # Validate required fields
        if not supplier_id or not supplier_name:
            messagebox.showerror("Error", "Supplier ID and Supplier Name are required.")
            return

        # Validate that supplier ID is an integer
        if not supplier_id.isdigit():
            messagebox.showerror("Error", "Supplier ID must be an integer.")
            return

        # Check if supplier_id or supplier_name already exists in the database
        if check_supplier_exists(connection, supplier_id, supplier_name):
            messagebox.showerror("Error", "A supplier with this ID or Name already exists.")
            return

        # Validate that the supplier name contains only alphabetic characters
        if not supplier_name.replace(" ", "").isalpha():
            messagebox.showerror("Error", "Supplier Name must contain only alphabetic characters.")
            return

        # Validate that the contact person contains only alphabetic characters
        if not contact_person.replace(" ", "").isalpha():
            messagebox.showerror("Error", "Contact Person must contain only alphabetic characters.")
            return

        # Validate phone number if provided
        if phone and not phone.replace("-", "").isdigit():
            messagebox.showerror("Error", "Phone number must contain only numeric digits and hyphens.")
            return

        # Validate email format if provided
        if email:
            email_pattern = r'^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}$'
            if not re.match(email_pattern, email):
                messagebox.showerror("Error", "Invalid email format. Please enter a valid email.")
                return

        try:
            # Insert the new supplier into the `suppliers` table
            query = '''
            INSERT INTO suppliers (supplier_id, name, contact_person, phone, email, address)
            VALUES (%s, %s, %s, %s, %s, %s)
            '''
            params = (supplier_id, supplier_name, contact_person, phone or None, email or None, address or None)
            execute_query(query, params)

            messagebox.showinfo("Success", "Supplier added successfully.")
            show_supplier_management()
        except Error as e:
            print(f"Error during database insertion: {e}")
            messagebox.showerror("Error", f"Failed to add supplier: {e}")

    ctk.CTkButton(root, text="Add Supplier", command=add_supplier).pack(pady=10, padx=20)
def show_edit_supplier():
    clear_widgets(root)

    def search_supplier():
        supplier_id = entry_supplier_id.get()

        if not supplier_id.isdigit():
            messagebox.showerror("Error", "Supplier ID must be an integer.")
            return

        # Query to fetch supplier details
        query = "SELECT * FROM suppliers WHERE supplier_id = %s"
        cursor = connection.cursor()
        cursor.execute(query, (supplier_id,))
        supplier = cursor.fetchone()
        cursor.close()

        if supplier:
            # Populate entries with existing supplier details
            entry_supplier_name.delete(0, 'end')
            entry_supplier_name.insert(0, supplier[1])  # name
            entry_contact_person.delete(0, 'end')
            entry_contact_person.insert(0, supplier[2])  # contact_person
            entry_phone.delete(0, 'end')
            entry_phone.insert(0, supplier[3])  # phone
            entry_email.delete(0, 'end')
            entry_email.insert(0, supplier[4])  # email
            entry_address.delete(0, 'end')
            entry_address.insert(0, supplier[5])  # address
        else:
            messagebox.showerror("Error", "Supplier not found.")

    def update_supplier():
        supplier_id = entry_supplier_id.get()  # Get supplier ID
        supplier_name = entry_supplier_name.get()  # Get Supplier Name
        contact_person = entry_contact_person.get()  # Get Contact Person
        phone = entry_phone.get()  # Get Phone
        email = entry_email.get()  # Get Email
        address = entry_address.get()  # Get Address

        # Validate required fields
        if not all([supplier_id, supplier_name]):
            messagebox.showerror("Error", "Supplier ID and Supplier Name are required.")
            return

        if not supplier_id.isdigit():
            messagebox.showerror("Error", "Supplier ID must be an integer.")
            return

        try:
            # Update the supplier in the `suppliers` table
            query = '''
            UPDATE suppliers 
            SET name = %s, contact_person = %s, phone = %s, email = %s, address = %s
            WHERE supplier_id = %s
            '''
            params = (supplier_name, contact_person, phone, email, address, supplier_id)
            execute_query(query, params)
            connection.commit()  # Add this after the query execution


            messagebox.showinfo("Success", "Supplier updated successfully.")
            show_supplier_management()  # Navigate back to the main management screen
        except Error as e:
            print(f"Error during database update: {e}")
            messagebox.showerror("Error", f"Failed to update supplier: {e}")

    def delete_supplier():
        delete_window = ctk.CTkToplevel(root)
        delete_window.title("Delete Supplier")
        delete_window.geometry("300x300")
        delete_window.resizable(False, False)
        # Keep the delete window on top of the main window
        delete_window.attributes('-topmost', True)
        delete_window.grab_set()  # Restrict input to the pop-up

        ctk.CTkLabel(delete_window, text="Enter Supplier ID to delete:").pack(pady=10, padx=10)
        entry_delete_id = ctk.CTkEntry(delete_window)
        entry_delete_id.pack(pady=10, padx=10)

        def confirm_delete():
            supplier_id = entry_delete_id.get()
            if not supplier_id.isdigit():
                messagebox.showerror("Error", "Supplier ID must be an integer.")
                return

            # Check if the supplier ID exists
            try:
                query = "SELECT * FROM suppliers WHERE supplier_id = %s"
                cursor = connection.cursor()
                cursor.execute(query, (supplier_id,))
                supplier = cursor.fetchone()
                cursor.close()
                connection.commit()  # Add this after the query execution

                if not supplier:
                    messagebox.showerror("Error", "Supplier ID not found.")
                    return

                # Perform delete operation
                query = "DELETE FROM suppliers WHERE supplier_id = %s"
                execute_query(query, (supplier_id,))
                messagebox.showinfo("Success", "Supplier deleted successfully.")
                delete_window.destroy()  # Close the delete window
                show_supplier_management()
            except Error as e:
                print(f"Error during database operation: {e}")
                messagebox.showerror("Error", f"Failed to delete supplier: {e}")

        ctk.CTkButton(delete_window, text="Delete Supplier", command=confirm_delete).pack(pady=10, padx=10)
        ctk.CTkButton(delete_window, text="Cancel", command=delete_window.destroy).pack(pady=5, padx=10)



    back_icon_image = ctk.CTkImage(Image.open("back.png"), size=(40, 40))

    search_frame = ctk.CTkFrame(root, fg_color="transparent")
    search_frame.pack(side="left", padx=10, pady=10, fill="y")
    
    back_button = ctk.CTkButton(
        search_frame,
        image=back_icon_image,
        text="",
        width=40,
        command=show_supplier_management,
        fg_color="transparent",
        hover_color="#E5E7E9"
    )
    back_button.grid(row=0, column=0, padx=(0, 5), pady=0, sticky="nw")
    
    # Create Supplier ID input
    ctk.CTkLabel(root, text="Supplier ID").pack(pady=0, padx=20)
    entry_supplier_id = ctk.CTkEntry(root, placeholder_text="1")  # Placeholder for Supplier ID
    entry_supplier_id.pack(pady=0, padx=20)
    ctk.CTkButton(root, text="Search Supplier", command=search_supplier).pack(pady=10, padx=20)

    # Create labels and entries for supplier details
    ctk.CTkLabel(root, text="Supplier Name").pack(pady=0, padx=20)
    entry_supplier_name = ctk.CTkEntry(root)
    entry_supplier_name.pack(pady=0, padx=20)

    ctk.CTkLabel(root, text="Contact Person").pack(pady=0, padx=20)
    entry_contact_person = ctk.CTkEntry(root)
    entry_contact_person.pack(pady=0, padx=20)

    ctk.CTkLabel(root, text="Phone").pack(pady=0, padx=20)
    entry_phone = ctk.CTkEntry(root)
    entry_phone.pack(pady=0, padx=20)

    ctk.CTkLabel(root, text="Email").pack(pady=0, padx=20)
    entry_email = ctk.CTkEntry(root)
    entry_email.pack(pady=0, padx=20)

    ctk.CTkLabel(root, text="Address").pack(pady=0, padx=20)
    entry_address = ctk.CTkEntry(root)
    entry_address.pack(pady=0, padx=20)

    # Create Search and Update buttons
    ctk.CTkButton(root, text="Update Supplier", command=update_supplier).pack(pady=10, padx=20)

    # Add Delete Supplier button
    ctk.CTkButton(root, text="Delete Supplier", command=delete_supplier).pack(pady=10, padx=20)
def fetch_suppliers(search_term=""):
    try:
        query = "SELECT supplier_id, name, contact_person, phone, email, address FROM suppliers"
        if search_term:
            query += " WHERE name LIKE %s OR contact_person LIKE %s"
            search_term = f"%{search_term}%"
            cursor = connection.cursor()
            cursor.execute(query, (search_term, search_term))
            connection.commit()  # Add this after the query execution

        else:
            cursor = connection.cursor()
            cursor.execute(query)

        
        suppliers = cursor.fetchall()
        cursor.close()
        return suppliers
    except Error as e:
        print(f"Error fetching suppliers: {e}")
        return []

def show_suppliers_report(start_row=2, search_term=""):
    clear_widgets(root)
    reset_grid_configuration(root)
    
    # Configure column weights for proper alignment
    for col in range(6):  # Adjusted for the number of columns in suppliers
        root.grid_columnconfigure(col, weight=1)

    # Title Section
    title_label = ctk.CTkLabel(
        root, 
        text="Suppliers Report", 
        font=("Arial", 24, "bold"),
        text_color="#2E86C1"
    )
    title_label.grid(row=0, column=0, columnspan=6, pady=20, sticky="nsew")

    # Back button and search section
    back_icon_image = ctk.CTkImage(Image.open("back.png"), size=(30, 30))
    search_icon_image = ctk.CTkImage(Image.open("search-icon-lob.png"), size=(20, 20))
    
    # Search frame setup
    search_frame = ctk.CTkFrame(root, fg_color="transparent")
    search_frame.grid(row=0, column=0, padx=20, pady=10, sticky="nw")
    
    # Back button, search entry, and search button
    back_button = ctk.CTkButton(
        search_frame,
        image=back_icon_image,
        text="",
        width=40,
        command=show_supplier_management,  # Replace with your relevant function
        fg_color="transparent",
        hover_color="#E5E7E9"
    )
    back_button.grid(row=0, column=0, padx=(0, 5), pady=0, sticky="nw")

    search_entry = ctk.CTkEntry(
        search_frame,
        placeholder_text="Search Supplier",
        width=150,
        height=32,
        border_color="#2E86C1",
        corner_radius=8
    )
    search_entry.grid(row=0, column=1, padx=(0, 5), pady=6, sticky="nw")

    search_button = ctk.CTkButton(
        search_frame,
        image=search_icon_image,
        text="",
        width=40,
        command=lambda: show_suppliers_report(search_term=search_entry.get()),
        fg_color="transparent",
        hover_color="#E5E7E9"
    )
    search_button.grid(row=0, column=2, padx=0, pady=6, sticky="nw")

    # Header Section with corrected anchors
    headers_suppliers = [
        ("Supplier ID", 10, "center"),
        ("Name", 25, "center"),
        ("Contact Person", 20, "center"),
        ("Phone", 15, "center"),
        ("Email", 20, "center"),
        ("Address", 20, "center"),
    ]

    # Create header frame
    header_frame = ctk.CTkFrame(root, fg_color="transparent")
    header_frame.grid(row=1, column=0, columnspan=6, sticky="nsew", padx=10)
    
    # Configure header columns
    total_weight = sum(weight for _, weight, _ in headers_suppliers)
    for col, (_, weight, _) in enumerate(headers_suppliers):
        header_frame.grid_columnconfigure(col, weight=weight)

    # Create headers
    for col, (header, _, _) in enumerate(headers_suppliers):
        header_label = ctk.CTkLabel(
            header_frame,
            text=header,
            font=("Arial", 14, "bold"),
            text_color="#FFFFFF",
            fg_color="#2E86C1",
            corner_radius=8,
            height=35,
            anchor="center"
        )
        header_label.grid(row=0, column=col, padx=2, pady=10, sticky="nsew")

    # Scrollable frame for data
    scroll_frame = ctk.CTkScrollableFrame(
        root,
        height=600,
        width=800,
        fg_color="#F8F9F9",
        corner_radius=10
    )
    scroll_frame.grid(row=2, column=0, columnspan=6, padx=10, pady=10, sticky="nsew")

    # Configure scroll frame columns
    for col, (_, weight, _) in enumerate(headers_suppliers):
        scroll_frame.grid_columnconfigure(col, weight=weight)

    try:
        # Fetch suppliers data
        suppliers = fetch_suppliers(search_term)  # Use the fetch_suppliers function

        if suppliers:
            for i, supplier in enumerate(suppliers, start=start_row):
                if isinstance(supplier, tuple) and len(supplier) == 6:
                    supplier_id, name, contact_person, phone, email, address = supplier

                    # Alternating row colors
                    bg_color = "#F4F6F7" if i % 2 == 0 else "#FFFFFF"

                    # Display data with corrected anchors
                    fields = [
                        (str(supplier_id), "center"),
                        (name, "center"),
                        (contact_person, "center"),
                        (phone, "center"),
                        (email, "center"),
                        (address, "center")
                    ]
                    
                    for col, (field, align) in enumerate(fields):
                        supplier_label = ctk.CTkLabel(
                            scroll_frame,
                            text=field,
                            font=("Arial", 13),
                            text_color="#34495E",
                            fg_color="#FFFFFF" if i % 2 == 0 else "#F7F9F9",
                            corner_radius=8,
                            anchor=align,
                            height=30
                        )
                        # Add padding based on alignment
                        padx = (10, 5) if align == "w" else (5, 10) if align == "e" else (5, 5)
                        supplier_label.grid(
                            row=i-start_row,
                            column=col,
                            padx=padx,
                            pady=2,
                            sticky="nsew"
                        )
        else:
            ctk.CTkLabel(
                scroll_frame,
                text="No suppliers found.",
                font=("Arial", 13, "bold"),
                text_color="#E74C3C",
                anchor="center",
                height=40
            ).grid(row=0, column=0, columnspan=len(headers_suppliers), pady=20, sticky="nsew")

    except Exception as e:
        ctk.CTkLabel(
            scroll_frame,
            text=f"Error fetching suppliers data: {e}",
            font=("Arial", 13, "bold"),
            text_color="#E74C3C",
            anchor="center",
            height=40
        ).grid(row=0, column=0, columnspan=len(headers_suppliers), pady=20, sticky="nsew") 

def show_medicine_stock():
    clear_widgets(root)
    reset_grid_configuration(root)

    # Configure column weights for proper alignment
    for col in range(10):
        root.grid_columnconfigure(col, weight=1)

    # Title Section
    title_label = ctk.CTkLabel(
        root, 
        text="Medicine Stock Report", 
        font=("Arial", 24, "bold"),
        text_color="#2E86C1"
    )
    title_label.grid(row=0, column=0, columnspan=10, pady=20, sticky="nsew")

    # Back button and search section
    back_icon_image = ctk.CTkImage(Image.open("back.png"), size=(30, 30))
    
    # Search frame setup
    search_frame = ctk.CTkFrame(root, fg_color="transparent")
    search_frame.grid(row=0, column=0, padx=20, pady=10, sticky="nw")
    
    # Back button and search entry
    back_button = ctk.CTkButton(
        search_frame,
        image=back_icon_image,
        text="",
        width=40,
        command=show_inventory_management,
        fg_color="transparent",
        hover_color="#E5E7E9"
    )
    back_button.grid(row=0, column=0, padx=(0, 5), pady=0, sticky="nw")

    search_var = tk.StringVar()
    search_entry = ctk.CTkEntry(
        search_frame,
        placeholder_text="🔍 Search Medicine",
        width=200,
        height=32,
        border_color="#2E86C1",
        corner_radius=8,
        textvariable=search_var
    )
    search_entry.grid(row=0, column=1, padx=(0, 5), pady=6, sticky="nw")
    search_entry.focus_set()

    # Function to update the report based on search term
    def update_report(*args):
        search_term = search_var.get().strip()
        show_filtered_stock_report(filter_combobox.get(), search_term=search_term)

    # Bind the search_var to update_report function
    search_var.trace_add("write", update_report)

    # ComboBox for filtering options
    filter_options = ["All Medicines", "Low Stock", "Full Stock", "Out of Stock"]
    filter_combobox = ctk.CTkComboBox(
        root,
        values=filter_options,
        command=lambda choice: show_filtered_stock_report(choice, search_term=search_var.get()),
        width=150,
        height=32,
        border_color="#2E86C1",
        button_color="#2E86C1",
        dropdown_hover_color="#3498DB"
    )
    filter_combobox.set("All Medicines")
    filter_combobox.grid(row=0, column=9, padx=(0, 20), pady=10, sticky="e")

    # Label to display the total number of medicines
    total_medicines_label = ctk.CTkLabel(
        root,
        text="Total: 0",  # Default value, updated later
        font=("Arial", 14, "bold"),
        text_color="#2E86C1"
    )
    total_medicines_label.grid(row=0, column=8, padx=(0, 10), pady=10, sticky="e")

    # Header Section with fixed widths and proper alignment
    headers = [
        ("Medicine ID", 10),
        ("Supplier ID", 10),   # Supplier ID after Medicine ID
        ("Name", 25),
        ("Company Name", 25),  # Company Name after Medicine Name
        ("Quantity", 10),
        ("Expiry Date", 15),
        ("Store Date", 15),
        ("Rack No", 10)
    ]

    # Create a frame for headers to ensure proper alignment
    header_frame = ctk.CTkFrame(root, fg_color="transparent")
    header_frame.grid(row=1, column=0, columnspan=10, sticky="nsew", padx=10)
    
    # Configure header frame columns with proper weights
    total_weight = sum(weight for _, weight in headers)
    for col in range(len(headers)):
        header_frame.grid_columnconfigure(col, weight=headers[col][1])

    # Create headers with proper widths
    for col, (header, weight) in enumerate(headers):
        header_label = ctk.CTkLabel(
            header_frame,
            text=header,
            font=("Arial", 14, "bold"),
            text_color="#FFFFFF",
            fg_color="#2E86C1",
            corner_radius=8,
            height=35,
            anchor="center"
        )
        header_label.grid(
            row=0,
            column=col,
            padx=2,
            pady=10,
            sticky="nsew"
        )

    # Scrollable frame for data
    scroll_frame = ctk.CTkScrollableFrame(
        root,
        height=600,
        width=900,
        fg_color="#F8F9F9",
        corner_radius=10
    )
    scroll_frame.grid(row=2, column=0, columnspan=10, padx=10, pady=10, sticky="nsew")

    # Configure scroll frame columns with the same weights as headers
    for col in range(len(headers)):
        scroll_frame.grid_columnconfigure(col, weight=headers[col][1])

    # Function to show filtered stock report
    def show_filtered_stock_report(filter_choice, search_term=""):
        # Clear previous data in scroll frame
        for widget in scroll_frame.winfo_children():
            widget.destroy()

        try:
            query = """
                SELECT medicine_id, supplier_id, name, company_name, quantity, expiry, store, rack 
                FROM medicines 
                WHERE 1=1
            """
            filter_params = []

            if filter_choice == "Low Stock":
                query += " AND quantity > 0 AND quantity < 70"
            elif filter_choice == "Full Stock":
                query += " AND quantity >= 70"
            elif filter_choice == "Out of Stock":
                query += " AND quantity = 0"

            # If search term is provided, add it to the query
            if search_term:
                query += " AND name LIKE %s"
                filter_params.append(f"%{search_term}%")

            query += " ORDER BY quantity ASC"
            medicines = fetch_data(query, filter_params)

            # Update total medicines label
            total_medicines_label.configure(text=f"Total: {len(medicines)}")
            if medicines:
                for row, info in enumerate(medicines):
                    # Condition: Set the row color based on quantity
                    if info[4] == 0:  # 'quantity' is in the 5th column (index 4)
                        row_color = "#FF6347"  # Red for out of stock
                    elif info[4] < 5:
                        row_color = "#FF9966"  # Orange for low stock
                    else:
                        row_color = "#90EE90"  # Green for sufficient stock

                    data = [
                        info[0],  # medicine_id
                        info[1],  # supplier_id
                        info[2],  # name
                        info[3],  # company_name
                        info[4],  # quantity
                        info[5].strftime('%Y-%m-%d'),  # expiry date
                        info[6].strftime('%Y-%m-%d'),  # store date
                        info[7]   # rack
                    ]

                    # Display data with consistent column widths
                    for col, field in enumerate(data):
                        medicine_label = ctk.CTkLabel(
                            scroll_frame,
                            text=str(field),
                            font=("Arial", 12),
                            fg_color=row_color,
                            corner_radius=8,
                            anchor="center",
                            height=30
                        )
                        medicine_label.grid(
                            row=row,
                            column=col,
                            padx=2,
                            pady=2,
                            sticky="nsew"
                        )
            else:
                ctk.CTkLabel(
                    scroll_frame,
                    text="No medicines available.",
                    font=("Arial", 13, "bold"),
                    text_color="#E74C3C",
                    anchor="center",
                    height=40
                ).grid(row=0, column=0, columnspan=len(headers), pady=20, sticky="nsew")

        except Exception as e:
            error_message = f"An error occurred: {str(e)}"
            print(error_message)  # Print to console for debugging
            ctk.CTkLabel(
                scroll_frame,
                text=error_message,
                font=("Arial", 13, "bold"),
                text_color="#E74C3C",
                anchor="center",
                height=40,
                wraplength=800
            ).grid(row=0, column=0, columnspan=len(headers), pady=20, sticky="nsew")

    # Initial load of all medicines
    show_filtered_stock_report("All Medicines")

    # Function to handle Backspace key event
    def on_backspace(event):
        show_inventory_management()

    # Bind Backspace key to trigger the back button functionality
    backspace_binding = root.bind("<BackSpace>", on_backspace)

    # Function to unbind the Backspace key when leaving this function
    def cleanup():
        root.unbind("<BackSpace>", backspace_binding)

    # Schedule the cleanup function to run when leaving this function
    root.after(0, cleanup)

def show_medicine_expiry():
    clear_widgets(root)
    reset_grid_configuration(root)
    
    # Configure column weights for proper alignment
    for col in range(8):
        root.grid_columnconfigure(col, weight=1)

    # Title Section
    title_label = ctk.CTkLabel(
        root, 
        text="Medicine Expiry Report", 
        font=("Arial", 24, "bold"),
        text_color="#2E86C1"
    )
    title_label.grid(row=0, column=0, columnspan=8, pady=20, sticky="nsew")

    # Back button and search section
    back_icon_image = ctk.CTkImage(Image.open("back.png"), size=(30, 30))
    
    # Search frame setup
    search_frame = ctk.CTkFrame(root, fg_color="transparent")
    search_frame.grid(row=0, column=0, padx=20, pady=10, sticky="nw")
    
    # Back button and search entry
    back_button = ctk.CTkButton(
        search_frame,
        image=back_icon_image,
        text="",
        width=40,
        command=show_inventory_management,
        fg_color="transparent",
        hover_color="#E5E7E9"
    )
    back_button.grid(row=0, column=0, padx=(0, 5), pady=0, sticky="nw")

    search_var = tk.StringVar()
    search_entry = ctk.CTkEntry(
        search_frame,
        placeholder_text="🔍 Search Medicine",
        width=200,
        height=32,
        border_color="#2E86C1",
        corner_radius=8,
        textvariable=search_var
    )
    search_entry.grid(row=0, column=1, padx=(0, 5), pady=6, sticky="nw")
    search_entry.focus_set()

    # Function to update the report based on search term
    def update_report(*args):
        search_term = search_var.get().strip()
        show_filtered_expiry_report(filter_combobox.get(), search_term=search_term)

    # Bind the search_var to update_report function
    search_var.trace_add("write", update_report)

    # ComboBox for filtering options
    filter_options = ["All Medicines", "Expired Medicines", "Expiring Soon"]
    filter_combobox = ctk.CTkComboBox(
        root,
        values=filter_options,
        command=lambda choice: show_filtered_expiry_report(choice, search_term=search_var.get()),
        width=150,
        height=32,
        border_color="#2E86C1",
        button_color="#2E86C1",
        dropdown_hover_color="#3498DB"
    )
    filter_combobox.set("All Medicines")
    filter_combobox.grid(row=0, column=7, padx=(0, 20), pady=10, sticky="e")

    # Label to display the total number of medicines
    total_medicines_label = ctk.CTkLabel(
        root,
        text="Total Medicines: 0",  # Default text
        font=("Arial", 14, "bold"),
        text_color="#2E86C1"
    )
    total_medicines_label.grid(row=0, column=6, padx=(0, 10), pady=10, sticky="e")  # Placed beside the ComboBox

    # Header Section with fixed widths and proper alignment
    headers = [
        ("Medicine ID", 10),
        ("Supplier ID", 10),   # Supplier ID added after Medicine ID
        ("Name", 25),
        ("Company Name", 25),  # Company Name added after Name
        ("Expiry Date", 15),
        ("Status", 15),
        ("Quantity", 10),
        ("Rack No", 10)
    ]

    # Create a frame for headers to ensure proper alignment
    header_frame = ctk.CTkFrame(root, fg_color="transparent")
    header_frame.grid(row=1, column=0, columnspan=8, sticky="nsew", padx=10)
    
    # Configure header frame columns with proper weights
    for col in range(len(headers)):
        header_frame.grid_columnconfigure(col, weight=headers[col][1])

    # Create headers with proper widths
    for col, (header, weight) in enumerate(headers):
        header_label = ctk.CTkLabel(
            header_frame,
            text=header,
            font=("Arial", 14, "bold"),
            text_color="#FFFFFF",
            fg_color="#2E86C1",
            corner_radius=8,
            height=35,
            anchor="center"
        )
        header_label.grid(
            row=0,
            column=col,
            padx=2,
            pady=10,
            sticky="nsew"
        )

    # Scrollable frame for data
    scroll_frame = ctk.CTkScrollableFrame(
        root,
        height=600,
        width=900,
        fg_color="#F8F9F9",
        corner_radius=10
    )
    scroll_frame.grid(row=2, column=0, columnspan=8, padx=10, pady=10, sticky="nsew")

    # Configure scroll frame columns with the same weights as headers
    for col in range(len(headers)):
        scroll_frame.grid_columnconfigure(col, weight=headers[col][1])

    # Function to show filtered expiry report
    def show_filtered_expiry_report(filter_choice, search_term=""):
        # Clear previous data in scroll frame
        for widget in scroll_frame.winfo_children():
            widget.destroy()

        try:
            today = datetime.now().date()
            one_week_later = today + timedelta(weeks=1)

            query = """
                SELECT medicine_id, supplier_id, name, company_name, expiry, quantity, rack 
                FROM medicines WHERE 1=1
            """

            filter_params = ()

            if filter_choice == "Expired Medicines":
                query += " AND expiry < %s"
                filter_params = (today,)
            elif filter_choice == "Expiring Soon":
                query += " AND expiry <= %s AND expiry >= %s"
                filter_params = (one_week_later, today)
            
            # If search term is provided, add it to the query
            if search_term:
                query += " AND name LIKE %s"
                filter_params += (f"%{search_term}%",)

            query += " ORDER BY expiry ASC"
            expiry_info = fetch_data(query, filter_params)

            # Update the total medicines label
            total_medicines_label.configure(text=f"Total Medicines: {len(expiry_info)}")



            if expiry_info:
                for row, info in enumerate(expiry_info):
                    expiry_date = info[4].date() if isinstance(info[4], datetime) else info[4]

                    if expiry_date < today:
                        status = "Expired"
                        color = "#FF6347"
                    elif today <= expiry_date <= one_week_later:
                        status = "Expiring Soon"
                        color = "#FF9966"
                    else:
                        status = "Normal"
                        color = "#90EE90"

                    data = [
                        info[0],  # medicine_id
                        info[1],  # supplier_id
                        info[2],  # name
                        info[3],  # company_name
                        expiry_date.strftime('%Y-%m-%d'),
                        status,
                        info[5],  # quantity
                        info[6]   # rack
                    ]

                    # Display data with consistent column widths
                    for col, field in enumerate(data):
                        medicine_label = ctk.CTkLabel(
                            scroll_frame,
                            text=str(field),
                            font=("Arial", 12),
                            fg_color=color,
                            corner_radius=8,
                            anchor="center",
                            height=30
                        )
                        medicine_label.grid(
                            row=row,
                            column=col,
                            padx=2,
                            pady=2,
                            sticky="nsew"
                        )
            else:
                ctk.CTkLabel(
                    scroll_frame,
                    text="No medicines match the selected filter.",
                    font=("Arial", 13, "bold"),
                    text_color="#E74C3C",
                    anchor="center",
                    height=40
                ).grid(row=0, column=0, columnspan=len(headers), pady=20, sticky="nsew")

        except Exception as e:
            error_message = f"An error occurred: {str(e)}"
            print(error_message)  # Print to console for debugging
            ctk.CTkLabel(
                scroll_frame,
                text=error_message,
                font=("Arial", 13, "bold"),
                text_color="#E74C3C",
                anchor="center",
                height=40,
                wraplength=800
            ).grid(row=0, column=0, columnspan=len(headers), pady=20, sticky="nsew")

    # Initial load of all medicines
    show_filtered_expiry_report("All Medicines")

    # Function to handle Backspace key event
    def on_backspace(event):
        show_inventory_management()

    # Bind Backspace key to trigger the back button functionality
    backspace_binding = root.bind("<BackSpace>", on_backspace)

    # Function to unbind the Backspace key when leaving this function
    def cleanup():
        root.unbind("<BackSpace>", backspace_binding)

    # Schedule the cleanup function to run when leaving this function
    root.after(0, cleanup)

def show_reporting():
    clear_widgets(root)
    
    # Create main container with gradient background
    container = ctk.CTkFrame(
        root,
        fg_color=("#EEF2FF", "#1a1a2e")  # Light purple/blue gradient
    )
    container.pack(expand=True, fill="both", padx=40, pady=30)
    
    # Decorative header section
    header_frame = ctk.CTkFrame(container, fg_color="transparent")
    header_frame.pack(fill="x", pady=(20, 40))
    
    # Main title with modern styling
    title_label = ctk.CTkLabel(
        header_frame, 
        text="Reports",
        font=("Helvetica", 32, "bold"),
        text_color=("#1e40af", "#93c5fd")  # Dark blue for light mode, light blue for dark mode
    )
    title_label.pack()
    
    # Subtitle with softer color
    subtitle_label = ctk.CTkLabel(
        header_frame,
        text="Select an option to manage your reports",
        font=("Helvetica", 16),
        text_color=("#6b7280", "#9ca3af")  # Gray that works in both modes
    )
    subtitle_label.pack(pady=(5, 0))
    
    # Create a card-like frame for buttons
    card_frame = ctk.CTkFrame(
        container,
        fg_color=("#ffffff", "#1e1e2d"),  # White for light mode, dark blue for dark mode
        corner_radius=20
    )
    card_frame.pack(padx=40, pady=20, ipady=30, fill="x")
    
    # Button configurations
    button_width = 300
    button_height = 50
    button_padding = 15
    
    # Add Medicine Button with icon-like prefix
    add_button = ctk.CTkButton(
        card_frame,
        text="➕ Added Medicine",
        width=button_width,
        height=button_height,
        fg_color=("#3b82f6", "#2563eb"),  # Blue that works in both modes
        text_color="white",
        font=("Helvetica", 18),
        corner_radius=15,
        hover_color=("#2563eb", "#1d4ed8"),  # Slightly darker blue on hover
        command=show_added_medicines_report
    )
    add_button.pack(pady=button_padding)
    
    # Edit Medicine Button
    edit_button = ctk.CTkButton(
        card_frame,
        text="💵 Sold Medicine",
        width=button_width,
        height=button_height,
        fg_color=("#3b82f6", "#2563eb"),
        text_color="white",
        font=("Helvetica", 18),
        corner_radius=15,
        hover_color=("#2563eb", "#1d4ed8"),
        command=show_sold_medicines_report
    )
    edit_button.pack(pady=button_padding)
    
    # Shifted Medicine Button
    show_button = ctk.CTkButton(
        card_frame,
        text="🚛 Shifted Medicine",
        width=button_width,
        height=button_height,
        fg_color=("#3b82f6", "#2563eb"),
        text_color="white",
        font=("Helvetica", 18),
        corner_radius=15,
        hover_color=("#2563eb", "#1d4ed8"),
        command=show_shifted_medicines_report
    )
    show_button.pack(pady=button_padding)
    edit_button = ctk.CTkButton(
        card_frame,
        text="🪙 Profit Management",
        width=button_width,
        height=button_height,
        fg_color=("#3b82f6", "#2563eb"),
        text_color="white",
        font=("Helvetica", 18),
        corner_radius=15,
        hover_color=("#2563eb", "#1d4ed8"),
        command=show_profit_management
    )
    edit_button.pack(pady=button_padding)
    # Back Button with refined styling
    back_button = ctk.CTkButton(
        container,
        text="← Back to Main Menu",
        width=button_width,
        height=40,
        fg_color="transparent",
        text_color=("#6b7280", "#9ca3af"),
        font=("Helvetica", 16),
        border_width=2,
        border_color=("#d1d5db", "#4b5563"),
        corner_radius=15,
        hover_color=("#f3f4f6", "#1f2937"),
        command=show_main_menu
    )
    back_button.pack(pady=(40, 20))
def get_db_connection():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="123@123",
        database="pharmacy"
    )
def load_medicines():
    medicines = {}
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT * FROM medicines")
        for row in cursor.fetchall():
            medicines[str(row['medicine_id'])] = row
    except mysql.connector.Error as e:
        print(f"Error loading medicines: {e}")
    finally:
        cursor.close()
        conn.close()
    return medicines
def fetch_data(query, params=None):
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        if params:
            cursor.execute(query, params)
        else:
            cursor.execute(query)
        return cursor.fetchall()
    finally:
        cursor.close()
        conn.close()

from tkcalendar import DateEntry
import tkinter as tk
from datetime import datetime, timedelta
def fetch_filtered_added_medicines(period, page=1, items_per_page=15, search_term="", start_date=None, end_date=None):
    """Function to fetch filtered added medicines with pagination."""
    current_date = datetime.now().date()
    offset = (page - 1) * items_per_page

    # Base query for the snapshot_medicines table
    base_query = """
        SELECT medicine_id, name, store, store_time, version, quantity
        FROM snapshot_medicines
        WHERE store IS NOT NULL
    """

    # Count query for total records
    count_query = "SELECT COUNT(*) FROM snapshot_medicines WHERE store IS NOT NULL"

    # Adjust the filter based on the selected period
    if period == "Today":
        filter_date = current_date
        where_clause = " AND DATE(store) = %s"
        params = (filter_date,)
    elif period == "Last Week":
        filter_date = current_date - timedelta(days=7)
        where_clause = " AND DATE(store) >= %s"
        params = (filter_date,)
    elif period == "Last Month":
        filter_date = current_date - timedelta(days=30)
        where_clause = " AND DATE(store) >= %s"
        params = (filter_date,)
    elif period == "6 Months":
        filter_date = current_date - timedelta(days=180)
        where_clause = " AND DATE(store) >= %s"
        params = (filter_date,)
    elif period == "1 Year":
        filter_date = current_date - timedelta(days=365)
        where_clause = " AND DATE(store) >= %s"
        params = (filter_date,)
    elif period == "Other" and start_date and end_date:
        where_clause = " AND DATE(store) BETWEEN %s AND %s"
        params = (start_date, end_date)
    else:  # Lifetime
        where_clause = ""
        params = ()

    # If a search term is provided, add it to the query
    if search_term:
        where_clause += " AND name LIKE %s"
        params += (f"%{search_term}%",)

    # Add where clause to queries
    if where_clause:
        base_query += where_clause
        count_query += where_clause

    # Get total count
    total_records = fetch_data(count_query, params)[0][0]
    total_pages = ceil(total_records / items_per_page)

    # Add pagination to base query
    base_query += " ORDER BY store DESC LIMIT %s OFFSET %s"
    params += (items_per_page, offset)

    return fetch_data(base_query, params), total_records, total_pages

def show_added_medicines_report(page=1, selected_period="Lifetime", search_term="", selected_date=None):
    clear_widgets(root)
    reset_grid_configuration(root)
    
    ITEMS_PER_PAGE = 15  # Number of items per page
    
    # Configure column weights for proper alignment
    for col in range(6):
        root.grid_columnconfigure(col, weight=1)

    # Title Section
    title_label = ctk.CTkLabel(
        root, 
        text="Medicine Added Report", 
        font=("Arial", 24, "bold"),
        text_color="#2E86C1"
    )
    title_label.grid(row=0, column=0, columnspan=6, pady=20, sticky="nsew")

    # Back button and search section
    back_icon_image = ctk.CTkImage(Image.open("back.png"), size=(30, 30))
    
    # Search frame setup
    search_frame = ctk.CTkFrame(root, fg_color="transparent")
    search_frame.grid(row=0, column=0, padx=20, pady=10, sticky="nw")
    
    # Back button and search entry
    back_button = ctk.CTkButton(
        search_frame,
        image=back_icon_image,
        text="",
        width=40,
        command=show_reporting,
        fg_color="transparent",
        hover_color="#E5E7E9"
    )
    back_button.grid(row=0, column=0, padx=(0, 5), pady=0, sticky="nw")

    search_var = tk.StringVar()
    search_entry = ctk.CTkEntry(
        search_frame,
        placeholder_text="🔍 Search Medicine",
        width=200,
        height=32,
        border_color="#2E86C1",
        corner_radius=8,
        textvariable=search_var
    )
    search_entry.grid(row=0, column=1, padx=(0, 5), pady=6, sticky="nw")
    search_entry.insert(0, search_term)
    search_entry.focus_set()

    def update_report(*args):
        show_added_medicines_report(page=1, selected_period=selected_period, search_term=search_var.get(), selected_date=selected_date)

    search_var.trace_add("write", update_report)
    
    # Calendar popup function
    def show_calendar_popup():
        calendar_window = tk.Toplevel(root)
        calendar_window.title("Select Date Range")
        calendar_window.geometry("300x300")

        start_date_label = ctk.CTkLabel(calendar_window, text="From:", font=("Arial", 12, "bold"))
        start_date_label.pack(pady=(10, 0))

        start_cal = DateEntry(calendar_window, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        start_cal.pack(pady=(5, 10))

        end_date_label = ctk.CTkLabel(calendar_window, text="End Date:", font=("Arial", 12, "bold"))
        end_date_label.pack(pady=(10, 0))

        end_cal = DateEntry(calendar_window, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        end_cal.pack(pady=(5, 10))

        def on_select_dates():
            start_date = start_cal.get_date()
            end_date = end_cal.get_date()
            calendar_window.destroy()
            show_added_medicines_report(page=1, selected_period="Other", search_term=search_var.get(), selected_date=(start_date, end_date))

        select_button = ctk.CTkButton(calendar_window, text="Select Dates", command=on_select_dates)
        select_button.pack(pady=20)

    def toggle_calendar(choice):
        if choice == "Other":
            show_calendar_popup()
        else:
            show_added_medicines_report(page=1, selected_period=choice, search_term=search_var.get())

    # Period combobox
    period_options = ["Today", "Last Week", "Last Month", "6 Months", "1 Year", "Lifetime", "Other"]
    period_combobox = ctk.CTkComboBox(
        root,
        values=period_options,
        command=toggle_calendar,
        width=150,
        height=32,
        border_color="#2E86C1",
        button_color="#2E86C1",
        dropdown_hover_color="#3498DB"
    )
    period_combobox.set(selected_period)
    period_combobox.grid(row=0, column=5, padx=20, pady=10, sticky="ne")

    # Header Section
    headers_added = [
        ("Medicine ID", 10, "center"),
        ("Name", 25, "center"),
        ("Version", 15, "center"),
        ("Quantity", 20, "center"),
        ("Store Date", 15, "center"),
        ("Store Time", 15, "center")
    ]

    # Create header frame
    header_frame = ctk.CTkFrame(root, fg_color="transparent")
    header_frame.grid(row=1, column=0, columnspan=6, sticky="nsew", padx=10)
    
    # Configure header columns
    total_weight = sum(weight for _, weight, _ in headers_added)
    for col, (_, weight, _) in enumerate(headers_added):
        header_frame.grid_columnconfigure(col, weight=weight)

    # Create headers
    for col, (header, _, _) in enumerate(headers_added):
        header_label = ctk.CTkLabel(
            header_frame,
            text=header,
            font=("Arial", 14, "bold"),
            text_color="#FFFFFF",
            fg_color="#2E86C1",
            corner_radius=8,
            height=35,
            anchor="center"
        )
        header_label.grid(row=0, column=col, padx=2, pady=10, sticky="nsew")

    # Data frame
    data_frame = ctk.CTkFrame(
        root,
        fg_color="#F8F9F9",
        corner_radius=10
    )
    data_frame.grid(row=2, column=0, columnspan=6, padx=10, pady=10, sticky="nsew")

    # Configure data frame columns
    for col, (_, weight, _) in enumerate(headers_added):
        data_frame.grid_columnconfigure(col, weight=weight)

    try:
        if selected_period == "Other" and selected_date:
            start_date, end_date = selected_date
            added_medicines, total_records, total_pages = fetch_filtered_added_medicines(
                "Other", page, ITEMS_PER_PAGE, search_term, start_date, end_date
            )
        else:
            added_medicines, total_records, total_pages = fetch_filtered_added_medicines(
                selected_period, page, ITEMS_PER_PAGE, search_term
            )

        # Update total medicines label
        total_medicines_label = ctk.CTkLabel(
            root,
            text=f"Total Medicines: {total_records}",
            font=("Arial", 14, "bold"),
            text_color="#2E86C1"
        )
        total_medicines_label.grid(row=0, column=4, padx=(0, 10), pady=10, sticky="ne")

        if isinstance(added_medicines, list):
            if added_medicines:
                for i, medicine in enumerate(added_medicines):
                    if isinstance(medicine, tuple) and len(medicine) == 6:
                        medicine_id, name, store_date, store_time, version, quantity = medicine

                        # Format store time
                        if store_time:
                            try:
                                store_time = datetime.strptime(str(store_time), '%H:%M:%S').strftime('%I:%M:%S %p')
                            except ValueError:
                                store_time = "Invalid Time"
                        else:
                            store_time = "N/A"

                        # Handle version display
                        version = str(version) if version is not None else "None"

                        fields = [
                            (str(medicine_id), "center"),
                            (name, "center"),
                            (version, "center"),
                            (str(quantity), "center"),
                            (str(store_date), "center"),
                            (store_time, "center")
                        ]
                        
                        for col, (field, align) in enumerate(fields):
                            medicine_label = ctk.CTkLabel(
                                data_frame,
                                text=field,
                                font=("Arial", 13),
                                text_color="#34495E",
                                fg_color="#FFFFFF" if i % 2 == 0 else "#F7F9F9",
                                corner_radius=8,
                                anchor=align,
                                height=30
                            )
                            padx = (10, 5) if align == "w" else (5, 10) if align == "e" else (5, 5)
                            medicine_label.grid(
                                row=i,
                                column=col,
                                padx=padx,
                                pady=2,
                                sticky="nsew"
                            )
            else:
                ctk.CTkLabel(
                    data_frame,
                    text="No medicines have been added.",
                    font=("Arial", 13, "bold"),
                    text_color="#E74C3C",
                    anchor="center",
                    height=40
                ).grid(row=0, column=0, columnspan=len(headers_added), pady=20, sticky="nsew")

            # Pagination frame
            pagination_frame = ctk.CTkFrame(root, fg_color="transparent")
            pagination_frame.grid(row=3, column=0, columnspan=6, pady=10, sticky="nsew")

            def change_page(new_page):
                show_added_medicines_report(
                    page=new_page,
                    selected_period=selected_period,
                    search_term=search_var.get(),
                    selected_date=selected_date
                )

            # Previous page button
            prev_button = ctk.CTkButton(
                pagination_frame,
                text="←",
                width=40,
                command=lambda: change_page(page - 1) if page > 1 else None,
                fg_color="#2E86C1" if page > 1 else "#B3B6B7",
                state="normal" if page > 1 else "disabled"
            )
            prev_button.grid(row=0, column=0, padx=5)

            # Page numbers
            page_label = ctk.CTkLabel(
                pagination_frame,
                text=f"Page {page} of {total_pages}",
                font=("Arial", 12, "bold")
            )
            page_label.grid(row=0, column=1, padx=10)

            # Next page button
            next_button = ctk.CTkButton(
                pagination_frame,
                text="→",
                width=40,
                command=lambda: change_page(page + 1) if page < total_pages else None,
                fg_color="#2E86C1" if page < total_pages else "#B3B6B7",
                state="normal" if page < total_pages else "disabled"
            )
            next_button.grid(row=0, column=2, padx=5)

        else:
            raise TypeError("Expected a list of tuples from fetch_filtered_added_medicines")

    except Exception as e:
        error_label = ctk.CTkLabel(
            data_frame,
            text=f"Error fetching added medicines data: {e}",
            font=("Arial", 13, "bold"),
            text_color="#E74C3C",
            anchor="center",
            height=40
        )
        error_label.grid(row=0, column=0, columnspan=len(headers_added), pady=20, sticky="nsew")
        
        # Empty pagination frame in case of error
        pagination_frame = ctk.CTkFrame(root, fg_color="transparent")
        pagination_frame.grid(row=3, column=0, columnspan=6, pady=10, sticky="nsew")

    status_frame = ctk.CTkFrame(root, fg_color="transparent")
    status_frame.grid(row=4, column=0, columnspan=6, pady=(5, 10), sticky="nsew")

    records_label = ctk.CTkLabel(
        status_frame,
        text=f"Showing {ITEMS_PER_PAGE} records per page",
        font=("Arial", 12),
        text_color="#7F8C8D"
    )
    records_label.pack(side="right", padx=20)

    # Add export button
    def export_to_excel():
        try:
            # Check if required modules are installed
            try:
                import pandas as pd
                import openpyxl
            except ImportError as e:
                if 'openpyxl' in str(e):
                    if messagebox.askyesno("Missing Dependency", 
                        "The openpyxl module is required for Excel export. Would you like to install it now?"):
                        import subprocess
                        try:
                            subprocess.check_call(['pip', 'install', 'openpyxl'])
                            messagebox.showinfo("Success", "openpyxl installed successfully. Please try exporting again.")
                        except subprocess.CalledProcessError:
                            messagebox.showerror("Installation Failed", 
                                "Could not install openpyxl. Please install it manually using 'pip install openpyxl'")
                    return
                else:
                    messagebox.showerror("Missing Dependency", f"Required module not found: {str(e)}")
                    return

            # Fetch all records for export
            all_records, _, _ = fetch_filtered_added_medicines(
                selected_period,
                page=1,
                items_per_page=1000000,  # Large number to get all records
                search_term=search_term,
                start_date=selected_date[0] if selected_date else None,
                end_date=selected_date[1] if selected_date else None
            )
            
            if not all_records:
                messagebox.showwarning("Export Failed", "No records to export.")
                return
            def format_time(store_time):
                if not store_time:
                    return "N/A"
                
                try:
                    if isinstance(store_time, str):
                        # Try parsing as string
                        parsed_time = datetime.strptime(store_time, '%H:%M:%S').time()
                    elif isinstance(store_time, datetime):
                        # If it's already a datetime, just get the time
                        parsed_time = store_time.time()
                    elif isinstance(store_time, time):
                        # If it's already a time object, use it directly
                        parsed_time = store_time
                    else:
                        # If it's another type, try converting to string first
                        parsed_time = datetime.strptime(str(store_time), '%H:%M:%S').time()
                    
                    return parsed_time.strftime('%I:%M:%S %p')
                except ValueError as e:
                    print(f"Error parsing time {store_time}: {e}")
                    return str(store_time)
                        # Convert records to list of dicts for better handling
            formatted_records = []
            for record in all_records:
                medicine_id, name, store_date, store_time, version, quantity = record
                
                formatted_records.append({
                    'Medicine ID': medicine_id,
                    'Name': name,
                    'Version': version if version is not None else "None",
                    'Quantity': quantity,
                    'Store Date': store_date,
                    'Store Time': format_time(store_time)
                })

            # Create DataFrame from formatted records
            df = pd.DataFrame(formatted_records)

            # Create DataFrame from formatted records
            df = pd.DataFrame(formatted_records)

            # Ask user for save location and format
            file_types = [
                ("Excel files", "*.xlsx"),
                ("CSV files", "*.csv"),
                ("Text files", "*.txt")
            ]
            file_path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=file_types,
                title="Export Added Medicines Report"
            )

            if file_path:
                if file_path.endswith('.xlsx'):
                    # Excel export with formatting
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='Added Medicines Report')
                        workbook = writer.book
                        worksheet = writer.sheets['Added Medicines Report']
                        
                        # Auto-adjust column widths
                        for column in worksheet.columns:
                            max_length = 0
                            column = [cell for cell in column]
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(cell.value)
                                except:
                                    pass
                            adjusted_width = (max_length + 2)
                            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

                        # Add some basic formatting
                        header = openpyxl.styles.PatternFill(
                            start_color='1E88E5',
                            end_color='1E88E5',
                            fill_type='solid'
                        )
                        for cell in worksheet[1]:
                            cell.fill = header
                            cell.font = openpyxl.styles.Font(color='FFFFFF', bold=True)

                elif file_path.endswith('.csv'):
                    df.to_csv(file_path, index=False)
                elif file_path.endswith('.txt'):
                    df.to_csv(file_path, index=False, sep='\t')

                messagebox.showinfo("Export Successful", 
                    f"Report exported successfully to:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Export Failed", 
                f"Error exporting data: {str(e)}\n\nPlease make sure you have permission to write to the selected location.")

    # Update the export button
    export_button = ctk.CTkButton(
        status_frame,
        text="Export Report",
        command=export_to_excel,
        width=120,
        height=30,
        fg_color="#27AE60",
        hover_color="#219A52"
    )
    export_button.pack(side="left", padx=20)

    # Add keyboard shortcuts for navigation
    def handle_keypress(event):
        if event.keysym == "Left" and page > 1:
            show_added_medicines_report(
                page=page-1,
                selected_period=selected_period,
                search_term=search_var.get(),
                selected_date=selected_date
            )
        elif event.keysym == "Right" and page < total_pages:
            show_added_medicines_report(
                page=page+1,
                selected_period=selected_period,
                search_term=search_var.get(),
                selected_date=selected_date
            )

    root.bind("<Left>", handle_keypress)
    root.bind("<Right>", handle_keypress)

    # Keep search entry focused
    search_entry.focus_set()
def fetch_filtered_sold_medicines(period, page=1, items_per_page=15, search_term="", start_date=None, end_date=None):
    """Function to fetch filtered sold medicines with pagination."""
    current_date = datetime.now().date()
    offset = (page - 1) * items_per_page

    # Base query for the sales table
    base_query = """
        SELECT sale_id, medicine_id, medicine_name, quantity_sold, 
               selling_price, discount, total_price, amount_paid, change_given,
               DATE(sale_date) as sale_date, TIME(sale_date) as sale_time
        FROM sales
    """

    # Count query for total records
    count_query = "SELECT COUNT(*) FROM sales"

    # Adjust the filter based on the selected period
    if period == "Today":
        filter_date = current_date
        where_clause = " WHERE DATE(sale_date) = %s"
        params = (filter_date,)
    elif period == "Last Week":
        filter_date = current_date - timedelta(days=7)
        where_clause = " WHERE DATE(sale_date) >= %s"
        params = (filter_date,)
    elif period == "Last Month":
        filter_date = current_date - timedelta(days=30)
        where_clause = " WHERE DATE(sale_date) >= %s"
        params = (filter_date,)
    elif period == "6 Months":
        filter_date = current_date - timedelta(days=180)
        where_clause = " WHERE DATE(sale_date) >= %s"
        params = (filter_date,)
    elif period == "1 Year":
        filter_date = current_date - timedelta(days=365)
        where_clause = " WHERE DATE(sale_date) >= %s"
        params = (filter_date,)
    elif period == "Other" and start_date and end_date:
        where_clause = " WHERE DATE(sale_date) BETWEEN %s AND %s"
        params = (start_date, end_date)
    else:  # Lifetime
        where_clause = ""
        params = ()

    # If a search term is provided, add it to the query
    if search_term:
        if where_clause:
            where_clause += " AND medicine_name LIKE %s"
        else:
            where_clause = " WHERE medicine_name LIKE %s"
        params += (f"%{search_term}%",)

    # Add where clause to queries
    if where_clause:
        base_query += where_clause
        count_query += where_clause

    # Get total count
    total_records = fetch_data(count_query, params)[0][0]
    total_pages = ceil(total_records / items_per_page)

    # Add pagination to base query
    base_query += " LIMIT %s OFFSET %s"
    params += (items_per_page, offset)

    return fetch_data(base_query, params), total_records, total_pages

def show_sold_medicines_report(page=1, selected_period="Lifetime", search_term="", selected_date=None):
    clear_widgets(root)
    reset_grid_configuration(root)
    
    ITEMS_PER_PAGE = 15  # Number of items per page
    
    # Configure column weights for proper alignment
    for col in range(11):
        root.grid_columnconfigure(col, weight=1)

    # Title Section
    title_label = ctk.CTkLabel(
        root, 
        text="Medicine Sales Report", 
        font=("Arial", 24, "bold"),
        text_color="#2E86C1"
    )
    title_label.grid(row=0, column=0, columnspan=12, pady=20, sticky="nsew")

    # Back button and search section
    back_icon_image = ctk.CTkImage(Image.open("back.png"), size=(30, 30))
    
    # Search frame setup
    search_frame = ctk.CTkFrame(root, fg_color="transparent")
    search_frame.grid(row=0, column=0, padx=20, pady=10, sticky="nw")
    
    # Back button and search entry
    back_button = ctk.CTkButton(
        search_frame,
        image=back_icon_image,
        text="",
        width=40,
        command=show_reporting,
        fg_color="transparent",
        hover_color="#E5E7E9"
    )
    back_button.grid(row=0, column=0, padx=(0, 5), pady=0, sticky="nw")

    search_var = tk.StringVar()
    search_entry = ctk.CTkEntry(
        search_frame,
        placeholder_text="🔍 Search Medicine",
        width=200,
        height=32,
        border_color="#2E86C1",
        corner_radius=8,
        textvariable=search_var
    )
    search_entry.grid(row=0, column=1, padx=(0, 5), pady=6, sticky="nw")
    search_entry.insert(0, search_term)
    search_entry.focus_set()

    def update_report(*args):
        show_sold_medicines_report(page=1, selected_period=selected_period, search_term=search_var.get(), selected_date=selected_date)

    search_var.trace_add("write", update_report)
    
    # Calendar popup function remains the same
    def show_calendar_popup():
        calendar_window = tk.Toplevel(root)
        calendar_window.title("Select Date Range")
        calendar_window.geometry("300x300")

        start_date_label = ctk.CTkLabel(calendar_window, text="From:", font=("Arial", 12, "bold"))
        start_date_label.pack(pady=(10, 0))

        start_cal = DateEntry(calendar_window, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        start_cal.pack(pady=(5, 10))

        end_date_label = ctk.CTkLabel(calendar_window, text="End Date:", font=("Arial", 12, "bold"))
        end_date_label.pack(pady=(10, 0))

        end_cal = DateEntry(calendar_window, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        end_cal.pack(pady=(5, 10))

        def on_select_dates():
            start_date = start_cal.get_date()
            end_date = end_cal.get_date()
            calendar_window.destroy()
            show_sold_medicines_report(page=1, selected_period="Other", search_term=search_var.get(), selected_date=(start_date, end_date))

        select_button = ctk.CTkButton(calendar_window, text="Select Dates", command=on_select_dates)
        select_button.pack(pady=20)

    def toggle_calendar(choice):
        if choice == "Other":
            show_calendar_popup()
        else:
            show_sold_medicines_report(page=1, selected_period=choice, search_term=search_var.get())

    # Period combobox
    period_options = ["Today", "Last Week", "Last Month", "6 Months", "1 Year", "Lifetime", "Other"]
    period_combobox = ctk.CTkComboBox(
        root,
        values=period_options,
        command=toggle_calendar,
        width=150,
        height=32,
        border_color="#2E86C1",
        button_color="#2E86C1",
        dropdown_hover_color="#3498DB"
    )
    period_combobox.set(selected_period)
    period_combobox.grid(row=0, column=11, padx=20, pady=10, sticky="ne")

    # Header Section
    headers_sold = [
        ("Sales | Med ID", 15, "center"),
        ("Name", 10, "center"),
        ("Quantity", 8, "center"),
        ("Price", 10, "center"),
        ("Discount", 8, "center"),
        ("Total", 10, "center"),
        ("Paid", 10, "center"),
        ("Change", 10, "center"),
        ("Date", 8, "center"),
        ("Time", 8, "center")
    ]

    # Create header frame
    header_frame = ctk.CTkFrame(root, fg_color="transparent")
    header_frame.grid(row=1, column=0, columnspan=12, sticky="nsew", padx=10)
    
    # Configure header columns
    total_weight = sum(weight for _, weight, _ in headers_sold)
    for col, (_, weight, _) in enumerate(headers_sold):
        header_frame.grid_columnconfigure(col, weight=weight)

    # Create headers
    for col, (header, _, _) in enumerate(headers_sold):
        header_label = ctk.CTkLabel(
            header_frame,
            text=header,
            font=("Arial", 14, "bold"),
            text_color="#FFFFFF",
            fg_color="#2E86C1",
            corner_radius=8,
            height=35,
            anchor="center"
        )
        header_label.grid(row=0, column=col, padx=2, pady=10, sticky="nsew")

    # Data frame
    data_frame = ctk.CTkFrame(
        root,
        fg_color="#F8F9F9",
        corner_radius=10
    )
    data_frame.grid(row=2, column=0, columnspan=12, padx=10, pady=10, sticky="nsew")

    # Configure data frame columns
    for col, (_, weight, _) in enumerate(headers_sold):
        data_frame.grid_columnconfigure(col, weight=weight)

    try:
        if selected_period == "Other" and selected_date:
            start_date, end_date = selected_date
            sold_medicines, total_records, total_pages = fetch_filtered_sold_medicines(
                "Other", page, ITEMS_PER_PAGE, search_term, start_date, end_date
            )
        else:
            sold_medicines, total_records, total_pages = fetch_filtered_sold_medicines(
                selected_period, page, ITEMS_PER_PAGE, search_term
            )

        # Update total sales label
        total_sold_medicines_label = ctk.CTkLabel(
            root,
            text=f"Total Sales: {total_records}",
            font=("Arial", 14, "bold"),
            text_color="#2E86C1"
        )
        total_sold_medicines_label.grid(row=0, column=10, padx=(0, 10), pady=10, sticky="ne")

        if isinstance(sold_medicines, list):
            if sold_medicines:
                for i, sale in enumerate(sold_medicines):
                    if isinstance(sale, tuple) and len(sale) == 11:
                        sale_id, medicine_id, name, quantity, price, discount, total, paid, change, sale_date, sale_time = sale

                        # Format sale time
                        if sale_time:
                            try:
                                sale_time = datetime.strptime(str(sale_time), '%H:%M:%S').strftime('%I:%M:%S %p')
                            except ValueError:
                                sale_time = "Invalid Time"
                        else:
                            sale_time = "N/A"

                        # Calculate change
                        change = float(paid) - float(total)

                        # Format monetary values
                        price_fmt = f"Rs{float(price):>8,.2f}"
                        discount_fmt = f"Rs{float(discount):>8,.2f}"
                        total_fmt = f"Rs{float(total):>8,.2f}"
                        paid_fmt = f"Rs{float(paid):>8,.2f}"
                        change_fmt = f"Rs{change:>8,.2f}"

                        # Alternating row colors
                        bg_color = "#F4F6F7" if i % 2 == 0 else "#FFFFFF"
                        
                        fields = [
                            (f"{sale_id} | {medicine_id}", "center"),
                            (name, "center"),
                            (str(quantity), "center"),
                            (price_fmt, "center"),
                            (discount_fmt, "center"),
                            (total_fmt, "center"),
                            (paid_fmt, "center"),
                            (change_fmt, "center"),
                            (str(sale_date), "center"),
                            (sale_time, "center")
                        ]
                        
                        for col, (field, align) in enumerate(fields):
                            medicine_label = ctk.CTkLabel(
                                data_frame,
                                text=field,
                                font=("Arial", 13),
                                text_color="#34495E",
                                fg_color="#FFFFFF" if i % 2 == 0 else "#F7F9F9",
                                corner_radius=8,
                                anchor=align,
                                height=30
                            )
                            padx = (10, 5) if align == "w" else (5, 10) if align == "e" else (5, 5)
                            medicine_label.grid(
                                row=i,
                                column=col,
                                padx=padx,
                                pady=2,
                                sticky="nsew"
                            )
            else:
                ctk.CTkLabel(
                    data_frame,
                    text="No sales records found.",
                    font=("Arial", 13, "bold"),
                    text_color="#E74C3C",
                    anchor="center",
                    height=40
                ).grid(row=0, column=0, columnspan=len(headers_sold), pady=20, sticky="nsew")

            # Pagination frame
            pagination_frame = ctk.CTkFrame(root, fg_color="transparent")
            pagination_frame.grid(row=3, column=0, columnspan=12, pady=10, sticky="nsew")

            def change_page(new_page):
                show_sold_medicines_report(
                    page=new_page,
                    selected_period=selected_period,
                    search_term=search_var.get(),
                    selected_date=selected_date
                )

            # Previous page button
            prev_button = ctk.CTkButton(
                pagination_frame,
                text="←",
                width=40,
                command=lambda: change_page(page - 1) if page > 1 else None,
                fg_color="#2E86C1" if page > 1 else "#B3B6B7",
                state="normal" if page > 1 else "disabled"
            )
            prev_button.grid(row=0, column=0, padx=5)

            # Page numbers
            page_label = ctk.CTkLabel(
                pagination_frame,
                text=f"Page {page} of {total_pages}",
                font=("Arial", 12, "bold")
            )
            page_label.grid(row=0, column=1, padx=10)

            # Next page button
            next_button = ctk.CTkButton(
                pagination_frame,
                text="→",
                width=40,
                command=lambda: change_page(page + 1) if page < total_pages else None,
                fg_color="#2E86C1" if page < total_pages else "#B3B6B7",
                state="normal" if page < total_pages else "disabled"
            )
            next_button.grid(row=0, column=2, padx=5)

        else:
            raise TypeError("Expected a list of tuples from fetch_filtered_sold_medicines")

    except Exception as e:
        error_label = ctk.CTkLabel(
            data_frame,
            text=f"Error fetching sales data: {e}",
            font=("Arial", 13, "bold"),
            text_color="#E74C3C",
            anchor="center",
            height=40
        )
        error_label.grid(row=0, column=0, columnspan=len(headers_sold), pady=20, sticky="nsew")
        
        # Empty pagination frame in case of error
        pagination_frame = ctk.CTkFrame(root, fg_color="transparent")
        pagination_frame.grid(row=3, column=0, columnspan=12, pady=10, sticky="nsew")

    # Add a status bar showing records per page
    status_frame = ctk.CTkFrame(root, fg_color="transparent")
    status_frame.grid(row=4, column=0, columnspan=12, pady=(5, 10), sticky="nsew")

    records_label = ctk.CTkLabel(
        status_frame,
        text=f"Showing {ITEMS_PER_PAGE} records per page",
        font=("Arial", 12),
        text_color="#7F8C8D"
    )
    records_label.pack(side="right", padx=20)

    # Add export button (optional feature)
    def export_to_excel():
        try:
            # First check if required modules are installed
            try:
                import pandas as pd
                import openpyxl
            except ImportError as e:
                if 'openpyxl' in str(e):
                    if messagebox.askyesno("Missing Dependency", 
                        "The openpyxl module is required for Excel export. Would you like to install it now?"):
                        import subprocess
                        try:
                            subprocess.check_call(['pip', 'install', 'openpyxl'])
                            messagebox.showinfo("Success", "openpyxl installed successfully. Please try exporting again.")
                        except subprocess.CalledProcessError:
                            messagebox.showerror("Installation Failed", 
                                "Could not install openpyxl. Please install it manually using 'pip install openpyxl'")
                    return
                else:
                    messagebox.showerror("Missing Dependency", f"Required module not found: {str(e)}")
                    return

            # Fetch all records for export
            all_records, _, _ = fetch_filtered_sold_medicines(
                selected_period,
                page=1,
                items_per_page=1000000,  # Large number to get all records
                search_term=search_term,
                start_date=selected_date[0] if selected_date else None,
                end_date=selected_date[1] if selected_date else None
            )
            
            if not all_records:
                messagebox.showwarning("Export Failed", "No records to export.")
                return

            # Convert records to list of dicts for better handling
            formatted_records = []
            for record in all_records:
                sale_id, medicine_id, name, quantity, price, discount, total, paid, change, sale_date, sale_time = record
                
                # Debug: Print the type and value of sale_time
                print(f"Debug - sale_time: type={type(sale_time)}, value={sale_time}")

                # Improved time formatting
                formatted_time = "N/A"
                if sale_time:
                    try:
                        if isinstance(sale_time, str):
                            # If it's a string, try to parse it
                            formatted_time = datetime.strptime(sale_time, '%H:%M:%S').strftime('%I:%M:%S %p')
                        elif isinstance(sale_time, datetime):
                            # If it's already a datetime object, just format it
                            formatted_time = sale_time.strftime('%I:%M:%S %p')
                        elif isinstance(sale_time, time):
                            # If it's a time object, convert to string
                            formatted_time = sale_time.strftime('%I:%M:%S %p')
                        else:
                            # If it's none of the above, convert to string
                            formatted_time = str(sale_time)
                    except ValueError as e:
                        print(f"Debug - Error formatting time: {e}")
                        # If parsing fails, keep the original string
                        formatted_time = str(sale_time)

                # Debug: Print the formatted time
                print(f"Debug - formatted_time: {formatted_time}")

                formatted_records.append({
                    'Sale ID': sale_id,
                    'Medicine ID': medicine_id,
                    'Medicine Name': name,
                    'Quantity': quantity,
                    'Price': f"Rs{float(price):,.2f}",
                    'Discount': f"Rs{float(discount):,.2f}",
                    'Total': f"Rs{float(total):,.2f}",
                    'Paid': f"Rs{float(paid):,.2f}",
                    'Change': f"Rs{float(change):,.2f}",
                    'Date': sale_date,
                    'Time': formatted_time
                })

            # Create DataFrame from formatted records
            df = pd.DataFrame(formatted_records)

            # Ask user for save location and format
            file_types = [
                ("Excel files", "*.xlsx"),
                ("CSV files", "*.csv"),
                ("Text files", "*.txt")
            ]
            file_path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=file_types,
                title="Export Sales Report"
            )

            if file_path:
                if file_path.endswith('.xlsx'):
                    # Excel export with formatting
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='Sales Report')
                        workbook = writer.book
                        worksheet = writer.sheets['Sales Report']
                        
                        # Auto-adjust column widths
                        for column in worksheet.columns:
                            max_length = 0
                            column = [cell for cell in column]
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(cell.value)
                                except:
                                    pass
                            adjusted_width = (max_length + 2)
                            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

                        # Add some basic formatting
                        header = openpyxl.styles.PatternFill(
                            start_color='1E88E5',
                            end_color='1E88E5',
                            fill_type='solid'
                        )
                        for cell in worksheet[1]:
                            cell.fill = header
                            cell.font = openpyxl.styles.Font(color='FFFFFF', bold=True)

                elif file_path.endswith('.csv'):
                    df.to_csv(file_path, index=False)
                elif file_path.endswith('.txt'):
                    df.to_csv(file_path, index=False, sep='\t')

                messagebox.showinfo("Export Successful", 
                    f"Report exported successfully to:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Export Failed", 
                f"Error exporting data: {str(e)}\n\nPlease make sure you have permission to write to the selected location.")

    # Update the export button
    export_button = ctk.CTkButton(
        status_frame,
        text="Export Report",
        command=export_to_excel,
        width=120,
        height=30,
        fg_color="#27AE60",
        hover_color="#219A52"
    )
    export_button.pack(side="left", padx=20)

    # Add keyboard shortcuts for navigation
    def handle_keypress(event):
        if event.keysym == "Left" and page > 1:
            show_sold_medicines_report(
                page=page-1,
                selected_period=selected_period,
                search_term=search_var.get(),
                selected_date=selected_date
            )
        elif event.keysym == "Right" and page < total_pages:
            show_sold_medicines_report(
                page=page+1,
                selected_period=selected_period,
                search_term=search_var.get(),
                selected_date=selected_date
            )

    root.bind("<Left>", handle_keypress)
    root.bind("<Right>", handle_keypress)

    # Keep search entry focused
    search_entry.focus_set()
def fetch_filtered_shifted_medicines(period, page=1, items_per_page=15, search_term="", start_date=None, end_date=None):
    """Function to fetch filtered shifted medicines with pagination."""
    current_date = datetime.now().date()
    offset = (page - 1) * items_per_page

    # Base query for the shifted_medicines table
    base_query = """
        SELECT sm.shift_id, sm.medicine_id, sm.medicine_name, sm.quantity_shifted, 
               DATE(sm.shifted_at) as shift_date, TIME(sm.shifted_at) as shift_time, 
               sm.company_name 
        FROM shifted_medicines sm
    """

    # Count query for total records
    count_query = "SELECT COUNT(*) FROM shifted_medicines sm"

    # Adjust the filter based on the selected period
    if period == "Today":
        filter_date = current_date
        where_clause = " WHERE DATE(sm.shifted_at) = %s"
        params = (filter_date,)
    elif period == "Last Week":
        filter_date = current_date - timedelta(days=7)
        where_clause = " WHERE DATE(sm.shifted_at) >= %s"
        params = (filter_date,)
    elif period == "Last Month":
        filter_date = current_date - timedelta(days=30)
        where_clause = " WHERE DATE(sm.shifted_at) >= %s"
        params = (filter_date,)
    elif period == "6 Months":
        filter_date = current_date - timedelta(days=180)
        where_clause = " WHERE DATE(sm.shifted_at) >= %s"
        params = (filter_date,)
    elif period == "1 Year":
        filter_date = current_date - timedelta(days=365)
        where_clause = " WHERE DATE(sm.shifted_at) >= %s"
        params = (filter_date,)
    elif period == "Other" and start_date and end_date:
        where_clause = " WHERE DATE(sm.shifted_at) BETWEEN %s AND %s"
        params = (start_date, end_date)
    else:  # Lifetime
        where_clause = ""
        params = ()

    # If a search term is provided, add it to the query
    if search_term:
        if where_clause:
            where_clause += " AND (sm.medicine_name LIKE %s OR sm.company_name LIKE %s)"
        else:
            where_clause = " WHERE (sm.medicine_name LIKE %s OR sm.company_name LIKE %s)"
        params += (f"%{search_term}%", f"%{search_term}%")

    # Add where clause to queries
    if where_clause:
        base_query += where_clause
        count_query += where_clause

    # Get total count
    total_records = fetch_data(count_query, params)[0][0]
    total_pages = ceil(total_records / items_per_page)

    # Add pagination to base query
    base_query += " ORDER BY sm.shifted_at DESC LIMIT %s OFFSET %s"
    params += (items_per_page, offset)

    return fetch_data(base_query, params), total_records, total_pages

def show_shifted_medicines_report(page=1, selected_period="Lifetime", search_term="", selected_date=None):
    clear_widgets(root)
    reset_grid_configuration(root)
    
    ITEMS_PER_PAGE = 15  # Number of items per page
    
    # Configure column weights for proper alignment
    for col in range(7):
        root.grid_columnconfigure(col, weight=1)

    # Title Section
    title_label = ctk.CTkLabel(
        root, 
        text="Medicine Shifted Report", 
        font=("Arial", 24, "bold"),
        text_color="#2E86C1"
    )
    title_label.grid(row=0, column=0, columnspan=7, pady=20, sticky="nsew")

    # Back button and search section
    back_icon_image = ctk.CTkImage(Image.open("back.png"), size=(30, 30))
    
    # Search frame setup
    search_frame = ctk.CTkFrame(root, fg_color="transparent")
    search_frame.grid(row=0, column=0, padx=20, pady=10, sticky="nw")
    
    # Back button and search entry
    back_button = ctk.CTkButton(
        search_frame,
        image=back_icon_image,
        text="",
        width=40,
        command=show_reporting,
        fg_color="transparent",
        hover_color="#E5E7E9"
    )
    back_button.grid(row=0, column=0, padx=(0, 5), pady=0, sticky="nw")

    search_var = tk.StringVar()
    search_entry = ctk.CTkEntry(
        search_frame,
        placeholder_text="🔍 Search Medicine or Company",
        width=200,
        height=32,
        border_color="#2E86C1",
        corner_radius=8,
        textvariable=search_var
    )
    search_entry.grid(row=0, column=1, padx=(0, 5), pady=6, sticky="nw")
    search_entry.insert(0, search_term)
    search_entry.focus_set()

    def update_report(*args):
        show_shifted_medicines_report(page=1, selected_period=selected_period, search_term=search_var.get(), selected_date=selected_date)

    search_var.trace_add("write", update_report)
    
    # Calendar popup function
    def show_calendar_popup():
        calendar_window = tk.Toplevel(root)
        calendar_window.title("Select Date Range")
        calendar_window.geometry("300x300")

        start_date_label = ctk.CTkLabel(calendar_window, text="From:", font=("Arial", 12, "bold"))
        start_date_label.pack(pady=(10, 0))

        start_cal = DateEntry(calendar_window, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        start_cal.pack(pady=(5, 10))

        end_date_label = ctk.CTkLabel(calendar_window, text="End Date:", font=("Arial", 12, "bold"))
        end_date_label.pack(pady=(10, 0))

        end_cal = DateEntry(calendar_window, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        end_cal.pack(pady=(5, 10))

        def on_select_dates():
            start_date = start_cal.get_date()
            end_date = end_cal.get_date()
            calendar_window.destroy()
            show_shifted_medicines_report(page=1, selected_period="Other", search_term=search_var.get(), selected_date=(start_date, end_date))

        select_button = ctk.CTkButton(calendar_window, text="Select Dates", command=on_select_dates)
        select_button.pack(pady=20)

    def toggle_calendar(choice):
        if choice == "Other":
            show_calendar_popup()
        else:
            show_shifted_medicines_report(page=1, selected_period=choice, search_term=search_var.get())

    # Period combobox
    period_options = ["Today", "Last Week", "Last Month", "6 Months", "1 Year", "Lifetime", "Other"]
    period_combobox = ctk.CTkComboBox(
        root,
        values=period_options,
        command=toggle_calendar,
        width=150,
        height=32,
        border_color="#2E86C1",
        button_color="#2E86C1",
        dropdown_hover_color="#3498DB"
    )
    period_combobox.set(selected_period)
    period_combobox.grid(row=0, column=6, padx=20, pady=10, sticky="ne")

    # Header Section
    headers_shifted = [
        ("Shift ID", 10, "center"),
        ("Medicine ID", 10, "center"),
        ("Medicine Name", 20, "center"),
        ("Quantity Shifted", 15, "center"),
        ("Shift Date", 15, "center"),
        ("Shift Time", 15, "center"),
        ("Company Name", 15, "center")
    ]

    # Create header frame
    header_frame = ctk.CTkFrame(root, fg_color="transparent")
    header_frame.grid(row=1, column=0, columnspan=7, sticky="nsew", padx=10)
    
    # Configure header columns
    total_weight = sum(weight for _, weight, _ in headers_shifted)
    for col, (_, weight, _) in enumerate(headers_shifted):
        header_frame.grid_columnconfigure(col, weight=weight)

    # Create headers
    for col, (header, _, _) in enumerate(headers_shifted):
        header_label = ctk.CTkLabel(
            header_frame,
            text=header,
            font=("Arial", 14, "bold"),
            text_color="#FFFFFF",
            fg_color="#2E86C1",
            corner_radius=8,
            height=35,
            anchor="center"
        )
        header_label.grid(row=0, column=col, padx=2, pady=10, sticky="nsew")

    # Data frame
    data_frame = ctk.CTkFrame(
        root,
        fg_color="#F8F9F9",
        corner_radius=10
    )
    data_frame.grid(row=2, column=0, columnspan=7, padx=10, pady=10, sticky="nsew")

    # Configure data frame columns
    for col, (_, weight, _) in enumerate(headers_shifted):
        data_frame.grid_columnconfigure(col, weight=weight)

    try:
        if selected_period == "Other" and selected_date:
            start_date, end_date = selected_date
            shifted_medicines, total_records, total_pages = fetch_filtered_shifted_medicines(
                "Other", page, ITEMS_PER_PAGE, search_term, start_date, end_date
            )
        else:
            shifted_medicines, total_records, total_pages = fetch_filtered_shifted_medicines(
                selected_period, page, ITEMS_PER_PAGE, search_term
            )

        # Update total medicines label
        total_medicines_label = ctk.CTkLabel(
            root,
            text=f"Total Medicines: {total_records}",
            font=("Arial", 14, "bold"),
            text_color="#2E86C1"
        )
        total_medicines_label.grid(row=0, column=5, padx=(0, 10), pady=10, sticky="ne")

        if isinstance(shifted_medicines, list):
            if shifted_medicines:
                for i, medicine in enumerate(shifted_medicines):
                    if isinstance(medicine, tuple) and len(medicine) == 7:
                        shift_id, medicine_id, name, quantity_shifted, shift_date, shift_time, company_name = medicine

                        # Format shift time
                        if shift_time:
                            try:
                                shift_time = datetime.strptime(str(shift_time), '%H:%M:%S').strftime('%I:%M:%S %p')
                            except ValueError:
                                shift_time = "Invalid Time"
                        else:
                            shift_time = "N/A"

                        fields = [
                            (str(shift_id), "center"),
                            (str(medicine_id), "center"),
                            (name, "center"),
                            (str(quantity_shifted), "center"),
                            (str(shift_date), "center"),
                            (shift_time, "center"),
                            (company_name, "center")
                        ]
                        
                        for col, (field, align) in enumerate(fields):
                            medicine_label = ctk.CTkLabel(
                                data_frame,
                                text=field,
                                font=("Arial", 13),
                                text_color="#34495E",
                                fg_color="#FFFFFF" if i % 2 == 0 else "#F7F9F9",
                                corner_radius=8,
                                anchor=align,
                                height=30
                            )
                            padx = (10, 5) if align == "w" else (5, 10) if align == "e" else (5, 5)
                            medicine_label.grid(
                                row=i,
                                column=col,
                                padx=padx,
                                pady=2,
                                sticky="nsew"
                            )
            else:
                ctk.CTkLabel(
                    data_frame,
                    text="No medicines have been shifted.",
                    font=("Arial", 13, "bold"),
                    text_color="#E74C3C",
                    anchor="center",
                    height=40
                ).grid(row=0, column=0, columnspan=len(headers_shifted), pady=20, sticky="nsew")

            # Pagination frame
            pagination_frame = ctk.CTkFrame(root, fg_color="transparent")
            pagination_frame.grid(row=3, column=0, columnspan=7, pady=10, sticky="nsew")

            def change_page(new_page):
                show_shifted_medicines_report(
                    page=new_page,
                    selected_period=selected_period,
                    search_term=search_var.get(),
                    selected_date=selected_date
                )

            # Previous page button
            prev_button = ctk.CTkButton(
                pagination_frame,
                text="←",
                width=40,
                command=lambda: change_page(page - 1) if page > 1 else None,
                fg_color="#2E86C1" if page > 1 else "#B3B6B7",
                state="normal" if page > 1 else "disabled"
            )
            prev_button.grid(row=0, column=0, padx=5)

            # Page numbers
            page_label = ctk.CTkLabel(
                pagination_frame,
                text=f"Page {page} of {total_pages}",
                font=("Arial", 12, "bold")
            )
            page_label.grid(row=0, column=1, padx=10)

            # Next page button
            next_button = ctk.CTkButton(
                pagination_frame,
                text="→",
                width=40,
                command=lambda: change_page(page + 1) if page < total_pages else None,
                fg_color="#2E86C1" if page < total_pages else "#B3B6B7",
                state="normal" if page < total_pages else "disabled"
            )
            next_button.grid(row=0, column=2, padx=5)

        else:
            raise TypeError("Expected a list of tuples from fetch_filtered_shifted_medicines")

    except Exception as e:
        error_label = ctk.CTkLabel(
            data_frame,
            text=f"Error fetching shifted medicines data: {e}",
            font=("Arial", 13, "bold"),
            text_color="#E74C3C",
            anchor="center",
            height=40
        )
        error_label.grid(row=0, column=0, columnspan=len(headers_shifted), pady=20, sticky="nsew")
        
        # Empty pagination frame in case of error
        pagination_frame = ctk.CTkFrame(root, fg_color="transparent")
        pagination_frame.grid(row=3, column=0, columnspan=7, pady=10, sticky="nsew")

    status_frame = ctk.CTkFrame(root, fg_color="transparent")
    status_frame.grid(row=4, column=0, columnspan=7, pady=(5, 10), sticky="nsew")

    records_label = ctk.CTkLabel(
        status_frame,
        text=f"Showing {ITEMS_PER_PAGE} records per page",
        font=("Arial", 12),
        text_color="#7F8C8D"
    )
    records_label.pack(side="right", padx=20)

    # Add export button
    def export_to_excel():
        try:
            # Check if required modules are installed
            try:
                import pandas as pd
                import openpyxl
            except ImportError as e:
                if 'openpyxl' in str(e):
                    if messagebox.askyesno("Missing Dependency", 
                        "The openpyxl module is required for Excel export. Would you like to install it now?"):
                        import subprocess
                        try:
                            subprocess.check_call(['pip', 'install', 'openpyxl'])
                            messagebox.showinfo("Success", "openpyxl installed successfully. Please try exporting again.")
                        except subprocess.CalledProcessError:
                            messagebox.showerror("Installation Failed", 
                                "Could not install openpyxl. Please install it manually using 'pip install openpyxl'")
                    return
                else:
                    messagebox.showerror("Missing Dependency", f"Required module not found: {str(e)}")
                    return

            # Fetch all records for export
            all_records, _, _ = fetch_filtered_shifted_medicines(
                selected_period,
                page=1,
                items_per_page=1000000,  # Large number to get all records
                search_term=search_term,
                start_date=selected_date[0] if selected_date else None,
                end_date=selected_date[1] if selected_date else None
            )
            
            if not all_records:
                messagebox.showwarning("Export Failed", "No records to export.")
                return

            # Convert records to list of dicts for better handling
            formatted_records = []
            for record in all_records:
                shift_id, medicine_id, name, quantity_shifted, shift_date, shift_time, company_name = record
                
                # Improved time formatting logic
                formatted_time = "N/A"
                if shift_time:
                    try:
                        if isinstance(shift_time, str):
                            # Handle string time format
                            formatted_time = datetime.strptime(shift_time, '%H:%M:%S').strftime('%I:%M:%S %p')
                        elif isinstance(shift_time, datetime.time):
                            # Handle time object directly
                            formatted_time = shift_time.strftime('%I:%M:%S %p')
                        elif isinstance(shift_time, datetime.datetime):
                            # Handle datetime object
                            formatted_time = shift_time.strftime('%I:%M:%S %p')
                    except (ValueError, TypeError) as e:
                        # If any error occurs, try to convert to string and format
                        try:
                            time_str = str(shift_time).strip()
                            if ':' in time_str:
                                # Try to parse time in HH:MM:SS format
                                time_parts = time_str.split(':')
                                if len(time_parts) == 3:
                                    formatted_time = datetime.strptime(time_str, '%H:%M:%S').strftime('%I:%M:%S %p')
                                elif len(time_parts) == 2:
                                    formatted_time = datetime.strptime(time_str, '%H:%M').strftime('%I:%M:%S %p')
                        except:
                            formatted_time = str(shift_time)

                formatted_records.append({
                    'Shift ID': shift_id,
                    'Medicine ID': medicine_id,
                    'Medicine Name': name,
                    'Quantity Shifted': quantity_shifted,
                    'Shift Date': shift_date,
                    'Shift Time': formatted_time,
                    'Company Name': company_name
                })

            # Create DataFrame from formatted records
            df = pd.DataFrame(formatted_records)

            # Ask user for save location and format
            file_types = [
                ("Excel files", "*.xlsx"),
                ("CSV files", "*.csv"),
                ("Text files", "*.txt")
            ]
            file_path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=file_types,
                title="Export Shifted Medicines Report"
            )

            if file_path:
                if file_path.endswith('.xlsx'):
                    # Excel export with formatting
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='Shifted Medicines Report')
                        workbook = writer.book
                        worksheet = writer.sheets['Shifted Medicines Report']
                        
                        # Auto-adjust column widths
                        for column in worksheet.columns:
                            max_length = 0
                            column = [cell for cell in column]
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(cell.value)
                                except:
                                    pass
                            adjusted_width = (max_length + 2)
                            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

                        # Add some basic formatting
                        header = openpyxl.styles.PatternFill(
                            start_color='1E88E5',
                            end_color='1E88E5',
                            fill_type='solid'
                        )
                        for cell in worksheet[1]:
                            cell.fill = header
                            cell.font = openpyxl.styles.Font(color='FFFFFF', bold=True)

                elif file_path.endswith('.csv'):
                    df.to_csv(file_path, index=False)
                elif file_path.endswith('.txt'):
                    df.to_csv(file_path, index=False, sep='\t')

                messagebox.showinfo("Export Successful", 
                    f"Report exported successfully to:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Export Failed", 
                f"Error exporting data: {str(e)}\n\nPlease make sure you have permission to write to the selected location.")

    # Update the export button
    export_button = ctk.CTkButton(
        status_frame,
        text="Export Report",
        command=export_to_excel,
        width=120,
        height=30,
        fg_color="#27AE60",
        hover_color="#219A52"
    )
    export_button.pack(side="left", padx=20)

    # Add keyboard shortcuts for navigation
    def handle_keypress(event):
        if event.keysym == "Left" and page > 1:
            show_shifted_medicines_report(
                page=page-1,
                selected_period=selected_period,
                search_term=search_var.get(),
                selected_date=selected_date
            )
        elif event.keysym == "Right" and page < total_pages:
            show_shifted_medicines_report(
                page=page+1,
                selected_period=selected_period,
                search_term=search_var.get(),
                selected_date=selected_date
            )

    root.bind("<Left>", handle_keypress)
    root.bind("<Right>", handle_keypress)

    # Keep search entry focused
    search_entry.focus_set()
def sales_and_billing():
    clear_widgets(root)

    # Set theme and colors
    ctk.set_appearance_mode("light")
    COLORS = {
        'primary': "#2D5AF0",
        'primary_hover': "#1E3FB3",
        'danger': "#FF5252",
        'danger_hover': "#FF1744",
        'background': "#F5F7FF",
        'surface': "#FFFFFF",
        'text': "#1A1A1A",
        'text_secondary': "#666666",
        'border': "#E0E0E0"
    }

    # Database connection (keeping existing code)
    try:
        conn = mysql.connector.connect(
            host="localhost",
            user="root",
            password="123@123",
            database="pharmacy"
        )
        cursor = conn.cursor()
    except Error as e:
        print(f"Error connecting to MySQL Database: {e}")
        messagebox.showerror("Database Error", f"Could not connect to the database: {e}")
        return

    selected_medicines = []
    total_price = 0
    final_total = 0
    discount = 0

    # Database connection
    try:
        conn = mysql.connector.connect(
            host="localhost",
            user="root",
            password="123@123",
            database="pharmacy"
        )
        cursor = conn.cursor()
    except Error as e:
        print(f"Error connecting to MySQL Database: {e}")
        messagebox.showerror("Database Error", f"Could not connect to the database: {e}")

    def fetch_medicines():
        """Fetches medicines from the database including quantity."""
        try:
            cursor.execute("SELECT medicine_id, name, selling_price, quantity FROM medicines")
            return cursor.fetchall()
        except Error as e:
            print(f"An error occurred: {e}")
            return []

    def update_medicine_table(*args):
        """Updates the medicine table based on the search query."""
        search_query = search_var.get().lower().strip()
        medicine_table.delete(*medicine_table.get_children())

        try:
            if search_query:
                cursor.execute("""
                    SELECT medicine_id, name, selling_price, quantity 
                    FROM medicines 
                    WHERE LOWER(name) LIKE %s OR LOWER(medicine_id) LIKE %s
                    """, (f'%{search_query}%', f'%{search_query}%'))
            else:
                cursor.execute("SELECT medicine_id, name, selling_price, quantity FROM medicines")
            
            medicines = cursor.fetchall()

            for i, medicine in enumerate(medicines):
                medicine_table.insert(
                    "", 
                    "end", 
                    values=medicine, 
                    tags=('oddrow' if i % 2 else 'evenrow',)
                )

        except Error as e:
            print(f"Database error while searching: {e}")
            messagebox.showerror("Search Error", "An error occurred while searching the database.")

    def add_medicine_to_selection(event=None):
        """Adds the selected medicine to the selected medicines table."""
        selected_item = medicine_table.focus()
        if not selected_item:
            return

        selected_values = medicine_table.item(selected_item, "values")
        if selected_values:
            medicine_code, medicine_name, medicine_price, available_quantity = selected_values
            medicine_price = float(medicine_price)
            available_quantity = int(available_quantity)

            if available_quantity <= 0:
                messagebox.showwarning("Stock Warning", "This medicine is out of stock!")
                return

            for med in selected_medicines:
                if med['Medicine ID'] == medicine_code:
                    if med['quantity'] >= available_quantity:
                        messagebox.showwarning("Quantity Error", "You cannot add more than the available quantity!")
                        return
                    med['quantity'] += 1
                    med['price'] += medicine_price
                    break
            else:
                selected_medicines.append({
                    "Medicine ID": medicine_code,
                    "name": medicine_name,
                    "quantity": 1,
                    "price": medicine_price,
                    "available_quantity": available_quantity
                })

            update_selected_medicine_table()
            update_totals()

    def remove_medicine(medicine):
        """Removes the specified medicine from the selection."""
        selected_medicines.remove(medicine)
        update_selected_medicine_table()
        update_totals()

    def update_selected_medicine_table():
        """Updates the table of selected medicines."""
        for widget in inner_frame.winfo_children():
            widget.destroy()

        scrollable_frame = ctk.CTkScrollableFrame(
            inner_frame,
            fg_color="#1E1E1E",
            corner_radius=8
        )
        scrollable_frame.pack(fill="both", expand=True, padx=5, pady=5)

        table_frame = ctk.CTkFrame(scrollable_frame, fg_color="#242424", corner_radius=6)
        table_frame.pack(fill="x", padx=2, pady=2)

        headers = [
            ("Medicine ID", 80), 
            ("Medicine Name", 120), 
            ("Unit Price", 80), 
            ("Quantity", 70), 
            ("Total", 80),
            ("", 40)
        ]

        header_frame = ctk.CTkFrame(table_frame, fg_color="#303030", height=40)
        header_frame.pack(fill="x", pady=(0, 1))
        header_frame.pack_propagate(False)

        for idx, (header, width) in enumerate(headers):
            header_frame.grid_columnconfigure(idx, weight=1, minsize=width)
            label = ctk.CTkLabel(
                header_frame,
                text=header,
                font=("Roboto", 11, "bold"),
                text_color="#E0E0E0",
                anchor="w" if idx == 1 else "center"
            )
            label.grid(row=0, column=idx, padx=5, pady=5, sticky="ew")

        for i, medicine in enumerate(selected_medicines):
            row_container = ctk.CTkFrame(
                table_frame,
                fg_color="#2A2A2A" if i % 2 == 0 else "#242424",
                height=45,
                corner_radius=0
            )
            row_container.pack(fill="x", pady=(0, 1))
            row_container.pack_propagate(False)

            for idx, (_, width) in enumerate(headers):
                row_container.grid_columnconfigure(idx, weight=1, minsize=width)

            ctk.CTkLabel(
                row_container,
                text=medicine["Medicine ID"],
                font=("Roboto", 11),
                text_color="#E0E0E0",
                anchor="w"
            ).grid(row=0, column=0, padx=5, pady=5, sticky="w")

            ctk.CTkLabel(
                row_container,
                text=medicine["name"],
                font=("Roboto", 11),
                text_color="#E0E0E0",
                anchor="w"
            ).grid(row=0, column=1, padx=5, pady=5, sticky="w")

            unit_price = medicine['price'] / medicine['quantity'] if medicine['quantity'] > 0 else medicine['price']
            ctk.CTkLabel(
                row_container,
                text=f"{unit_price:.2f}",
                font=("Roboto", 11),
                text_color="#E0E0E0",
                anchor="e"
            ).grid(row=0, column=2, padx=5, pady=5, sticky="e")

            quantity_frame = ctk.CTkFrame(row_container, fg_color="transparent")
            quantity_frame.grid(row=0, column=3, padx=5, pady=3, sticky="ew")
            
            quantity_entry = ctk.CTkEntry(
                quantity_frame,
                width=50,
                height=25,
                corner_radius=4,
                fg_color="#1E1E1E",
                border_color="#404040",
                text_color="#E0E0E0",
                font=("Roboto", 11)
            )
            quantity_entry.pack(expand=True)
            quantity_entry.insert(0, str(medicine["quantity"]))

            ctk.CTkLabel(
                row_container,
                text=f"{medicine['price']:.2f}",
                font=("Roboto", 11),
                text_color="#E0E0E0",
                anchor="e"
            ).grid(row=0, column=4, padx=5, pady=5, sticky="e")

            remove_button = ctk.CTkButton(
                row_container,
                text="×",
                width=25,
                height=25,
                corner_radius=12,
                fg_color="#D32F2F",
                hover_color="#B71C1C",
                text_color="white",
                font=("Roboto", 12, "bold"),
                command=lambda med=medicine: remove_medicine(med)
            )
            remove_button.grid(row=0, column=5, padx=5, pady=5)

            quantity_entry.bind(
                "<KeyRelease>",
                lambda e, med=medicine, entry=quantity_entry: update_quantity(med, entry)
            )

        ctk.CTkFrame(table_frame, fg_color="transparent", height=5).pack(fill="x")

    def update_quantity(medicine, quantity_entry):
        """Updates the quantity and price based on the entry."""
        try:
            new_quantity = int(quantity_entry.get())
            if new_quantity < 0:
                new_quantity = 0
            elif new_quantity > medicine['available_quantity']:
                messagebox.showwarning("Quantity Error", f"You cannot add more than the available quantity ({medicine['available_quantity']})!")
                new_quantity = medicine['available_quantity']
                quantity_entry.delete(0, tk.END)
                quantity_entry.insert(0, str(new_quantity))

            price_per_unit = medicine['price'] / medicine['quantity'] if medicine['quantity'] > 0 else medicine['price']
            medicine['quantity'] = new_quantity
            medicine['price'] = price_per_unit * new_quantity

            update_selected_medicine_table()
            update_totals()
        except ValueError:
            pass

    def update_totals(event=None):
        global total_price, final_total, discount
        
        total_price = sum(med['price'] for med in selected_medicines)
        
        try:
            discount = float(discount_entry.get()) if discount_entry.get() else 0
            # Now discount is directly in PKR, no percentage calculation needed
        except ValueError:
            discount = 0
        
        final_total = max(total_price - discount, 0)
        
        total_value_label.configure(text=f"{total_price:.2f}")
        final_total_value_label.configure(text=f"{final_total:.2f}")
        
        try:
            amount_paid = float(amount_paid_entry.get()) if amount_paid_entry.get() else 0
            change = max(amount_paid - final_total, 0)
            change_value_label.configure(text=f"{change:.2f}")
        except ValueError:
            change_value_label.configure(text="0.00")

    def record_sale(medicine, quantity, sale_date, sale_time):
        """Records the sale in the database."""
        try:
            unit_price = medicine['price'] / medicine['quantity']
            total_price = unit_price * quantity
            discount_amount = float(discount_entry.get() or 0)
            # Now discount_amount is directly in PKR, no percentage calculation needed
            final_price = total_price - discount_amount
            
            amount_paid = float(amount_paid_entry.get() or 0)
            change_given = max(0, amount_paid - final_price)
            cursor.execute("""
                INSERT INTO sales (
                    medicine_id, quantity_sold, selling_price, discount, 
                    total_price, amount_paid, change_given, sale_date, medicine_name
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                medicine['Medicine ID'], quantity, unit_price, discount_amount,
                final_price, amount_paid, change_given, 
                f"{sale_date} {sale_time}", medicine['name']
            ))
            
            cursor.execute("""
                UPDATE medicines SET quantity = quantity - %s WHERE medicine_id = %s
            """, (quantity, medicine['Medicine ID']))
            
            conn.commit()
        except Error as e:
            print(f"An error occurred while recording the sale: {e}")
            messagebox.showerror("Database Error", f"Failed to record sale: {e}")
            conn.rollback()

    def print_bill():
        global total_price, final_total, discount
        
        if not selected_medicines:
            messagebox.showwarning("Selection Error", "No medicines selected for billing.")
            return

        sale_date = datetime.now().strftime('%Y-%m-%d')
        sale_time = datetime.now().strftime('%H:%M:%S')

        header = "Mustafa Eye Care"
        datetime_line = f"Date: {sale_date} Time: {sale_time}"
        bill_text = f"{header}\n{datetime_line}\n=============================="
        
        bill_text += "\nItems:\n"
        for medicine in selected_medicines:
            unit_price = medicine['price'] / medicine['quantity']
            bill_text += f"{medicine['name'][:20]}: {medicine['quantity']}x{unit_price:.2f}={medicine['price']:.2f}\n"
        
        bill_text += "------------------------------\n"
        
        bill_text += f"Total: {total_price:.2f}\n"
        
        try:
            discount_percentage = float(discount_entry.get() or 0)
            discount_amount = total_price * discount_percentage / 100
            if discount_percentage > 0:
                bill_text += f"Disc({discount_percentage}%): -{discount_amount:.2f}\n"
        except ValueError:
            discount_percentage = 0
            discount_amount = 0
        
        bill_text += f"Net Total: {final_total:.2f}\n"
        
        try:
            amount_paid = float(amount_paid_entry.get() or 0)
            if amount_paid > 0:
                change = amount_paid - final_total
                bill_text += f"Paid: {amount_paid:.2f}\nChange: {max(0, change):.2f}\n"
        except ValueError:
            amount_paid = 0
            change = 0
        
        bill_text += "\nدواٸ واپس کرنے کیلۓ بل کا ساتھ ہونا ضروری ہے"
        bill_text += "\nThank you for visiting us!"
        bill_text += "\n====================================================\n"
        bill_text += "\nSoftware By Little Devs of CIT!"
        bill_text += "\nGovernment College of Technology BHAKKAR"
        bill_text += "\nContact us: 03424387577"
        

        try:
            for medicine in selected_medicines:
                record_sale(medicine, medicine['quantity'], sale_date, sale_time)
            
            messagebox.showinfo("Sale Recorded", "Sale has been recorded successfully.")
        except Error as e:
            print(f"An error occurred while recording the sale: {e}")
            messagebox.showerror("Database Error", f"Failed to record sale: {e}")
            conn.rollback()
            return

        print_confirmation = messagebox.askyesno("Print Confirmation", "Do you want to print the bill?")
        
        if print_confirmation:
            print_actual_bill(header, datetime_line, bill_text)
        else:
            messagebox.showinfo("Print Cancelled", "Bill printing was cancelled.")

        sales_and_billing()
    def print_actual_bill(header, datetime_line, bill_text):
        try:
            hprinter = win32print.OpenPrinter(win32print.GetDefaultPrinter())
            try:
                hdc = win32ui.CreateDC()
                hdc.CreatePrinterDC(win32print.GetDefaultPrinter())
                hdc.StartDoc("Bill")
                hdc.StartPage()
                
                # Large bold font for header
                header_font = win32ui.CreateFont({
                    "name": "Courier New",
                    "height": 40,  # Larger size for header
                    "weight": 700,
                })
                
                # Regular bold font for rest of content
                regular_font = win32ui.CreateFont({
                    "name": "Courier New",
                    "height": 25,
                    "weight": 700,
                })
                
                # Print header with larger font
                hdc.SelectObject(header_font)
                hdc.TextOut(0, 0, header)
                
                # Switch to regular font for the rest
                hdc.SelectObject(regular_font)
                
                # Start y position after header
                y = 50  # Adjusted spacing for larger header
                
                # Print datetime line
                hdc.TextOut(0, y, datetime_line)
                y += 35
                
                # Print rest of bill content
                for line in bill_text.split('\n')[2:]:  # Skip header and datetime line
                    hdc.TextOut(0, y, line)
                    y += 35
                
                hdc.EndPage()
                hdc.EndDoc()
            finally:
                win32print.ClosePrinter(hprinter)
            messagebox.showinfo("Success", "Bill printed successfully.")
        except Exception as e:
            messagebox.showerror("Printer Error", f"Printing error: {str(e)}")
        sales_and_billing()
    def focus_search_bar(event):
        """Focuses the search bar when 'S' is pressed.""" 
        search_bar.focus_set()

    def handle_enter(event):
        """Handles Enter key press on medicine table."""
        add_medicine_to_selection()
        quantity_entries = [widget for widget in inner_frame.winfo_children() if isinstance(widget, ctk.CTkEntry)]
        if quantity_entries:
            quantity_entries[-1].focus_set()

    def handle_delete(event):
        """Handles Delete key press on selected medicines."""
        focused_widget = root.focus_get()
        if isinstance(focused_widget, ctk.CTkEntry):
            parent_frame = focused_widget.master
            for medicine in selected_medicines:
                if medicine['name'] in [child.cget('text') for child in parent_frame.winfo_children() if isinstance(child, tk.Label)]:
                    remove_medicine(medicine)
                    break

    def handle_ctrl_p(event):
        """Handles Ctrl+P key press to print bill."""
        print_bill()

    # Keep all existing functions (fetch_medicines, update_medicine_table, etc.)
    # ... (keep all the function definitions from the original code)

    # Main layout
    main_container = ctk.CTkFrame(root, fg_color=COLORS['background'])
    main_container.pack(fill="both", expand=True, padx=20, pady=20)

    # Main content container with shadow effect
    content_frame = ctk.CTkFrame(
        main_container,
        fg_color=COLORS['surface'],
        corner_radius=15,
        border_width=1,
        border_color=COLORS['border']
    )
    content_frame.pack(fill="both", expand=True)

    # Left panel (65% width)
    left_panel = ctk.CTkFrame(content_frame, fg_color="transparent")
    left_panel.pack(side="left", fill="both", expand=True, padx=20, pady=20)

    # Search bar with icon
    search_frame = ctk.CTkFrame(left_panel, fg_color="transparent")
    search_frame.pack(fill="x", pady=(0, 10))
    
    search_var = tk.StringVar()
    search_bar = ctk.CTkEntry(
        search_frame,
        placeholder_text="🔍 Search medicines...",
        height=40,
        corner_radius=10,
        font=("Roboto", 13),
        fg_color=COLORS['background'],
        border_color=COLORS['border'],
        text_color=COLORS['text'],
        textvariable=search_var  # Connect the StringVar
    )
    search_bar.pack(fill="x")

    # Make sure to trace the StringVar for real-time updates
    search_var.trace_add("write", update_medicine_table)

    # Style Treeview
    style = ttk.Style()
    style.theme_use('clam')
    style.configure(
        "Treeview",
        background=COLORS['surface'],
        fieldbackground=COLORS['surface'],
        foreground=COLORS['text'],
        rowheight=35,
        font=("Roboto", 11)
    )
    style.configure(
        "Treeview.Heading",
        background=COLORS['background'],
        foreground=COLORS['text'],
        font=("Roboto", 12, "bold")
    )

    # Medicine table
    table_frame = ctk.CTkFrame(left_panel, fg_color=COLORS['surface'])
    table_frame.pack(fill="both", expand=True)

    medicine_table = ttk.Treeview(
        table_frame,
        columns=("medicine_id", "name", "selling_price", "quantity"),
        show="headings",
        style="Treeview"
    )
    
    # Configure columns
    medicine_table.heading("medicine_id", text="Medicine ID")
    medicine_table.heading("name", text="Medicine Name")
    medicine_table.heading("selling_price", text="Price")
    medicine_table.heading("quantity", text="Stock")
    
    medicine_table.column("medicine_id", width=80)
    medicine_table.column("name", width=200)
    medicine_table.column("selling_price", width=100)
    medicine_table.column("quantity", width=80)

    # Add scrollbar
    scrollbar = ctk.CTkScrollbar(table_frame, command=medicine_table.yview)
    scrollbar.pack(side="right", fill="y")
    medicine_table.configure(yscrollcommand=scrollbar.set)
    medicine_table.pack(fill="both", expand=True)

    # Right panel (35% width)
    right_panel = ctk.CTkFrame(content_frame, fg_color=COLORS['background'], corner_radius=10)
    right_panel.pack(side="right", fill="both", padx=20, pady=20, expand=True, ipadx=20)

    # Selected Medicine section
    ctk.CTkLabel(
        right_panel,
        text="Selected Medicines",
        font=("Roboto", 16, "bold"),
        text_color=COLORS['text']
    ).pack(pady=(0, 10))

    inner_frame = ctk.CTkFrame(right_panel, fg_color=COLORS['surface'], corner_radius=10)
    inner_frame.pack(fill="both", expand=True)

    # Summary section
    summary_frame = ctk.CTkFrame(right_panel, fg_color="transparent")
    summary_frame.pack(fill="x", pady=20)

    # Style for all summary items
    def create_summary_row(parent, label_text, widget_type="label", **kwargs):
        frame = ctk.CTkFrame(parent, fg_color="transparent")
        frame.pack(fill="x", pady=5)
        
        ctk.CTkLabel(
            frame,
            text=label_text,
            font=("Roboto", 14),
            text_color=COLORS['text_secondary']
        ).pack(side="left")
        
        if widget_type == "label":
            widget = ctk.CTkLabel(
                frame,
                text=kwargs.get("text", "0.00"),
                font=("Roboto", 14, "bold"),
                text_color=COLORS['text']
            )
        else:  # entry
            widget = ctk.CTkEntry(
                frame,
                width=100,
                height=32,
                corner_radius=8,
                font=("Roboto", 14),
                fg_color=COLORS['surface'],
                border_color=COLORS['border'],
                placeholder_text=kwargs.get("placeholder_text", "")
            )
        widget.pack(side="right")
        return widget

    # Create summary rows
    total_value_label = create_summary_row(summary_frame, "Total:")
    discount_entry = create_summary_row(summary_frame, "Discount (%):", "entry", placeholder_text="0%")
    final_total_value_label = create_summary_row(summary_frame, "Net Total:")
    amount_paid_entry = create_summary_row(summary_frame, "Amount Paid:", "entry", placeholder_text="Amount")
    change_value_label = create_summary_row(summary_frame, "Change:")

    # Bind events
    discount_entry.bind("<KeyRelease>", update_totals)
    amount_paid_entry.bind("<KeyRelease>", update_totals)

    # Button frame
    button_frame = ctk.CTkFrame(right_panel, fg_color="transparent")
    button_frame.pack(fill="x", pady=(20, 0))

    # Styled buttons
    print_button = ctk.CTkButton(
        button_frame,
        text="Print Bill",
        font=("Roboto", 14, "bold"),
        height=40,
        corner_radius=10,
        fg_color=COLORS['primary'],
        hover_color=COLORS['primary_hover'],
        command=print_bill
    )
    print_button.pack(side="right", padx=5)

    back_button = ctk.CTkButton(
        button_frame,
        text="Back",
        font=("Roboto", 14, "bold"),
        height=40,
        corner_radius=10,
        fg_color=COLORS['danger'],
        hover_color=COLORS['danger_hover'],
        command=show_main_menu
    )
    back_button.pack(side="left", padx=5)

    # Bind keyboard shortcuts
    root.bind("<s>", focus_search_bar)
    root.bind("<Delete>", handle_delete)
    root.bind("<Control-p>", handle_ctrl_p)
    medicine_table.bind("<Double-Button-1>", add_medicine_to_selection)
    medicine_table.bind("<Return>", handle_enter)

    # Initial update of the medicine table
    update_medicine_table()
def Shifting():
    clear_widgets(root)

    # Initialize variables
    selected_medicines = []

    # MySQL Connection Configuration
    db_config = {
        'host': 'localhost',
        'user': 'root',
        'password': '123@123',
        'database': 'pharmacy'
    }

    def fetch_medicines():
        """Fetches medicines from the database including quantity."""
        try:
            with mysql.connector.connect(**db_config) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT medicine_id, name, selling_price, quantity FROM medicines")
                return cursor.fetchall()
        except mysql.connector.Error as e:
            print(f"An error occurred while fetching medicines: {e}")
            return []

    def update_medicine_table(*args):
        """Updates the medicine table based on the search query."""
        search_query = search_var.get().lower()
        medicine_table.delete(*medicine_table.get_children())

        medicines = fetch_medicines()

        for i, medicine in enumerate(medicines):
            if search_query in medicine[1].lower() or search_query == "":
                medicine_table.insert("", "end", values=medicine, tags=('oddrow' if i % 2 else 'evenrow',))

    def add_medicine_to_selection(event=None):
        """Adds the selected medicine to the selected medicines table."""
        selected_item = medicine_table.focus()
        if not selected_item:
            return

        selected_values = medicine_table.item(selected_item, "values")
        if selected_values:  
            medicine_code, medicine_name, medicine_price, available_quantity = selected_values
            medicine_price = float(medicine_price)
            available_quantity = int(available_quantity)

            if available_quantity <= 0:
                messagebox.showwarning("Stock Warning", "This medicine is out of stock!")
                return

            for med in selected_medicines:
                if med['Medicine ID'] == medicine_code:
                    if med['quantity'] >= available_quantity:
                        messagebox.showwarning("Quantity Error", "You cannot add more than the available quantity!")
                        return
                    med['quantity'] += 1
                    med['price'] += medicine_price
                    break
            else:
                selected_medicines.append({
                    "Medicine ID": medicine_code,
                    "name": medicine_name,
                    "quantity": 1,
                    "price": medicine_price,
                    "available_quantity": available_quantity
                })

            update_selected_medicine_table()

    def remove_medicine(medicine):
        """Removes the specified medicine from the selection."""
        selected_medicines.remove(medicine)
        update_selected_medicine_table()

    def update_selected_medicine_table():
        """Updates the table of selected medicines with responsive design."""
        # Clear existing widgets
        for widget in inner_frame.winfo_children():
            widget.destroy()

        # Create scrollable frame
        scrollable_frame = ctk.CTkScrollableFrame(
            inner_frame,
            fg_color="#1E1E1E",  # Dark background
            corner_radius=8
        )
        scrollable_frame.pack(fill="both", expand=True, padx=5, pady=5)

        # Create table container
        table_frame = ctk.CTkFrame(scrollable_frame, fg_color="#242424", corner_radius=6)
        table_frame.pack(fill="both", expand=True, padx=2, pady=2)

        # Header row
        headers = [
            ("Medicine ID", 1), 
            ("Medicine Name", 3), 
            ("Quantity", 1), 
            ("Price", 1),
            ("", 1)
        ]

        # Create header
        header_frame = ctk.CTkFrame(table_frame, fg_color="#303030")
        header_frame.pack(fill="x", pady=(0, 1))

        # Configure header grid
        for idx, (header, weight) in enumerate(headers):
            header_frame.grid_columnconfigure(idx, weight=weight)
            label = ctk.CTkLabel(
                header_frame,
                text=header,
                font=("Roboto", 11, "bold"),
                text_color="#E0E0E0",
                anchor="w" if idx == 1 else "center"  # Left align medicine name, center others
            )
            label.grid(row=0, column=idx, padx=5, pady=5, sticky="ew")

        # Add items
        for i, medicine in enumerate(selected_medicines):
            # Row container with border effect
            row_container = ctk.CTkFrame(
                table_frame,
                fg_color="#2A2A2A" if i % 2 == 0 else "#242424",  # Alternate row colors
                corner_radius=0
            )
            row_container.pack(fill="x", pady=(0, 1))

            # Configure grid columns
            for idx, (_, weight) in enumerate(headers):
                row_container.grid_columnconfigure(idx, weight=weight)

            # Serial number
            ctk.CTkLabel(
                row_container,
                text=medicine["Medicine ID"],
                font=("Roboto", 11),
                text_color="#E0E0E0",
                anchor="w"
            ).grid(row=0, column=0, padx=5, pady=5, sticky="ew")

            # Medicine name
            ctk.CTkLabel(
                row_container,
                text=medicine["name"],
                font=("Roboto", 11),
                text_color="#E0E0E0",
                anchor="w"
            ).grid(row=0, column=1, padx=5, pady=5, sticky="ew")

            # Quantity entry
            quantity_entry = ctk.CTkEntry(
                row_container,
                height=25,
                corner_radius=4,
                fg_color="#1E1E1E",
                border_color="#404040",
                text_color="#E0E0E0",
                font=("Roboto", 11)
            )
            quantity_entry.grid(row=0, column=2, padx=5, pady=5, sticky="ew")
            quantity_entry.insert(0, str(medicine["quantity"]))

            # Price
            ctk.CTkLabel(
                row_container,
                text=f"{medicine['price']:.2f}",
                font=("Roboto", 11),
                text_color="#E0E0E0",
                anchor="e"
            ).grid(row=0, column=3, padx=5, pady=5, sticky="ew")

            # Remove button
            remove_button = ctk.CTkButton(
                row_container,
                text="×",
                width=25,
                height=25,
                corner_radius=12,
                fg_color="#D32F2F",
                hover_color="#B71C1C",
                text_color="white",
                font=("Roboto", 12, "bold"),
                command=lambda med=medicine: remove_medicine(med)
            )
            remove_button.grid(row=0, column=4, padx=5, pady=5, sticky="ew")

            # Bind quantity update
            quantity_entry.bind(
                "<KeyRelease>",
                lambda e, med=medicine, entry=quantity_entry: update_quantity(med, entry)
            )

        # Add bottom padding frame
        ctk.CTkFrame(table_frame, fg_color="transparent", height=5).pack(fill="x")

        # Make the table frame expand to fill available space
        table_frame.pack(fill="both", expand=True)
    def update_quantity(medicine, quantity_entry):
        """Updates the quantity of a medicine in the selection."""
        try:
            new_quantity = int(quantity_entry.get())
            if new_quantity <= 0:
                messagebox.showwarning("Quantity Error", "Quantity must be greater than 0.")
                quantity_entry.delete(0, tk.END)
                quantity_entry.insert(0, str(medicine["quantity"]))
            elif new_quantity > medicine["available_quantity"]:
                messagebox.showwarning("Quantity Error", f"You cannot select more than the available quantity ({medicine['available_quantity']}).")
                quantity_entry.delete(0, tk.END)
                quantity_entry.insert(0, str(medicine["quantity"]))
            else:
                medicine["quantity"] = new_quantity
                medicine["price"] = (medicine["price"] / medicine["quantity"]) * new_quantity
                update_selected_medicine_table()
        except ValueError:
            messagebox.showwarning("Input Error", "Please enter a valid number for quantity.")
            quantity_entry.delete(0, tk.END)
            quantity_entry.insert(0, str(medicine["quantity"]))

    def ship_medicines():
        
        if not selected_medicines:
            messagebox.showwarning("Selection Error", "No medicines selected for Shifting.")
            return

        company_name = company_entry.get().strip()
        if not company_name:
            messagebox.showwarning("Input Error", "Please enter a company name.")
            return

        # Check if the company name is a valid string (not just numbers)
        if company_name.isdigit():
            messagebox.showwarning("Input Error", "Company name cannot be just numbers. Please enter a valid name.")
            return

        # Check if any selected quantity exceeds available quantity
        for medicine in selected_medicines:
            if medicine['quantity'] > medicine['available_quantity']:
                messagebox.showwarning("Quantity Error", f"Selected quantity for {medicine['name']} exceeds available quantity.")
                return

        confirmation = messagebox.askyesno("Confirm Shifting", f"Are you sure you want to ship these medicines to {company_name}?")
        if confirmation:
            try:
                with mysql.connector.connect(**db_config) as conn:
                    cursor = conn.cursor()
                    
                    for medicine in selected_medicines:
                        # Deduct the quantity from the medicines table
                        cursor.execute("""UPDATE medicines 
                                        SET quantity = quantity - %s 
                                        WHERE medicine_id = %s""",
                                    (medicine['quantity'], medicine['Medicine ID']))

                        # Insert into shifted_medicines table with the correct quantity and company name
                        cursor.execute("""INSERT INTO shifted_medicines 
                                        (medicine_id, medicine_name, quantity_shifted, company_name) 
                                        VALUES (%s, %s, %s, %s)""",
                                    (medicine['Medicine ID'], medicine['name'], medicine['quantity'], company_name))

                    conn.commit()
                    messagebox.showinfo("Success", f"Medicines have been shipped successfully to {company_name}.")
                    
                    # Clear the selected medicines and reset tables
                    selected_medicines.clear()
                    update_selected_medicine_table()
                    update_medicine_table()
                    company_entry.delete(0, tk.END)

            except mysql.connector.Error as e:
                print(f"An error occurred while updating the database: {e}")
                conn.rollback()
                messagebox.showerror("Database Error", f"An error occurred while updating the database: {e}")

    # UI Setup
    left_frame = ctk.CTkFrame(root, height=500, fg_color="white")
    left_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)
    right_frame = ctk.CTkFrame(root, height=500, fg_color="white")
    right_frame.pack(side="right", fill="both", expand=True, padx=10, pady=10)

    search_var = tk.StringVar()
    search_bar = ctk.CTkEntry(left_frame, textvariable=search_var, placeholder_text="Search Medicines")
    search_bar.pack(ipadx=250,pady=10)
    search_bar.bind("<KeyRelease>", update_medicine_table)

    medicine_table = ttk.Treeview(left_frame, columns=("medicine_id", "name", "selling_price", "quantity"), show="headings")
    medicine_table.heading("medicine_id", text="Medicine ID")
    medicine_table.heading("name", text="Medicine Name")
    medicine_table.heading("selling_price", text="Price")
    medicine_table.heading("quantity", text="Available Quantity")
    medicine_table.pack(expand=True, fill="both")
    medicine_table.bind("<Double-Button-1>", add_medicine_to_selection)

    inner_frame = ctk.CTkFrame(right_frame, fg_color="white")
    inner_frame.pack(expand=True, fill="both")

    company_label = tk.Label(right_frame, text="Company Name:", bg="white")
    company_label.pack(pady=5)
    company_entry = ctk.CTkEntry(right_frame, width=20, placeholder_text="Enter Company Name")
    company_entry.pack(pady=5, padx=10, ipadx=50)

    ship_button = ctk.CTkButton(right_frame, text="Ship", command=ship_medicines)
    ship_button.pack(pady=10)

    ctk.CTkButton(right_frame, text="Back", hover_color="red", command=show_main_menu).pack(pady=10, padx=20)

    update_medicine_table()
def show_combined_medicines_report():
    clear_widgets(root)
    reset_grid_configuration(root)

    # Configure the root grid layout for full expansion
    root.grid_rowconfigure(1, weight=1)  # Make the tab view expandable
    for col in range(12):  # Configure 12 columns to ensure even distribution
        root.grid_columnconfigure(col, weight=1)

    # Title Section
    title_label = ctk.CTkLabel(
        root, 
        text="Dashboard", 
        font=("Arial", 26, "bold"),
        text_color="#2C3E50"  # Darker, more professional color
    )
    title_label.grid(row=0, column=0, columnspan=12, pady=10, sticky="nsew")

    # Back button section
    back_icon_image = ctk.CTkImage(Image.open("back.png"), size=(30, 30))
    
    # Back button
    back_button = ctk.CTkButton(
        root,
        image=back_icon_image,
        text="",
        width=40,
        height=40,  # Make the button square and compact
        command=show_main_menu,
        fg_color="#ECF0F1",  # Light background color
        hover_color="#BDC3C7",  # Subtle hover effect
        corner_radius=10,  # Rounded corners for a modern look
        border_width=2,  # Add a slight border for definition
        border_color="#34495E"
    )
    back_button.grid(row=0, column=0, padx=15, pady=10, sticky="nw")

    # Create tabs for Added, Sold, Shipped, and Expired medicines
    tab_view = ctk.CTkTabview(root, segmented_button_selected_color="#3498DB", segmented_button_unselected_color="#E5E8E8")
    tab_view.grid(row=1, column=0, columnspan=12, padx=20, pady=10, sticky="nsew")

    # Adding the tabs
    added_tab = tab_view.add("Added Medicines")
    sold_tab = tab_view.add("Sold Medicines")
    shipped_tab = tab_view.add("Shipped Medicines")
    expired_tab = tab_view.add("Expired Medicines")

    # Configure tab views to expand fully
    for tab in [added_tab, sold_tab, shipped_tab, expired_tab]:
        tab.grid_rowconfigure(0, weight=1)
        tab.grid_columnconfigure(0, weight=1)

    # Function to fetch data for each report
    def fetch_data(query):
        try:
            connection = mysql.connector.connect(
                host='localhost',
                database='pharmacy',
                user='root',
                password='123@123'
            )
            if connection.is_connected():
                cursor = connection.cursor()
                cursor.execute(query)
                result = cursor.fetchall()
                cursor.close()
                connection.close()
                return result
        except Error as e:
            print(f"Error: {e}")
            return []

    # Function to create and populate a report frame
    def create_report_frame(parent, headers, query):
        frame = ctk.CTkScrollableFrame(parent, fg_color="#F2F3F4")  # Subtle background for the scrollable area
        frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        frame.grid_columnconfigure(tuple(range(len(headers))), weight=1)
        
        # Add headers
        for col, header in enumerate(headers):
            ctk.CTkLabel(
                frame, 
                text=header, 
                font=("Arial", 15, "bold"), 
                text_color="#2C3E50",  # Darker text for headers
                fg_color="#D6DBDF",  # Light background for header row
                corner_radius=8,
                padx=5
            ).grid(row=0, column=col, padx=5, pady=10, sticky="nsew")
        
        # Fetch and display data
        data = fetch_data(query)
        if data:
            for i, row in enumerate(data, start=1):
                for j, value in enumerate(row):
                    if isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %I:%M:%S %p')
                    elif isinstance(value, date):
                        value = value.strftime('%Y-%m-%d')
                    elif isinstance(value, time):
                        value = value.strftime('%I:%M:%S %p')
                    ctk.CTkLabel(
                        frame, 
                        text=str(value), 
                        font=("Arial", 13), 
                        text_color="#34495E",  # Use a dark gray for the text
                        fg_color="#FFFFFF" if i % 2 == 0 else "#F7F9F9",  # Alternating row colors for better readability
                        corner_radius=8, 
                        padx=5
                    ).grid(row=i, column=j, padx=5, pady=5, sticky="nsew")
        else:
            ctk.CTkLabel(
                frame, 
                text="No data available for today.", 
                font=("Arial", 13, "italic"), 
                text_color="#E74C3C"  # Red text for no data message
            ).grid(row=1, column=0, columnspan=len(headers), pady=7, sticky="nsew")

        ctk.CTkLabel(frame, text="").grid(row=len(data)+1 if data else 2, column=0, columnspan=len(headers))

    # Added Medicines Report (from snapshot_medicines table)
    added_headers = ["Medicine ID", "Name", "Quantity", "Store Date", "Store Time", "Version"]
    added_query = """
        SELECT medicine_id, name, quantity, store, store_time, version
        FROM snapshot_medicines
        WHERE DATE(created_at) = CURDATE()
        ORDER BY created_at DESC
    """
    create_report_frame(added_tab, added_headers, added_query)

    # Sold Medicines Report (from sales table)
    sold_headers = ["Sale ID", "Medicine ID", "Medicine Name", "Quantity Sold", "Selling Price", "Total Price", "Sale Date"]
    sold_query = """
        SELECT sale_id, medicine_id, medicine_name, quantity_sold, selling_price, total_price, sale_date
        FROM sales
        WHERE DATE(sale_date) = CURDATE()
        ORDER BY sale_date DESC
    """
    create_report_frame(sold_tab, sold_headers, sold_query)

    # Shipped Medicines Report (from shifted_medicines table)
    shipped_headers = ["Shift ID", "Medicine ID", "Medicine Name", "Quantity Shifted", "Company Name", "Shifted At"]
    shipped_query = """
        SELECT shift_id, medicine_id, medicine_name, quantity_shifted, company_name, shifted_at
        FROM shifted_medicines
        WHERE DATE(shifted_at) = CURDATE()
        ORDER BY shifted_at DESC
    """
    create_report_frame(shipped_tab, shipped_headers, shipped_query)

    # Expired Medicines Report (from medicines table)
    expired_headers = ["Medicine ID", "Name", "Quantity", "Expiry Date", "Rack", "Company Name"]
    expired_query = """
        SELECT medicine_id, name, quantity, expiry, rack, company_name
        FROM medicines
        WHERE DATE(expiry) = CURDATE()
        ORDER BY name
    """
    create_report_frame(expired_tab, expired_headers, expired_query)

def show_add_return():
    clear_widgets(root)
    back_icon_image = ctk.CTkImage(Image.open("back.png"), size=(40, 40))

    search_frame = ctk.CTkFrame(root, fg_color="transparent")
    search_frame.pack(side="left", padx=10, pady=10, fill="y")
    back_button = ctk.CTkButton(
        search_frame,
        image=back_icon_image,
        text="",
        width=40,
        command=show_return_management,
        fg_color="transparent",
        hover_color="#E5E7E9"
    )
    back_button.grid(row=0, column=0, padx=(0, 5), pady=0, sticky="nw")

    ctk.CTkLabel(root, text="Sale ID").pack(pady=0, padx=20)
    entry_sale_id = ctk.CTkEntry(root, placeholder_text="1")
    entry_sale_id.pack(pady=0, padx=20)

    ctk.CTkLabel(root, text="Medicine ID").pack(pady=0, padx=20)
    entry_medicine_id = ctk.CTkEntry(root, placeholder_text="123")
    entry_medicine_id.pack(pady=0, padx=20)

    # Label to show the medicine name after lookup
    medicine_name_label = ctk.CTkLabel(root, text="Medicine Name: ")
    medicine_name_label.pack(pady=10, padx=20)

    # Function to get medicine name based on entered medicine_id
    def get_medicine_name(event=None):
        medicine_id = entry_medicine_id.get()  # Get the medicine ID from the entry field

        if not medicine_id.isdigit():  # Ensure medicine_id is valid
            medicine_name_label.configure(text="Medicine Name: Invalid ID")
            return

        try:
            # Query to fetch the medicine name from the database
            query = "SELECT name FROM medicines WHERE medicine_id = %s"
            cursor = connection.cursor()
            cursor.execute(query, (medicine_id,))
            result = cursor.fetchone()
            cursor.close()

            if result:
                medicine_name = result[0]  # Fetch the medicine name from the result
                medicine_name_label.configure(text=f"Medicine Name: {medicine_name}")
            else:
                medicine_name_label.configure(text="Medicine Name: Not Found")

        except Error as e:
            print(f"Error fetching medicine name: {e}")
            messagebox.showerror("Error", "Failed to fetch medicine name.")

    # Bind the get_medicine_name function to the key release event
    entry_medicine_id.bind("<KeyRelease>", get_medicine_name)

    ctk.CTkLabel(root, text="Return Quantity").pack(pady=0, padx=20)
    entry_return_quantity = ctk.CTkEntry(root, placeholder_text="10")
    entry_return_quantity.pack(pady=0, padx=20)

    ctk.CTkLabel(root, text="Refund Amount").pack(pady=0, padx=20)
    entry_refund_amount = ctk.CTkEntry(root, placeholder_text="100.00")
    entry_refund_amount.pack(pady=0, padx=20)

    ctk.CTkLabel(root, text="Return Reason").pack(pady=0, padx=20)
    entry_return_reason = ctk.CTkEntry(root, placeholder_text="Expired product")
    entry_return_reason.pack(pady=0, padx=20)

    # Check if the sale ID exists in the sales table
    def check_sale_exists(connection, sale_id):
        """Check if the sale ID exists in the sales table."""
        query = '''
        SELECT COUNT(*) FROM sales WHERE sale_id = %s
        '''
        cursor = connection.cursor()
        cursor.execute(query, (sale_id,))
        count = cursor.fetchone()[0]
        cursor.close()
        return count > 0  # Returns True if sale_id exists

    # Add return function
    def add_return():
        sale_id = entry_sale_id.get()  # Get Sale ID
        medicine_id = entry_medicine_id.get()  # Get Medicine ID
        return_quantity = entry_return_quantity.get()  # Get Return Quantity
        refund_amount = entry_refund_amount.get()  # Get Refund Amount
        return_reason = entry_return_reason.get()  # Get Return Reason

        # Validate required fields
        if not sale_id or not medicine_id or not return_quantity or not refund_amount:
            messagebox.showerror("Error", "All fields except Return Reason are required.")
            return

        # Validate that Sale ID, Medicine ID, and Return Quantity are integers
        if not sale_id.isdigit() or not medicine_id.isdigit() or not return_quantity.isdigit():
            messagebox.showerror("Error", "Sale ID, Medicine ID, and Return Quantity must be integers.")
            return

        # Validate that refund amount is a valid decimal
        try:
            float(refund_amount)
        except ValueError:
            messagebox.showerror("Error", "Refund Amount must be a valid decimal number.")
            return

        try:
            # Ensure the sale_id exists in the sales table
            if not check_sale_exists(connection, sale_id):
                messagebox.showerror("Error", "Sale ID does not exist.")
                return

            # Insert the new return into the `returns` table
            query = '''
            INSERT INTO returns (sale_id, medicine_id, return_quantity, refund_amount, return_reason)
            VALUES (%s, %s, %s, %s, %s)
            '''
            params = (sale_id, medicine_id, return_quantity, refund_amount, return_reason)
            execute_query(query, params)

            # Update the quantity in the medicines table
            update_query = '''
            UPDATE medicines
            SET quantity = quantity + %s
            WHERE medicine_id = %s
            '''
            update_params = (return_quantity, medicine_id)
            execute_query(update_query, update_params)

            messagebox.showinfo("Success", "Return processed successfully.")
            show_return_management()
        except Error as e:
            print(f"Error during database insertion: {e}")
            messagebox.showerror("Error", f"Failed to process return: {e}")

    ctk.CTkButton(root, text="Return", command=add_return).pack(pady=10, padx=20)

# Placeholder function for database execution
def execute_query(query, params):
    try:
        connection = connect(
            host="localhost",
            user="root",
            password="123@123",
            database="pharmacy"  # Your database name
        )
        cursor = connection.cursor()
        cursor.execute(query, params)
        connection.commit()
        cursor.close()
        connection.close()
    except Error as e:
        print(f"Database error: {e}")
import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk
from PIL import Image
from tkcalendar import DateEntry
from datetime import datetime, timedelta
from math import ceil

def fetch_filtered_returns(period, page=1, items_per_page=15, search_term="", start_date=None, end_date=None):
    """Function to fetch filtered returns with pagination."""
    current_date = datetime.now().date()
    offset = (page - 1) * items_per_page

    # Base query for the returns table
    base_query = """
        SELECT r.return_id, m.name, r.return_quantity, r.refund_amount, r.return_reason, r.return_date 
        FROM returns r
        JOIN medicines m ON r.medicine_id = m.medicine_id
    """

    # Count query for total records
    count_query = "SELECT COUNT(*) FROM returns r JOIN medicines m ON r.medicine_id = m.medicine_id"

    # Adjust the filter based on the selected period
    if period == "Today":
        filter_date = current_date
        where_clause = " WHERE DATE(r.return_date) = %s"
        params = (filter_date,)
    elif period == "Last Week":
        filter_date = current_date - timedelta(days=7)
        where_clause = " WHERE r.return_date >= %s"
        params = (filter_date,)
    elif period == "Last Month":
        filter_date = current_date - timedelta(days=30)
        where_clause = " WHERE r.return_date >= %s"
        params = (filter_date,)
    elif period == "6 Months":
        filter_date = current_date - timedelta(days=180)
        where_clause = " WHERE r.return_date >= %s"
        params = (filter_date,)
    elif period == "1 Year":
        filter_date = current_date - timedelta(days=365)
        where_clause = " WHERE r.return_date >= %s"
        params = (filter_date,)
    elif period == "Other" and start_date and end_date:
        where_clause = " WHERE r.return_date BETWEEN %s AND %s"
        params = (start_date, end_date)
    else:  # Lifetime
        where_clause = ""
        params = ()

    # If a search term is provided, add it to the query
    if search_term:
        if where_clause:
            where_clause += " AND m.name LIKE %s"
        else:
            where_clause = " WHERE m.name LIKE %s"
        params += (f"%{search_term}%",)

    # Add where clause to queries
    if where_clause:
        base_query += where_clause
        count_query += where_clause

    # Get total count
    total_records = fetch_data(count_query, params)[0][0]
    total_pages = ceil(total_records / items_per_page)

    # Add pagination to base query
    base_query += " ORDER BY r.return_date DESC LIMIT %s OFFSET %s"
    params += (items_per_page, offset)

    return fetch_data(base_query, params), total_records, total_pages

def show_returns_report(page=1, selected_period="Lifetime", search_term="", selected_date=None):
    clear_widgets(root)
    reset_grid_configuration(root)
    
    ITEMS_PER_PAGE = 15  # Number of items per page
    
    # Configure column weights for proper alignment
    for col in range(6):
        root.grid_columnconfigure(col, weight=1)

    # Title Section
    title_label = ctk.CTkLabel(
        root, 
        text="Medicine Returns Report", 
        font=("Arial", 24, "bold"),
        text_color="#2E86C1"
    )
    title_label.grid(row=0, column=0, columnspan=6, pady=20, sticky="nsew")

    # Back button and search section
    back_icon_image = ctk.CTkImage(Image.open("back.png"), size=(30, 30))
    
    # Search frame setup
    search_frame = ctk.CTkFrame(root, fg_color="transparent")
    search_frame.grid(row=0, column=0, padx=20, pady=10, sticky="nw")
    
    # Back button and search entry
    back_button = ctk.CTkButton(
        search_frame,
        image=back_icon_image,
        text="",
        width=40,
        command=show_return_management,
        fg_color="transparent",
        hover_color="#E5E7E9"
    )
    back_button.grid(row=0, column=0, padx=(0, 5), pady=0, sticky="nw")

    search_var = tk.StringVar()
    search_entry = ctk.CTkEntry(
        search_frame,
        placeholder_text="🔍 Search Medicine",
        width=200,
        height=32,
        border_color="#2E86C1",
        corner_radius=8,
        textvariable=search_var
    )
    search_entry.grid(row=0, column=1, padx=(0, 5), pady=6, sticky="nw")
    search_entry.insert(0, search_term)
    search_entry.focus_set()

    def update_report(*args):
        show_returns_report(page=1, selected_period=selected_period, search_term=search_var.get(), selected_date=selected_date)

    search_var.trace_add("write", update_report)
    
    # Calendar popup function
    def show_calendar_popup():
        calendar_window = tk.Toplevel(root)
        calendar_window.title("Select Date Range")
        calendar_window.geometry("300x300")

        start_date_label = ctk.CTkLabel(calendar_window, text="From:", font=("Arial", 12, "bold"))
        start_date_label.pack(pady=(10, 0))

        start_cal = DateEntry(calendar_window, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        start_cal.pack(pady=(5, 10))

        end_date_label = ctk.CTkLabel(calendar_window, text="End Date:", font=("Arial", 12, "bold"))
        end_date_label.pack(pady=(10, 0))

        end_cal = DateEntry(calendar_window, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        end_cal.pack(pady=(5, 10))

        def on_select_dates():
            start_date = start_cal.get_date()
            end_date = end_cal.get_date()
            calendar_window.destroy()
            show_returns_report(page=1, selected_period="Other", search_term=search_var.get(), selected_date=(start_date, end_date))

        select_button = ctk.CTkButton(calendar_window, text="Select Dates", command=on_select_dates)
        select_button.pack(pady=20)

    def toggle_calendar(choice):
        if choice == "Other":
            show_calendar_popup()
        else:
            show_returns_report(page=1, selected_period=choice, search_term=search_var.get())

    # Period combobox
    period_options = ["Today", "Last Week", "Last Month", "6 Months", "1 Year", "Lifetime", "Other"]
    period_combobox = ctk.CTkComboBox(
        root,
        values=period_options,
        command=toggle_calendar,
        width=150,
        height=32,
        border_color="#2E86C1",
        button_color="#2E86C1",
        dropdown_hover_color="#3498DB"
    )
    period_combobox.set(selected_period)
    period_combobox.grid(row=0, column=5, padx=20, pady=10, sticky="ne")

    # Header Section
    headers_returns = [
        ("Return ID", 10, "center"),
        ("Medicine Name", 20, "center"),
        ("Quantity", 15, "center"),
        ("Refund Amount", 15, "center"),
        ("Return Reason", 25, "center"),
        ("Return Date", 15, "center")
    ]

    # Create header frame
    header_frame = ctk.CTkFrame(root, fg_color="transparent")
    header_frame.grid(row=1, column=0, columnspan=6, sticky="nsew", padx=10)
    
    # Configure header columns
    total_weight = sum(weight for _, weight, _ in headers_returns)
    for col, (_, weight, _) in enumerate(headers_returns):
        header_frame.grid_columnconfigure(col, weight=weight)

    # Create headers
    for col, (header, _, _) in enumerate(headers_returns):
        header_label = ctk.CTkLabel(
            header_frame,
            text=header,
            font=("Arial", 14, "bold"),
            text_color="#FFFFFF",
            fg_color="#2E86C1",
            corner_radius=8,
            height=35,
            anchor="center"
        )
        header_label.grid(row=0, column=col, padx=2, pady=10, sticky="nsew")

    # Data frame
    data_frame = ctk.CTkFrame(
        root,
        fg_color="#F8F9F9",
        corner_radius=10
    )
    data_frame.grid(row=2, column=0, columnspan=6, padx=10, pady=10, sticky="nsew")

    # Configure data frame columns
    for col, (_, weight, _) in enumerate(headers_returns):
        data_frame.grid_columnconfigure(col, weight=weight)

    try:
        if selected_period == "Other" and selected_date:
            start_date, end_date = selected_date
            returns_data, total_records, total_pages = fetch_filtered_returns(
                "Other", page, ITEMS_PER_PAGE, search_term, start_date, end_date
            )
        else:
            returns_data, total_records, total_pages = fetch_filtered_returns(
                selected_period, page, ITEMS_PER_PAGE, search_term
            )

        # Update total returns label
        total_returns_label = ctk.CTkLabel(
            root,
            text=f"Total Returns: {total_records}",
            font=("Arial", 14, "bold"),
            text_color="#2E86C1"
        )
        total_returns_label.grid(row=0, column=4, padx=(0, 10), pady=10, sticky="ne")

        if isinstance(returns_data, list):
            if returns_data:
                for i, return_item in enumerate(returns_data):
                    if isinstance(return_item, tuple) and len(return_item) == 6:
                        return_id, medicine_name, return_quantity, refund_amount, return_reason, return_date = return_item

                        # Format return date
                        if return_date:
                            try:
                                return_date = return_date.strftime('%Y-%m-%d %I:%M:%S %p')
                            except AttributeError:
                                return_date = str(return_date)
                        else:
                            return_date = "N/A"

                        fields = [
                            (str(return_id), "center"),
                            (medicine_name, "center"),
                            (str(return_quantity), "center"),
                            (f"Rs{refund_amount:.2f}", "center"),
                            (return_reason, "center"),
                            (return_date, "center")
                        ]
                        
                        for col, (field, align) in enumerate(fields):
                            return_label = ctk.CTkLabel(
                                data_frame,
                                text=field,
                                font=("Arial", 13),
                                text_color="#34495E",
                                fg_color="#FFFFFF" if i % 2 == 0 else "#F7F9F9",
                                corner_radius=8,
                                anchor=align,
                                height=30
                            )
                            padx = (10, 5) if align == "w" else (5, 10) if align == "e" else (5, 5)
                            return_label.grid(
                                row=i,
                                column=col,
                                padx=padx,
                                pady=2,
                                sticky="nsew"
                            )
            else:
                ctk.CTkLabel(
                    data_frame,
                    text="No returns found for the selected period.",
                    font=("Arial", 13, "bold"),
                    text_color="#E74C3C",
                    anchor="center",
                    height=40
                ).grid(row=0, column=0, columnspan=len(headers_returns), pady=20, sticky="nsew")

            # Pagination frame
            pagination_frame = ctk.CTkFrame(root, fg_color="transparent")
            pagination_frame.grid(row=3, column=0, columnspan=6, pady=10, sticky="nsew")

            def change_page(new_page):
                show_returns_report(
                    page=new_page,
                    selected_period=selected_period,
                    search_term=search_var.get(),
                    selected_date=selected_date
                )

            # Previous page button
            prev_button = ctk.CTkButton(
                pagination_frame,
                text="←",
                width=40,
                command=lambda: change_page(page - 1) if page > 1 else None,
                fg_color="#2E86C1" if page > 1 else "#B3B6B7",
                state="normal" if page > 1 else "disabled"
            )
            prev_button.grid(row=0, column=0, padx=5)

            # Page numbers
            page_label = ctk.CTkLabel(
                pagination_frame,
                text=f"Page {page} of {total_pages}",
                font=("Arial", 12, "bold")
            )
            page_label.grid(row=0, column=1, padx=10)

            # Next page button
            next_button = ctk.CTkButton(
                pagination_frame,
                text="→",
                width=40,
                command=lambda: change_page(page + 1) if page < total_pages else None,
                fg_color="#2E86C1" if page < total_pages else "#B3B6B7",
                state="normal" if page < total_pages else "disabled"
            )
            next_button.grid(row=0, column=2, padx=5)

        else:
            raise TypeError("Expected a list of tuples from fetch_filtered_returns")

    except Exception as e:
        error_label = ctk.CTkLabel(
            data_frame,
            text=f"Error fetching returns data: {e}",
            font=("Arial", 13, "bold"),
            text_color="#E74C3C",
            anchor="center",
            height=40
        )
        error_label.grid(row=0, column=0, columnspan=len(headers_returns), pady=20, sticky="nsew")
        
# Empty pagination frame in case of error
        pagination_frame = ctk.CTkFrame(root, fg_color="transparent")
        pagination_frame.grid(row=3, column=0, columnspan=6, pady=10, sticky="nsew")

    status_frame = ctk.CTkFrame(root, fg_color="transparent")
    status_frame.grid(row=4, column=0, columnspan=6, pady=(5, 10), sticky="nsew")

    records_label = ctk.CTkLabel(
        status_frame,
        text=f"Showing {ITEMS_PER_PAGE} records per page",
        font=("Arial", 12),
        text_color="#7F8C8D"
    )
    records_label.pack(side="right", padx=20)

    # Add export button
    def export_to_excel():
        try:
            # Check if required modules are installed
            try:
                import pandas as pd
                import openpyxl
            except ImportError as e:
                if 'openpyxl' in str(e):
                    if messagebox.askyesno("Missing Dependency", 
                        "The openpyxl module is required for Excel export. Would you like to install it now?"):
                        import subprocess
                        try:
                            subprocess.check_call(['pip', 'install', 'openpyxl'])
                            messagebox.showinfo("Success", "openpyxl installed successfully. Please try exporting again.")
                        except subprocess.CalledProcessError:
                            messagebox.showerror("Installation Failed", 
                                "Could not install openpyxl. Please install it manually using 'pip install openpyxl'")
                    return
                else:
                    messagebox.showerror("Missing Dependency", f"Required module not found: {str(e)}")
                    return

            # Fetch all records for export
            all_records, _, _ = fetch_filtered_returns(
                selected_period,
                page=1,
                items_per_page=1000000,  # Large number to get all records
                search_term=search_term,
                start_date=selected_date[0] if selected_date else None,
                end_date=selected_date[1] if selected_date else None
            )
            
            if not all_records:
                messagebox.showwarning("Export Failed", "No records to export.")
                return

            # Convert records to list of dicts for better handling
            formatted_records = []
            for record in all_records:
                return_id, medicine_name, return_quantity, refund_amount, return_reason, return_date = record
                
                # Format date string
                try:
                    if isinstance(return_date, str):
                        formatted_date = datetime.strptime(return_date, '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %I:%M:%S %p')
                    else:
                        formatted_date = return_date.strftime('%Y-%m-%d %I:%M:%S %p') if return_date else "N/A"
                except:
                    formatted_date = "N/A"

                formatted_records.append({
                    'Return ID': return_id,
                    'Medicine Name': medicine_name,
                    'Quantity Returned': return_quantity,
                    'Refund Amount': refund_amount,
                    'Return Reason': return_reason,
                    'Return Date': formatted_date
                })

            # Create DataFrame from formatted records
            df = pd.DataFrame(formatted_records)

            # Ask user for save location and format
            file_types = [
                ("Excel files", "*.xlsx"),
                ("CSV files", "*.csv"),
                ("Text files", "*.txt")
            ]
            file_path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=file_types,
                title="Export Returns Report"
            )

            if file_path:
                if file_path.endswith('.xlsx'):
                    # Excel export with formatting
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='Returns Report')
                        workbook = writer.book
                        worksheet = writer.sheets['Returns Report']
                        
                        # Auto-adjust column widths
                        for column in worksheet.columns:
                            max_length = 0
                            column = [cell for cell in column]
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(cell.value)
                                except:
                                    pass
                            adjusted_width = (max_length + 2)
                            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

                        # Add some basic formatting
                        header = openpyxl.styles.PatternFill(
                            start_color='1E88E5',
                            end_color='1E88E5',
                            fill_type='solid'
                        )
                        for cell in worksheet[1]:
                            cell.fill = header
                            cell.font = openpyxl.styles.Font(color='FFFFFF', bold=True)

                elif file_path.endswith('.csv'):
                    df.to_csv(file_path, index=False)
                elif file_path.endswith('.txt'):
                    df.to_csv(file_path, index=False, sep='\t')

                messagebox.showinfo("Export Successful", 
                    f"Report exported successfully to:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Export Failed", 
                f"Error exporting data: {str(e)}\n\nPlease make sure you have permission to write to the selected location.")

    # Update the export button
    export_button = ctk.CTkButton(
        status_frame,
        text="Export Report",
        command=export_to_excel,
        width=120,
        height=30,
        fg_color="#27AE60",
        hover_color="#219A52"
    )
    export_button.pack(side="left", padx=20)

    # Add keyboard shortcuts for navigation
    def handle_keypress(event):
        if event.keysym == "Left" and page > 1:
            show_returns_report(
                page=page-1,
                selected_period=selected_period,
                search_term=search_var.get(),
                selected_date=selected_date
            )
        elif event.keysym == "Right" and page < total_pages:
            show_returns_report(
                page=page+1,
                selected_period=selected_period,
                search_term=search_var.get(),
                selected_date=selected_date
            )

    root.bind("<Left>", handle_keypress)
    root.bind("<Right>", handle_keypress)

    # Keep search entry focused
    search_entry.focus_set()
def show_add_transaction():
    clear_widgets(root)

    # Create main frame
    main_frame = ctk.CTkFrame(root)
    main_frame.pack(expand=True, fill="both", padx=20, pady=20)

    # Header
    header_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
    header_frame.pack(fill="x", pady=(0, 20))

    back_icon_image = ctk.CTkImage(Image.open("back.png"), size=(30, 30))
    back_button = ctk.CTkButton(
        header_frame,
        image=back_icon_image,
        text="",
        width=40,
        command=finance,
        fg_color="transparent",
        hover_color="#E5E7E9"
    )
    back_button.pack(side="left")

    ctk.CTkLabel(header_frame, text="Add New Bill", font=("Roboto", 24, "bold"), text_color="#1E3A8A").pack(side="left", padx=20)

    # Supplier Info Frame
    supplier_frame = ctk.CTkFrame(main_frame, fg_color="#F3F4F6", corner_radius=10)
    supplier_frame.pack(pady=5, padx=20, fill="x")

    ctk.CTkLabel(supplier_frame, text="Supplier ID:", font=("Roboto", 14)).grid(row=0, column=0, padx=10, pady=10, sticky="e")
    entry_supplier_id = ctk.CTkEntry(supplier_frame, placeholder_text="Enter ID", width=200)
    entry_supplier_id.grid(row=0, column=1, padx=10, pady=5, sticky="ew")

    supplier_name_label = ctk.CTkLabel(supplier_frame, text="Supplier Name:", font=("Roboto", 14))
    supplier_name_label.grid(row=0, column=2, padx=10, pady=5, sticky="w")

    supplier_name_value = ctk.CTkLabel(supplier_frame, text="", font=("Roboto", 18), text_color="Black",fg_color="#f0f0c9")
    supplier_name_value.grid(row=0, column=3, padx=0, pady=5, sticky="w")

    def get_supplier_name(event=None):
        supplier_id = entry_supplier_id.get()
        if not supplier_id.isdigit():
            supplier_name_value.configure(text="Invalid ID")
            return
        try:
            connection = connect(host="localhost", user="root", password="123@123", database="pharmacy")
            cursor = connection.cursor()
            query = "SELECT name FROM suppliers WHERE supplier_id = %s"
            cursor.execute(query, (supplier_id,))
            result = cursor.fetchone()
            if result:
                supplier_name_value.configure(text=f"{result[0]}")
            else:
                supplier_name_value.configure(text="Not Found")
        except Error as e:
            print(f"Error fetching supplier name: {e}")
            supplier_name_value.configure(text="Error")
        finally:
            if connection.is_connected():
                cursor.close()
                connection.close()

    entry_supplier_id.bind("<KeyRelease>", get_supplier_name)
    def calculate_grand_total_and_tax(event=None):
        try:
            # Calculate the sum of all Net Amounts from the medicine entries
            grand_total = 0.0
            for row_entries in medicine_entries:
                net_amount_entry = row_entries[-1]  # The last column is Net Amount
                net_amount = float(net_amount_entry.get())
                grand_total += net_amount
            
            entry_grand_total.delete(0, 'end')
            entry_grand_total.insert(0, f"{grand_total:.2f}")

            # Now calculate the net bill amount by adding the tax to the grand total
            tax = float(entry_advance_tax.get()) if entry_advance_tax.get() else 0.0
            net_bill_amount = grand_total + tax  # Add the tax to the grand total
            
            entry_total_amount.delete(0, 'end')
            entry_total_amount.insert(0, f"{net_bill_amount:.2f}")

        except ValueError:
            entry_grand_total.delete(0, 'end')
            entry_grand_total.insert(0, "0.00")
            entry_total_amount.delete(0, 'end')
            entry_total_amount.insert(0, "0.00")


    # Medicines Frame
    # Medicines Frame
    medicines_frame = ctk.CTkFrame(main_frame, fg_color="#F3F4F6", corner_radius=10)
    medicines_frame.pack(pady=5, padx=20, fill="x")

    ctk.CTkLabel(medicines_frame, text="Number of Medicines:", font=("Roboto", 14)).pack(side="left", padx=10, pady=10)
    entry_num_medicines = ctk.CTkEntry(medicines_frame, placeholder_text="Enter number", width=100)
    entry_num_medicines.pack(side="left", padx=5, pady=10)

    medicine_entries_frame = ctk.CTkScrollableFrame(main_frame, height=70, fg_color="#FFFFFF", corner_radius=10)
    medicine_entries_frame.pack(pady=5, padx=20, fill="both", expand=True)

    medicine_entries = []

    def calculate_gross_and_net_amount(entry_sale_qty, entry_trade_price, entry_discount, entry_tax, entry_gross_amount, entry_net_amount):
        try:
            sale_quantity = float(entry_sale_qty.get())
            trade_price = float(entry_trade_price.get())
            discount = float(entry_discount.get()) if entry_discount.get() else 0
            tax = float(entry_tax.get()) if entry_tax.get() else 0

            gross_amount = sale_quantity * trade_price
            net_amount = gross_amount - discount + tax

            entry_gross_amount.set(f"{gross_amount:.2f}")
            entry_net_amount.set(f"{net_amount:.2f}")
        except ValueError:
            entry_gross_amount.set("0.00")
            entry_net_amount.set("0.00")

    def create_medicine_entries(event=None):
        for widget in medicine_entries_frame.winfo_children():
            widget.destroy()
        medicine_entries.clear()

        try:
            num_medicines = int(entry_num_medicines.get())
        except ValueError:
            return

        columns = [
            "Medicine Name", "Sale Quantity", "Bonus", "Pack", "Batch",
            "Trade Price", "Gross Amount", "Discount", "Tax", "Net Amount"
        ]

        for idx, col in enumerate(columns):
            ctk.CTkLabel(medicine_entries_frame, text=col, font=("Roboto", 12, "bold"), fg_color="#E5E7EB").grid(row=0, column=idx, padx=5, pady=5, sticky="ew")

        for i in range(num_medicines):
            row_entries = []
            sale_qty_var = tk.StringVar()
            trade_price_var = tk.StringVar()
            gross_amount_var = tk.StringVar(value="0.00")
            discount_var = tk.StringVar()
            tax_var = tk.StringVar()
            net_amount_var = tk.StringVar(value="0.00")

            for j, col in enumerate(columns):
                if col == "Gross Amount":
                    entry = ctk.CTkEntry(medicine_entries_frame, textvariable=gross_amount_var, state="readonly", width=60)
                elif col == "Net Amount":
                    entry = ctk.CTkEntry(medicine_entries_frame, textvariable=net_amount_var, state="readonly", width=60)
                else:
                    entry = ctk.CTkEntry(medicine_entries_frame, width=60)
                    if col == "Sale Quantity":
                        entry.configure(textvariable=sale_qty_var)
                    elif col == "Trade Price":
                        entry.configure(textvariable=trade_price_var)
                    elif col == "Discount":
                        entry.configure(textvariable=discount_var)
                    elif col == "Tax":
                        entry.configure(textvariable=tax_var)

                entry.grid(row=i+1, column=j, padx=5, pady=5, sticky="ew")
                row_entries.append(entry)

            medicine_entries.append(row_entries)

            # Bind events to calculate gross and net amounts when quantity, price, discount, or tax is updated
            sale_qty_var.trace_add("write", lambda *args, sq=sale_qty_var, tp=trade_price_var, d=discount_var, t=tax_var, g=gross_amount_var, n=net_amount_var:
                                   calculate_gross_and_net_amount(sq, tp, d, t, g, n))
            trade_price_var.trace_add("write", lambda *args, sq=sale_qty_var, tp=trade_price_var, d=discount_var, t=tax_var, g=gross_amount_var, n=net_amount_var:
                                      calculate_gross_and_net_amount(sq, tp, d, t, g, n))
            discount_var.trace_add("write", lambda *args, sq=sale_qty_var, tp=trade_price_var, d=discount_var, t=tax_var, g=gross_amount_var, n=net_amount_var:
                                   calculate_gross_and_net_amount(sq, tp, d, t, g, n))
            tax_var.trace_add("write", lambda *args, sq=sale_qty_var, tp=trade_price_var, d=discount_var, t=tax_var, g=gross_amount_var, n=net_amount_var:
                              calculate_gross_and_net_amount(sq, tp, d, t, g, n))
            # Bind events to calculate gross and net amounts when quantity, price, discount, or tax is updated
            sale_qty_var.trace_add("write", lambda *args, sq=sale_qty_var, tp=trade_price_var, d=discount_var, t=tax_var, g=gross_amount_var, n=net_amount_var:
                (calculate_gross_and_net_amount(sq, tp, d, t, g, n), calculate_grand_total_and_tax()))
            trade_price_var.trace_add("write", lambda *args, sq=sale_qty_var, tp=trade_price_var, d=discount_var, t=tax_var, g=gross_amount_var, n=net_amount_var:
                (calculate_gross_and_net_amount(sq, tp, d, t, g, n), calculate_grand_total_and_tax()))
            discount_var.trace_add("write", lambda *args, sq=sale_qty_var, tp=trade_price_var, d=discount_var, t=tax_var, g=gross_amount_var, n=net_amount_var:
                (calculate_gross_and_net_amount(sq, tp, d, t, g, n), calculate_grand_total_and_tax()))
            tax_var.trace_add("write", lambda *args, sq=sale_qty_var, tp=trade_price_var, d=discount_var, t=tax_var, g=gross_amount_var, n=net_amount_var:
                (calculate_gross_and_net_amount(sq, tp, d, t, g, n), calculate_grand_total_and_tax()))

        for j in range(len(columns)):
            medicine_entries_frame.columnconfigure(j, weight=1)

    entry_num_medicines.bind("<KeyRelease>", create_medicine_entries)

    # Remaining frames for transaction date, amounts, and add button (same as original)

    date_frame = ctk.CTkFrame(main_frame, fg_color="#F3F4F6", corner_radius=10)
    date_frame.pack(pady=5, padx=20, fill="x")

    ctk.CTkLabel(date_frame, text="Transaction Date:", font=("Roboto", 14)).pack(side="left", padx=(10, 5), pady=5)
    date_picker = DateEntry(date_frame, width=12, background='#1E3A8A', foreground='white', borderwidth=2)
    date_picker.pack(side="left", padx=(5, 10), pady=10)
    # Amount Frame
    amount_frame = ctk.CTkFrame(main_frame, fg_color="#F3F4F6", corner_radius=10)
    amount_frame.pack(pady=20, padx=20, fill="x")

    ctk.CTkLabel(amount_frame, text="Grand Total:", font=("Roboto", 14)).grid(row=0, column=0, padx=5, pady=10, sticky="e")
    entry_grand_total = ctk.CTkEntry(amount_frame, placeholder_text="0.00", width=170)
    entry_grand_total.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

    ctk.CTkLabel(amount_frame, text="Advance Tax:", font=("Roboto", 14)).grid(row=0, column=2, padx=5, pady=10, sticky="e")
    entry_advance_tax = ctk.CTkEntry(amount_frame, placeholder_text="0.00", width=170)
    entry_advance_tax.grid(row=0, column=3, padx=5, pady=5, sticky="ew")
    entry_advance_tax.bind("<KeyRelease>", calculate_grand_total_and_tax)

    ctk.CTkLabel(amount_frame, text="Net Bill Amount:", font=("Roboto", 14)).grid(row=0, column=4, padx=5, pady=10, sticky="e")
    entry_total_amount = ctk.CTkEntry(amount_frame, placeholder_text="0.00", width=170)
    entry_total_amount.grid(row=0, column=5, padx=5, pady=5, sticky="ew")

    ctk.CTkLabel(amount_frame, text="Paid Amount:", font=("Roboto", 14)).grid(row=0, column=6, padx=5, pady=10, sticky="e")
    entry_paid_amount = ctk.CTkEntry(amount_frame, placeholder_text="0.00", width=170)
    entry_paid_amount.grid(row=0, column=7, padx=5, pady=5, sticky="ew")

    ctk.CTkLabel(amount_frame, text="Remaining:", font=("Roboto", 14)).grid(row=0, column=8, padx=5, pady=10, sticky="e")
    remaining_amount_var = tk.StringVar()
    entry_remaining_amount = ctk.CTkEntry(amount_frame, textvariable=remaining_amount_var, state="readonly", width=170)
    entry_remaining_amount.grid(row=0, column=9, padx=5, pady=5, sticky="ew")

    def calculate_remaining_amount(*args):
        try:
            total = float(entry_total_amount.get())
            paid = float(entry_paid_amount.get())
            remaining = max(0, total - paid)
            remaining_amount_var.set(f"{remaining:.2f}")
        except ValueError:
            remaining_amount_var.set("")

    entry_total_amount.bind("<KeyRelease>", calculate_remaining_amount)
    entry_paid_amount.bind("<KeyRelease>", calculate_remaining_amount)

    def add_transaction():
        supplier_id = entry_supplier_id.get()
        grand_total = entry_grand_total.get()  # Now this value will already be calculated
        advance_tax = entry_advance_tax.get()
        total_amount = entry_total_amount.get()  # The net bill amount
        paid_amount = entry_paid_amount.get()
        remaining_amount = remaining_amount_var.get()
        transaction_date = date_picker.get_date()

        if not all([supplier_id, grand_total, advance_tax, total_amount, paid_amount, remaining_amount]):
            messagebox.showerror("Error", "All fields are required.")
            return

        if not supplier_id.isdigit():
            messagebox.showerror("Error", "Supplier ID must be an integer.")
            return

        try:
            float(grand_total)
            float(advance_tax)
            float(total_amount)
            float(paid_amount)
            float(remaining_amount)
        except ValueError:
            messagebox.showerror("Error", "Amount fields must be valid decimal numbers.")
            return

        medicines = []
        for row_entries in medicine_entries:
            medicine_data = [entry.get() for entry in row_entries]
            if not all(medicine_data):
                messagebox.showerror("Error", "All medicine fields must be filled.")
                return
            medicines.append(medicine_data)

        try:
            connection = connect(host="localhost", user="root", password="123@123", database="pharmacy")
            cursor = connection.cursor()

            # Insert the new transaction
            query = '''
            INSERT INTO transactions (supplier_id, grand_total, advance_tax, total_amount, paid_amount, remaining_amount, transaction_date)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            '''
            cursor.execute(query, (supplier_id, grand_total, advance_tax, total_amount, paid_amount, remaining_amount, transaction_date))
            transaction_id = cursor.lastrowid

            # Insert medicine details
            medicine_query = '''
            INSERT INTO transaction_medicines (transaction_id, medicine_name, sale_quantity, bonus, pack, batch, 
                                            trade_price, gross_amount, discount, tax, net_amount)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            '''
            for medicine in medicines:
                cursor.execute(medicine_query, (transaction_id, *medicine))

            connection.commit()
            messagebox.showinfo("Success", "Transaction added successfully.")
            finance()

        except Error as e:
            print(f"Error during database insertion: {e}")
            messagebox.showerror("Error", f"Failed to add transaction: {e}")

        finally:
            if connection.is_connected():
                cursor.close()
                connection.close()

    add_button = ctk.CTkButton(main_frame, text="Add Transaction", command=add_transaction, font=("Roboto", 14, "bold"), fg_color="#1E3A8A", hover_color="#2563EB")
    add_button.pack(pady=5)
def show_edit_transaction():
    clear_widgets(root)
    
    # Create main frame
    main_frame = ctk.CTkScrollableFrame(root)
    main_frame.pack(expand=True, fill="both", padx=20, pady=20)

    # Header
    header_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
    header_frame.pack(fill="x", pady=(0, 20))

    back_icon_image = ctk.CTkImage(Image.open("back.png"), size=(30, 30))
    back_button = ctk.CTkButton(
        header_frame,
        image=back_icon_image,
        text="",
        width=40,
        command=finance,
        fg_color="transparent",
        hover_color="#E5E7E9"
    )
    back_button.pack(side="left")

    ctk.CTkLabel(header_frame, text="Edit Transaction", font=("Roboto", 24, "bold"), text_color="#1E3A8A").pack(side="left", padx=20)
    
    # Transaction ID Frame
    transaction_id_frame = ctk.CTkFrame(main_frame, fg_color="#F3F4F6", corner_radius=10)
    transaction_id_frame.pack(pady=5, padx=20, fill="x")

    ctk.CTkLabel(transaction_id_frame, text="Transaction ID:", font=("Roboto", 14)).pack(side="left", padx=10, pady=10)
    entry_transaction_id = ctk.CTkEntry(transaction_id_frame, placeholder_text="Enter ID", width=200)
    entry_transaction_id.pack(side="left", padx=5, pady=10)

    fetch_button = ctk.CTkButton(transaction_id_frame, text="Fetch Transaction", command=lambda: fetch_transaction_details(), 
                                 font=("Roboto", 14), fg_color="#1E3A8A", hover_color="#2563EB")
    fetch_button.pack(side="left", padx=10, pady=10)

    # Supplier Info Frame
    supplier_frame = ctk.CTkFrame(main_frame, fg_color="#F3F4F6", corner_radius=10)
    supplier_frame.pack(pady=5, padx=20, fill="x")

    ctk.CTkLabel(supplier_frame, text="Supplier ID:", font=("Roboto", 14)).grid(row=0, column=0, padx=10, pady=10, sticky="e")
    entry_supplier_id = ctk.CTkEntry(supplier_frame, placeholder_text="Enter ID", width=200)
    entry_supplier_id.grid(row=0, column=1, padx=10, pady=5, sticky="ew")

    supplier_name_label = ctk.CTkLabel(supplier_frame, text="Supplier Name: ", font=("Roboto", 14, "bold"), text_color="Green")
    supplier_name_label.grid(row=0, column=2, padx=10, pady=5, sticky="w")

    def get_supplier_name(event=None):
        supplier_id = entry_supplier_id.get()
        if not supplier_id.isdigit():
            supplier_name_label.configure(text="Supplier Name: Invalid ID")
            return
        try:
            connection = connect(host="localhost", user="root", password="123@123", database="pharmacy")
            cursor = connection.cursor()
            query = "SELECT name FROM suppliers WHERE supplier_id = %s"
            cursor.execute(query, (supplier_id,))
            result = cursor.fetchone()
            if result:
                supplier_name_label.configure(text=f"Supplier Name: {result[0]}")
            else:
                supplier_name_label.configure(text="Supplier Name: Not Found")
        except Error as e:
            print(f"Error fetching supplier name: {e}")
            supplier_name_label.configure(text="Supplier Name: Error")
        finally:
            if connection.is_connected():
                cursor.close()
                connection.close()

    entry_supplier_id.bind("<KeyRelease>", get_supplier_name)

    # Medicines Frame
    medicines_frame = ctk.CTkFrame(main_frame, fg_color="#F3F4F6", corner_radius=10)
    medicines_frame.pack(pady=5, padx=20, fill="x")

    ctk.CTkLabel(medicines_frame, text="Number of Medicines:", font=("Roboto", 14)).pack(side="left", padx=10, pady=10)
    entry_num_medicines = ctk.CTkEntry(medicines_frame, placeholder_text="Enter number", width=100)
    entry_num_medicines.pack(side="left", padx=5, pady=10)

    medicine_entries_frame = ctk.CTkScrollableFrame(main_frame, height=200, fg_color="#FFFFFF", corner_radius=10)
    medicine_entries_frame.pack(pady=5, padx=20, fill="both", expand=True)

    medicine_entries = []

    def create_medicine_entries(event=None):
        for widget in medicine_entries_frame.winfo_children():
            widget.destroy()
        medicine_entries.clear()

        try:
            num_medicines = int(entry_num_medicines.get())
        except ValueError:
            return

        columns = [
            "Medicine Name", "Sale Quantity", "Bonus", "Pack", "Batch",
            "Trade Price", "Gross Amount", "Discount", "Tax", "Net Amount"
        ]

        for idx, col in enumerate(columns):
            ctk.CTkLabel(medicine_entries_frame, text=col, font=("Roboto", 12, "bold"), fg_color="#E5E7EB").grid(row=0, column=idx, padx=5, pady=5, sticky="ew")

        for i in range(num_medicines):
            row_entries = []
            for j in range(len(columns)):
                entry = ctk.CTkEntry(medicine_entries_frame, placeholder_text=columns[j], width=100)
                entry.grid(row=i+1, column=j, padx=5, pady=5, sticky="ew")
                row_entries.append(entry)
            medicine_entries.append(row_entries)

        for j in range(len(columns)):
            medicine_entries_frame.columnconfigure(j, weight=1)

    entry_num_medicines.bind("<KeyRelease>", create_medicine_entries)

    # Date Frame
    date_frame = ctk.CTkFrame(main_frame, fg_color="#F3F4F6", corner_radius=10)
    date_frame.pack(pady=5, padx=20, fill="x")

    ctk.CTkLabel(date_frame, text="Transaction Date:", font=("Roboto", 14)).pack(side="left", padx=(10, 5), pady=5)
    date_picker = DateEntry(date_frame, width=12, background='#1E3A8A', foreground='white', borderwidth=2)
    date_picker.pack(side="left", padx=(5, 10), pady=10)

    # Amount Frame
    amount_frame = ctk.CTkFrame(main_frame, fg_color="#F3F4F6", corner_radius=10)
    amount_frame.pack(pady=20, padx=20, fill="x")

    ctk.CTkLabel(amount_frame, text="Grand Total:", font=("Roboto", 14)).grid(row=0, column=0, padx=5, pady=10, sticky="e")
    entry_grand_total = ctk.CTkEntry(amount_frame, placeholder_text="0.00", width=170)
    entry_grand_total.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

    ctk.CTkLabel(amount_frame, text="Advance Tax:", font=("Roboto", 14)).grid(row=0, column=2, padx=5, pady=10, sticky="e")
    entry_advance_tax = ctk.CTkEntry(amount_frame, placeholder_text="0.00", width=170)
    entry_advance_tax.grid(row=0, column=3, padx=5, pady=5, sticky="ew")

    ctk.CTkLabel(amount_frame, text="Net Bill Amount:", font=("Roboto", 14)).grid(row=0, column=4, padx=5, pady=10, sticky="e")
    entry_total_amount = ctk.CTkEntry(amount_frame, placeholder_text="0.00", width=170)
    entry_total_amount.grid(row=0, column=5, padx=5, pady=5, sticky="ew")

    ctk.CTkLabel(amount_frame, text="Paid Amount:", font=("Roboto", 14)).grid(row=0, column=6, padx=5, pady=10, sticky="e")
    entry_paid_amount = ctk.CTkEntry(amount_frame, placeholder_text="0.00", width=170)
    entry_paid_amount.grid(row=0, column=7, padx=5, pady=5, sticky="ew")

    ctk.CTkLabel(amount_frame, text="Remaining:", font=("Roboto", 14)).grid(row=0, column=8, padx=5, pady=10, sticky="e")
    remaining_amount_var = tk.StringVar()
    entry_remaining_amount = ctk.CTkEntry(amount_frame, textvariable=remaining_amount_var, state="readonly", width=170)
    entry_remaining_amount.grid(row=0, column=9, padx=5, pady=5, sticky="ew")

    def calculate_remaining_amount(*args):
        try:
            total = float(entry_total_amount.get())
            paid = float(entry_paid_amount.get())
            remaining = max(0, total - paid)
            remaining_amount_var.set(f"{remaining:.2f}")
        except ValueError:
            remaining_amount_var.set("")

    entry_total_amount.bind("<KeyRelease>", calculate_remaining_amount)
    entry_paid_amount.bind("<KeyRelease>", calculate_remaining_amount)

    def fetch_transaction_details():
        transaction_id = entry_transaction_id.get()
        if not transaction_id.isdigit():
            messagebox.showerror("Error", "Invalid Transaction ID")
            return

        try:
            connection = connect(host="localhost", user="root", password="123@123", database="pharmacy")
            cursor = connection.cursor()

            # Fetch transaction details
            query = """
            SELECT t.supplier_id, t.grand_total, t.advance_tax, t.total_amount, t.paid_amount, t.remaining_amount, t.transaction_date, s.name
            FROM transactions t
            JOIN suppliers s ON t.supplier_id = s.supplier_id
            WHERE t.transaction_id = %s
            """
            cursor.execute(query, (transaction_id,))
            result = cursor.fetchone()

            if result:
                supplier_id, grand_total, advance_tax, total, paid, remaining, trans_date, supplier_name = result
                entry_supplier_id.delete(0, tk.END)
                entry_supplier_id.insert(0, supplier_id)
                supplier_name_label.configure(text=f"Supplier Name: {supplier_name}")
                entry_grand_total.delete(0, tk.END)
                entry_grand_total.insert(0, grand_total)
                entry_advance_tax.delete(0, tk.END)
                entry_advance_tax.insert(0, advance_tax)
                entry_total_amount.delete(0, tk.END)
                entry_total_amount.insert(0, total)
                entry_paid_amount.delete(0, tk.END)
                entry_paid_amount.insert(0, paid)
                remaining_amount_var.set(str(remaining))
                date_picker.set_date(trans_date)

                # Fetch medicine details
                query = """
                SELECT medicine_name, sale_quantity, bonus, pack, batch, 
                       trade_price, gross_amount, discount, tax, net_amount
                FROM transaction_medicines
                WHERE transaction_id = %s
                """
                cursor.execute(query, (transaction_id,))
                medicines = cursor.fetchall()

                entry_num_medicines.delete(0, tk.END)
                entry_num_medicines.insert(0, str(len(medicines)))
                create_medicine_entries()

                for i, medicine in enumerate(medicines):
                    for j, value in enumerate(medicine):
                        medicine_entries[i][j].delete(0, tk.END)
                        medicine_entries[i][j].insert(0, str(value))
            else:
                messagebox.showerror("Error", "Transaction not found")

        except Error as e:
            print(f"Error fetching transaction details: {e}")
            messagebox.showerror("Error", "Failed to fetch transaction details")
        finally:
            if connection.is_connected():
                cursor.close()
                connection.close()

    def update_transaction():
        transaction_id = entry_transaction_id.get()
        supplier_id = entry_supplier_id.get()
        grand_total = entry_grand_total.get()
        advance_tax = entry_advance_tax.get()
        total_amount = entry_total_amount.get()
        paid_amount = entry_paid_amount.get()
        remaining_amount = remaining_amount_var.get()
        transaction_date = date_picker.get_date()

        if not all([transaction_id, supplier_id, grand_total, advance_tax, total_amount, paid_amount, remaining_amount]):
            messagebox.showerror("Error", "All fields are required.")
            return

        try:
            connection = connect(host="localhost", user="root", password="123@123", database="pharmacy")
            cursor = connection.cursor()

            # Update the transaction
            query = '''
            UPDATE transactions 
            SET supplier_id = %s, grand_total = %s, advance_tax = %s, total_amount = %s, paid_amount = %s, remaining_amount = %s, transaction_date = %s
            WHERE transaction_id = %s
            '''
            cursor.execute(query, (supplier_id, grand_total, advance_tax, total_amount, paid_amount, remaining_amount, transaction_date, transaction_id))

            # Delete existing medicine entries
            cursor.execute("DELETE FROM transaction_medicines WHERE transaction_id = %s", (transaction_id,))

            # Insert updated medicine details
            medicine_query = '''
            INSERT INTO transaction_medicines (transaction_id, medicine_name, sale_quantity, bonus, pack, batch, 
                                               trade_price, gross_amount, discount, tax, net_amount)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            '''
            for row_entries in medicine_entries:
                medicine_data = [entry.get() for entry in row_entries]
                if all(medicine_data):
                    cursor.execute(medicine_query, (transaction_id, *medicine_data))

            connection.commit()
            messagebox.showinfo("Success", "Transaction updated successfully.")
            finance()

        except Error as e:
            print(f"Error during database update: {e}")
            messagebox.showerror("Error", f"Failed to update transaction: {e}")

        finally:
            if connection.is_connected():
                cursor.close()
                connection.close()
    def delete_transaction():
        # Create a new top-level window for the delete prompt
        delete_window = tk.Toplevel()
        delete_window.title("Delete Transaction")
        delete_window.geometry("300x150")

        tk.Label(delete_window, text="Enter Transaction ID to Delete:", font=("Roboto", 12)).pack(pady=10)

        entry_delete_transaction_id = tk.Entry(delete_window, width=25)
        entry_delete_transaction_id.pack(pady=5)

        def confirm_delete():
            transaction_id = entry_delete_transaction_id.get()
            if not transaction_id.isdigit():
                messagebox.showerror("Error", "Invalid Transaction ID")
                return

            confirm = messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete transaction {transaction_id}?")
            if not confirm:
                return

            try:
                connection = connect(host="localhost", user="root", password="123@123", database="pharmacy")
                cursor = connection.cursor()

                # Delete related medicine details first to avoid foreign key constraint issues
                cursor.execute("DELETE FROM transaction_medicines WHERE transaction_id = %s", (transaction_id,))

                # Delete the transaction record
                cursor.execute("DELETE FROM transactions WHERE transaction_id = %s", (transaction_id,))

                connection.commit()
                messagebox.showinfo("Success", f"Transaction {transaction_id} deleted successfully.")
                delete_window.destroy()  # Close the delete window after successful deletion
                finance()  # Redirect to the finance screen after deletion

            except Error as e:
                print(f"Error during transaction deletion: {e}")
                messagebox.showerror("Error", f"Failed to delete transaction: {e}")
            finally:
                if connection.is_connected():
                    cursor.close()
                    connection.close()

        # Button to confirm delete
        delete_button = tk.Button(delete_window, text="Delete", command=confirm_delete, bg="#D32F2F", fg="white")
        delete_button.pack(pady=10)

        # Button to cancel and close the window
        cancel_button = tk.Button(delete_window, text="Cancel", command=delete_window.destroy, bg="#E5E7E9")
        cancel_button.pack()


    update_button = ctk.CTkButton(main_frame, text="Update Transaction", command=update_transaction, 
                                  font=("Roboto", 14, "bold"), fg_color="#1E3A8A", hover_color="#2563EB")
    update_button.pack(pady=10)

    # Button to delete transaction (which opens a message box with an entry field)
    delete_button = ctk.CTkButton(main_frame, text="Delete Transaction", command=delete_transaction, 
                                  font=("Roboto", 14, "bold"), fg_color="#D32F2F", hover_color="#B71C1C")
    delete_button.pack(pady=10)


def fetch_filtered_transactions(supplier_id=None, start_date=None, end_date=None):
    """Fetch filtered transactions including medicine details."""
    query = """
    SELECT t.transaction_id, s.supplier_id, s.name AS supplier_name, 
           tm.medicine_name, tm.sale_quantity, tm.bonus, tm.trade_price, tm.gross_amount, 
           tm.discount, tm.net_amount, t.total_amount, t.paid_amount, t.remaining_amount, 
           t.transaction_date, t.grand_total, t.advance_tax
    FROM transactions t
    JOIN suppliers s ON t.supplier_id = s.supplier_id
    JOIN transaction_medicines tm ON t.transaction_id = tm.transaction_id
    WHERE 1=1
    """
    params = []

    if supplier_id:
        query += " AND t.supplier_id = %s"
        params.append(supplier_id)

    if start_date and end_date:
        query += " AND t.transaction_date BETWEEN %s AND %s"
        params.extend([start_date, end_date])

    query += " ORDER BY t.transaction_date DESC"

    # Fetch data and print for debugging
    result = fetch_data(query, tuple(params))
    print(result)  # Add this line to debug the fetched data
    return result

def get_db_connection():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="123@123",
        database="pharmacy"
    )

def fetch_medicine_details(transaction_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    query = """
    SELECT medicine_name, sale_quantity, bonus, pack, batch, trade_price, 
           gross_amount, discount, tax, net_amount
    FROM transaction_medicines
    WHERE transaction_id = %s
    """
    cursor.execute(query, (transaction_id,))
    medicines = cursor.fetchall()
    cursor.close()
    conn.close()
    return medicines

def show_transactions_report(start_row=2, selected_supplier=None, selected_date=None):
    clear_widgets(root)
    reset_grid_configuration(root)

    def fetch_suppliers():
        """Fetch all suppliers from the database."""
        query = "SELECT supplier_id, name FROM suppliers ORDER BY name"
        suppliers = fetch_data(query)
        if suppliers is None:
            return []
        return suppliers

    def show_calendar_popup():
        """Displays a calendar popup for selecting date range."""
        calendar_window = ctk.CTkToplevel(root)
        calendar_window.title("Select Date Range")
        calendar_window.geometry("300x300")
        calendar_window.transient(root)
        calendar_window.grab_set()
        # Start Date Calendar
        start_date_label = ctk.CTkLabel(calendar_window, text="From:", font=("Arial", 12, "bold"))
        start_date_label.pack(pady=(10, 0))

        start_cal = DateEntry(calendar_window, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        start_cal.pack(pady=(5, 10))

        # End Date Calendar
        end_date_label = ctk.CTkLabel(calendar_window, text="End Date:", font=("Arial", 12, "bold"))
        end_date_label.pack(pady=(10, 0))

        end_cal = DateEntry(calendar_window, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        end_cal.pack(pady=(5, 10))

        # Select button to confirm the selected date range
        def on_select_dates():
            start_date = start_cal.get()
            end_date = end_cal.get()
            calendar_window.destroy()  # Close the calendar pop-up
            show_transactions_report(selected_supplier=selected_supplier, selected_date=(start_date, end_date))  # Update report with date range

        select_button = ctk.CTkButton(calendar_window, text="Select Dates", command=on_select_dates)
        select_button.pack(pady=20)


    for col in range(12):
        root.grid_columnconfigure(col, weight=1)

    title_label = ctk.CTkLabel(
        root, 
        text="Supplier Transactions Report", 
        font=("Arial", 24, "bold"),
        text_color="#2E86C1"
    )
    title_label.grid(row=0, column=0, columnspan=12, pady=20, sticky="nsew")

    back_icon_image = ctk.CTkImage(Image.open("back.png"), size=(30, 30))
    
    # Search frame setup
    search_frame = ctk.CTkFrame(root, fg_color="transparent")
    search_frame.grid(row=0, column=0, padx=20, pady=10, sticky="nw")
    
    # Back button and search entry
    back_button = ctk.CTkButton(
        search_frame,
        image=back_icon_image,
        text="",
        width=40,
        command=finance,
        fg_color="transparent",
        hover_color="#E5E7E9"
    )
    back_button.grid(row=0, column=0, padx=(0, 5), pady=0, sticky="nw")

    date_button = ctk.CTkButton(
        root,
        text="Select Date",
        command=show_calendar_popup,
        width=150,
        height=32,
        fg_color="#2E86C1",
        hover_color="#3498DB"
    )
    date_button.grid(row=0, column=11, padx=20, pady=10, sticky="ne")

    suppliers = fetch_suppliers()
    supplier_options = [f"{sid}: {name}" for sid, name, *_ in suppliers]

    supplier_combobox = ctk.CTkComboBox(
        root,
        values=supplier_options,
        command=lambda choice: show_transactions_report(selected_supplier=int(choice.split(':')[0]), selected_date=selected_date),
        width=200,
        height=32,
        border_color="#2E86C1",
        button_color="#2E86C1",
        dropdown_hover_color="#3498DB"
    )
    supplier_combobox.set("Select Supplier")
    supplier_combobox.grid(row=0, column=10, padx=20, pady=10, sticky="ne")

    headers = [
            ("", 5),  # For dropdown button
            ("Transaction ID", 10),
            ("Supplier Name", 20),
            ("Grand Total", 15),
            ("Paid Amount", 15),
            ("Remaining", 15),
            ("Date", 20)
        ]
    header_frame = ctk.CTkFrame(root, fg_color="transparent")
    header_frame.grid(row=1, column=0, columnspan=12, sticky="nsew", padx=10)
    
    # Configure column weights in header frame
    for col, (_, weight) in enumerate(headers):
        header_frame.grid_columnconfigure(col, weight=weight)
        header_label = ctk.CTkLabel(
            header_frame,
            text=headers[col][0],
            font=("Arial", 14, "bold"),
            text_color="#FFFFFF",
            fg_color="#2E86C1",
            corner_radius=8,
            height=35,
            anchor="center"
        )
        header_label.grid(row=0, column=col, padx=2, pady=10, sticky="nsew")

    scroll_frame = ctk.CTkScrollableFrame(
        root,
        height=600,
        width=1200,
        fg_color="#F8F9F9",
        corner_radius=10
    )
    scroll_frame.grid(row=2, column=0, columnspan=12, padx=10, pady=10, sticky="nsew")

    # Configure column weights in scroll frame
    for col, (_, weight) in enumerate(headers):
        scroll_frame.grid_columnconfigure(col, weight=weight)

    try:
        transactions_data = fetch_filtered_transactions(selected_supplier, *selected_date if selected_date else (None, None))

        if transactions_data:
            # Clear existing widgets and display new transaction data
            for widget in scroll_frame.winfo_children():
                widget.destroy()

            # Use a set to keep track of displayed transaction IDs
            displayed_transactions = set()

            for row_index, transaction in enumerate(transactions_data, start=1):
                transaction_id = transaction[0]
                
                # Check if this transaction has already been displayed
                if transaction_id in displayed_transactions:
                    continue
                
                displayed_transactions.add(transaction_id)

                bg_color = "#d3f3f1" if row_index % 2 == 0 else "#FFFFFF"
                formatted_date = transaction[13].strftime('%Y-%m-%d') if isinstance(transaction[13], date) else "N/A"

                # Create a frame for each transaction
                transaction_frame = ctk.CTkFrame(scroll_frame, fg_color=bg_color)
                transaction_frame.grid(row=row_index, column=0, columnspan=len(headers), sticky="nsew", padx=2, pady=2)

                # Configure column weights in transaction frame to match headers
                for col, (_, weight) in enumerate(headers):
                    transaction_frame.grid_columnconfigure(col, weight=weight)

                # Add dropdown button
                dropdown_button = ctk.CTkButton(
                    transaction_frame,
                    text="▼",
                    width=30,
                    command=lambda tf=transaction_frame, tid=transaction_id: toggle_dropdown(tf, tid),
                    fg_color="#86bbd8",
                    hover_color="#33658a"
                )
                dropdown_button.grid(row=0, column=0, padx=5, pady=5, sticky="w")

                # Add transaction details with proper alignment
                fields = [
                    (str(transaction_id), "center"),  # Transaction ID
                    (transaction[2], "center"),  # Supplier Name
                    (f"Rs{transaction[14]:.2f}", "center"),  # Grand Total
                    (f"Rs{transaction[10]:.2f}", "center"),  # Paid Amount
                    (f"Rs{transaction[11]:.2f}", "center"),  # Remaining Amount
                    (formatted_date, "center")  # Date
                ]

                for col, (field, alignment) in enumerate(fields):
                    label = ctk.CTkLabel(
                        transaction_frame,
                        text=str(field),
                        font=("Arial", 13),
                        text_color="#34495E",
                        fg_color="#ebf5df",
                        anchor=alignment,
                        height=30
                    )
                    label.grid(row=0, column=col+1, padx=2, pady=5, sticky="nsew")
                    if "Rs" in str(field):  # Add extra padding for monetary values
                        label.grid(padx=(20, 2))  # More left padding for currency values

        else:
            ctk.CTkLabel(
                scroll_frame,
                text="No transactions found for the selected criteria.",
                font=("Arial", 13, "bold"),
                text_color="#E74C3C",
                anchor="center",
                height=40
            ).grid(row=0, column=0, columnspan=len(headers), pady=20, sticky="nsew")

    except Exception as e:
        print(f"Error in show_transactions_report: {e}")
        ctk.CTkLabel(
            scroll_frame,
            text=f"Error fetching transaction data: {e}",
            font=("Arial", 13, "bold"),
            text_color="#E74C3C",
            anchor="center",
            height=40
        ).grid(row=0, column=0, columnspan=len(headers), pady=20, sticky="nsew")
def toggle_dropdown(transaction_frame, transaction_id):
    headers_length = 7  # Number of columns in the main transaction view

    medicine_frame = transaction_frame.grid_slaves(row=1)[0] if transaction_frame.grid_slaves(row=1) else None
    dropdown_button = transaction_frame.grid_slaves(row=0, column=0)[0]
    
    if medicine_frame:
        if medicine_frame.winfo_viewable():
            medicine_frame.grid_remove()
            dropdown_button.configure(text="▼")
        else:
            medicine_frame.grid()
            dropdown_button.configure(text="▲")
    else:
        # Fetch and display medicine details
        medicines = fetch_medicine_details(transaction_id)
        if medicines:
            medicine_frame = ctk.CTkFrame(transaction_frame, fg_color=transaction_frame.cget("fg_color"))
            medicine_frame.grid(row=1, column=0, columnspan=headers_length, sticky="nsew", padx=2, pady=2)

            # Configure column weights for proper alignment
            for i in range(10):  # 10 columns for medicine details
                medicine_frame.grid_columnconfigure(i, weight=1)

            # Define column widths (adjust these values as needed)
            column_widths = {
                "Medicine Name": 200,
                "Quantity": 80,
                "Bonus": 80,
                "Pack": 80,
                "Batch": 100,
                "Trade Price": 100,
                "Gross Amount": 120,
                "Discount": 100,
                "Tax": 80,
                "Net Amount": 120
            }

            # Add medicine headers
            medicine_headers = list(column_widths.keys())
            for col, header in enumerate(medicine_headers):
                header_label = ctk.CTkLabel(
                    medicine_frame,
                    text=header,
                    font=("Arial", 12, "bold"),
                    text_color="#34495E",
                    fg_color="#ebf4f5",
                    width=column_widths[header],
                    anchor="center"  # Center-align headers
                )
                header_label.grid(row=0, column=col, padx=5, pady=2, sticky="ew")

            # Add medicine details with alignment
            for row, medicine in enumerate(medicines, start=1):
                for col, (key, value) in enumerate(medicine.items()):
                    # Format numeric values
                    if isinstance(value, (int, float)):
                        value = f"{value:,.2f}" if isinstance(value, float) else f"{value:,}"
                    
                    # Determine text alignment based on column type
                    if col == 0:  # Medicine Name
                        anchor = "center"  # Left-align text
                    elif isinstance(medicine[key], (int, float)):  # Numeric columns
                        anchor = "center"  # Right-align numbers
                        value = f"Rs{value}" if "Amount" in medicine_headers[col] or "Price" in medicine_headers[col] else value
                    else:
                        anchor = "center"  # Center-align other content
                    
                    detail_label = ctk.CTkLabel(
                        medicine_frame,
                        text=str(value),
                        text_color="#34495E",
                        fg_color="transparent",
                        width=column_widths[medicine_headers[col]],
                        anchor=anchor
                    )
                    detail_label.grid(row=row, column=col, padx=5, pady=2, sticky="ew")

            dropdown_button.configure(text="▲")
        else:
            print("No medicine details available for this transaction.")
def calculate_and_insert_consolidated_daily_profit():
    connection = mysql.connector.connect(
        host="localhost",
        user="root",
        password="123@123",
        database="pharmacy"  
    )
    
    cursor = connection.cursor()
    
    try:
        # Get the current date
        current_date = datetime.now().date()
        
        # First, check if an entry for today already exists
        check_query = "SELECT profit_id FROM profit WHERE DATE(profit_date) = %s"
        cursor.execute(check_query, (current_date,))
        existing_entry = cursor.fetchone()
        
        if existing_entry:
            # Update existing entry
            update_query = """
            UPDATE profit SET
                total_sales = (
                    SELECT COALESCE(SUM(sales.total_price), 0)
                    FROM sales
                    WHERE DATE(sales.sale_date) = %s
                ),
                total_cost = (
                    SELECT COALESCE(SUM(medicines.buying_price * sales.quantity_sold), 0)
                    FROM sales
                    JOIN medicines ON sales.medicine_id = medicines.medicine_id
                    WHERE DATE(sales.sale_date) = %s
                ),
                profit_amount = total_sales - total_cost
            WHERE DATE(profit_date) = %s
            """
            cursor.execute(update_query, (current_date, current_date, current_date))
            
        else:
            # Insert new entry
            insert_query = """
            INSERT INTO profit (total_sales, total_cost, profit_amount, profit_date)
            SELECT 
                COALESCE(SUM(sales.total_price), 0) AS total_sales,
                COALESCE(SUM(medicines.buying_price * sales.quantity_sold), 0) AS total_cost,
                COALESCE(SUM(sales.total_price) - SUM(medicines.buying_price * sales.quantity_sold), 0) AS profit_amount,
                %s AS profit_date
            FROM sales
            JOIN medicines ON sales.medicine_id = medicines.medicine_id
            WHERE DATE(sales.sale_date) = %s
            """
            cursor.execute(insert_query, (current_date, current_date))
        
        connection.commit()
    
    except mysql.connector.Error as err:
        print(f"Error: {err}")
        connection.rollback()
    
    finally:
        cursor.close()
        connection.close()
def fetch_filtered_profit(selected_period="Lifetime", start_date=None, end_date=None):
    try:
        conn = mysql.connector.connect(
            host="localhost",
            user="root",
            password="123@123",
            database="pharmacy"
        )
        cursor = conn.cursor(prepared=True)

        query = """
        SELECT profit_id, total_sales, total_cost, profit_amount, DATE(profit_date) as profit_date
        FROM profit
        WHERE 1=1
        """
        params = []

        # Add date filtering based on selected period
        if selected_period != "Lifetime":
            end_date = datetime.now().date()
            if selected_period == "Today":
                start_date = end_date
            elif selected_period == "Last Week":
                start_date = end_date - timedelta(days=7)
            elif selected_period == "Last Month":
                start_date = end_date - timedelta(days=30)
            elif selected_period == "6 Months":
                start_date = end_date - timedelta(days=180)
            elif selected_period == "1 Year":
                start_date = end_date - timedelta(days=365)
            elif selected_period == "Other" and start_date and end_date:
                if isinstance(start_date, str):
                    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                if isinstance(end_date, str):
                    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            
            query += " AND DATE(profit_date) BETWEEN %s AND %s"
            params.extend([start_date, end_date])

        # Order by profit date, most recent first
        query += " ORDER BY profit_date DESC"

        print(f"Query: {query}")  # Debug print
        print(f"Params: {params}")  # Debug print

        cursor.execute(query, params)
        profits = cursor.fetchall()
        
        print(f"Fetched profits: {profits}")  # Debug print
        return profits

    except mysql.connector.Error as err:
        print(f"Database error in fetch_filtered_profit: {err}")
        return []

    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()
def show_profit_report(page=1, selected_period="Lifetime", search_term="", selected_date=None):
    clear_widgets(root)
    reset_grid_configuration(root)
    calculate_and_insert_consolidated_daily_profit()
    
    ITEMS_PER_PAGE = 15  # Number of items per page
    
    # Configure column weights for proper alignment
    for col in range(12):
        root.grid_columnconfigure(col, weight=1)

    # Title Section
    title_label = ctk.CTkLabel(
        root, 
        text="Profit Report", 
        font=("Arial", 24, "bold"),
        text_color="#2E86C1"
    )
    title_label.grid(row=0, column=0, columnspan=12, pady=20, sticky="nsew")

    # Back button section
    back_icon_image = ctk.CTkImage(Image.open("back.png"), size=(30, 30))
    
    # Back frame setup
    back_frame = ctk.CTkFrame(root, fg_color="transparent")
    back_frame.grid(row=0, column=0, padx=20, pady=10, sticky="nw")
    
    back_button = ctk.CTkButton(
        back_frame,
        image=back_icon_image,
        text="",
        width=40,
        command=show_profit_management,
        fg_color="transparent",
        hover_color="#E5E7E9"
    )
    back_button.grid(row=0, column=0, padx=(0, 5), pady=0, sticky="nw")

    def show_calendar_popup():
        calendar_window = tk.Toplevel(root)
        calendar_window.title("Select Date Range")
        calendar_window.geometry("300x300")

        start_date_label = ctk.CTkLabel(calendar_window, text="From:", font=("Arial", 12, "bold"))
        start_date_label.pack(pady=(10, 0))

        start_cal = DateEntry(calendar_window, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        start_cal.pack(pady=(5, 10))

        end_date_label = ctk.CTkLabel(calendar_window, text="End Date:", font=("Arial", 12, "bold"))
        end_date_label.pack(pady=(10, 0))

        end_cal = DateEntry(calendar_window, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        end_cal.pack(pady=(5, 10))

        def on_select_dates():
            start_date = start_cal.get_date()
            end_date = end_cal.get_date()
            calendar_window.destroy()
            show_profit_report(page=1, selected_period="Other", selected_date=(start_date, end_date))

        select_button = ctk.CTkButton(calendar_window, text="Select Dates", command=on_select_dates)
        select_button.pack(pady=20)

    def toggle_calendar(choice):
        if choice == "Other":
            show_calendar_popup()
        else:
            show_profit_report(page=1, selected_period=choice)

    # Period combobox
    period_options = ["Today", "Last Week", "Last Month", "6 Months", "1 Year", "Lifetime", "Other"]
    period_combobox = ctk.CTkComboBox(
        root,
        values=period_options,
        command=toggle_calendar,
        width=150,
        height=32,
        border_color="#2E86C1",
        button_color="#2E86C1",
        dropdown_hover_color="#3498DB"
    )
    period_combobox.set(selected_period)
    period_combobox.grid(row=0, column=11, padx=20, pady=10, sticky="ne")

    # Header Section
    headers_profit = [
        ("Profit ID", 15, "center"),
        ("Total Sales", 20, "center"),
        ("Total Cost", 20, "center"),
        ("Profit Amount", 20, "center"),
        ("Date", 15, "center")
    ]

    # Create header frame
    header_frame = ctk.CTkFrame(root, fg_color="transparent")
    header_frame.grid(row=1, column=0, columnspan=12, sticky="nsew", padx=10)
    
    # Configure header columns
    total_weight = sum(weight for _, weight, _ in headers_profit)
    for col, (_, weight, _) in enumerate(headers_profit):
        header_frame.grid_columnconfigure(col, weight=weight)

    # Create headers
    for col, (header, _, _) in enumerate(headers_profit):
        header_label = ctk.CTkLabel(
            header_frame,
            text=header,
            font=("Arial", 14, "bold"),
            text_color="#FFFFFF",
            fg_color="#2E86C1",
            corner_radius=8,
            height=35,
            anchor="center"
        )
        header_label.grid(row=0, column=col, padx=2, pady=10, sticky="nsew")

    # Data frame
    data_frame = ctk.CTkFrame(
        root,
        fg_color="#F8F9F9",
        corner_radius=10
    )
    data_frame.grid(row=2, column=0, columnspan=12, padx=10, pady=10, sticky="nsew")

    # Configure data frame columns
    for col, (_, weight, _) in enumerate(headers_profit):
        data_frame.grid_columnconfigure(col, weight=weight)

    try:
        # Calculate offset for pagination
        offset = (page - 1) * ITEMS_PER_PAGE

        if selected_period == "Other" and selected_date:
            start_date, end_date = selected_date
            profit_data = fetch_filtered_profit("Other", start_date, end_date)
        else:
            profit_data = fetch_filtered_profit(selected_period)

        # Implement pagination
        total_records = len(profit_data)
        total_pages = ceil(total_records / ITEMS_PER_PAGE)
        paginated_data = profit_data[offset:offset + ITEMS_PER_PAGE]

        # Update total entries label
        total_entries_label = ctk.CTkLabel(
            root,
            text=f"Total Entries: {total_records}",
            font=("Arial", 14, "bold"),
            text_color="#2E86C1"
        )
        total_entries_label.grid(row=0, column=10, padx=(0, 10), pady=10, sticky="ne")

        if paginated_data:
            for i, profit_item in enumerate(paginated_data):
                if isinstance(profit_item, tuple) and len(profit_item) == 6:
                    profit_id, total_sales, total_cost, profit_amount, unused, profit_date = profit_item

                    # Format profit date
                    profit_date = profit_date.strftime('%Y-%m-%d') if profit_date else "N/A"

                    # Format monetary values
                    sales_fmt = f"Rs{float(total_sales):,.2f}"
                    cost_fmt = f"Rs{float(total_cost):,.2f}"
                    profit_fmt = f"Rs{float(profit_amount):,.2f}"

                    fields = [
                        (str(profit_id), "center"),
                        (sales_fmt, "center"),
                        (cost_fmt, "center"),
                        (profit_fmt, "center"),
                        (profit_date, "center")
                    ]

                    for col, (field, align) in enumerate(fields):
                        profit_label = ctk.CTkLabel(
                            data_frame,
                            text=field,
                            font=("Arial", 13),
                            text_color="#34495E",
                            fg_color="#FFFFFF" if i % 2 == 0 else "#F7F9F9",
                            corner_radius=8,
                            anchor=align,
                            height=30
                        )
                        padx = (10, 5) if align == "w" else (5, 10) if align == "e" else (5, 5)
                        profit_label.grid(
                            row=i,
                            column=col,
                            padx=padx,
                            pady=2,
                            sticky="nsew"
                        )
        else:
            ctk.CTkLabel(
                data_frame,
                text="No profit entries found for the selected period.",
                font=("Arial", 13, "bold"),
                text_color="#E74C3C",
                anchor="center",
                height=40
            ).grid(row=0, column=0, columnspan=len(headers_profit), pady=20, sticky="nsew")

        # Pagination frame
        pagination_frame = ctk.CTkFrame(root, fg_color="transparent")
        pagination_frame.grid(row=3, column=0, columnspan=12, pady=10, sticky="nsew")

        def change_page(new_page):
            show_profit_report(
                page=new_page,
                selected_period=selected_period,
                selected_date=selected_date
            )

        # Previous page button
        prev_button = ctk.CTkButton(
            pagination_frame,
            text="←",
            width=40,
            command=lambda: change_page(page - 1) if page > 1 else None,
            fg_color="#2E86C1" if page > 1 else "#B3B6B7",
            state="normal" if page > 1 else "disabled"
        )
        prev_button.grid(row=0, column=0, padx=5)

        # Page numbers
        page_label = ctk.CTkLabel(
            pagination_frame,
            text=f"Page {page} of {total_pages}",
            font=("Arial", 12, "bold")
        )
        page_label.grid(row=0, column=1, padx=10)

        # Next page button
        next_button = ctk.CTkButton(
            pagination_frame,
            text="→",
            width=40,
            command=lambda: change_page(page + 1) if page < total_pages else None,
            fg_color="#2E86C1" if page < total_pages else "#B3B6B7",
            state="normal" if page < total_pages else "disabled"
        )
        next_button.grid(row=0, column=2, padx=5)

    except Exception as e:
        error_label = ctk.CTkLabel(
            data_frame,
            text=f"Error fetching profit data: {e}",
            font=("Arial", 13, "bold"),
            text_color="#E74C3C",
            anchor="center",
            height=40
        )
        error_label.grid(row=0, column=0, columnspan=len(headers_profit), pady=20, sticky="nsew")

        # Empty pagination frame in case of error
        pagination_frame = ctk.CTkFrame(root, fg_color="transparent")
        pagination_frame.grid(row=3, column=0, columnspan=12, pady=10, sticky="nsew")

    # Add a status bar showing records per page
    status_frame = ctk.CTkFrame(root, fg_color="transparent")
    status_frame.grid(row=4, column=0, columnspan=12, pady=(5, 10), sticky="nsew")

    records_label = ctk.CTkLabel(
        status_frame,
        text=f"Showing {ITEMS_PER_PAGE} records per page",
        font=("Arial", 12),
        text_color="#7F8C8D"
    )
    records_label.pack(side="right", padx=20)

    # Add export button
    def export_to_excel():
        try:
            # Check if required modules are installed
            try:
                import pandas as pd
                import openpyxl
            except ImportError as e:
                if 'openpyxl' in str(e):
                    if messagebox.askyesno("Missing Dependency", 
                        "The openpyxl module is required for Excel export. Would you like to install it now?"):
                        import subprocess
                        try:
                            subprocess.check_call(['pip', 'install', 'openpyxl'])
                            messagebox.showinfo("Success", "openpyxl installed successfully. Please try exporting again.")
                        except subprocess.CalledProcessError:
                            messagebox.showerror("Installation Failed", 
                                "Could not install openpyxl. Please install it manually using 'pip install openpyxl'")
                    return
                else:
                    messagebox.showerror("Missing Dependency", f"Required module not found: {str(e)}")
                    return

            # Fetch all records for export
            if selected_period == "Other" and selected_date:
                profit_data = fetch_filtered_profit("Other", selected_date[0], selected_date[1])
            else:
                profit_data = fetch_filtered_profit(selected_period)

            if not profit_data:
                messagebox.showwarning("Export Failed", "No records to export.")
                return

            # Convert records to list of dicts
            formatted_records = []
            for record in profit_data:
                profit_id, total_sales, total_cost, profit_amount, unused, profit_date = record
                formatted_records.append({
                    'Profit ID': profit_id,
                    'Total Sales': f"Rs{float(total_sales):,.2f}",
                    'Total Cost': f"Rs{float(total_cost):,.2f}",
                    'Profit Amount': f"Rs{float(profit_amount):,.2f}",
                    'Date': profit_date.strftime('%Y-%m-%d') if profit_date else "N/A"
                })

            # Create DataFrame
            df = pd.DataFrame(formatted_records)

            # Ask user for save location and format
            file_types = [
                ("Excel files", "*.xlsx"),
                ("CSV files", "*.csv"),
                ("Text files", "*.txt")
            ]
            file_path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=file_types,
                title="Export Profit Report"
            )

            if file_path:
                if file_path.endswith('.xlsx'):
                    # Excel export with formatting
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='Profit Report')
                        workbook = writer.book
                        worksheet = writer.sheets['Profit Report']
                        
                        # Auto-adjust column widths
                        for column in worksheet.columns:
                            max_length = 0
                            column = [cell for cell in column]
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(cell.value)
                                except:
                                    pass
                            adjusted_width = (max_length + 2)
                            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

                        # Add formatting
                        header = openpyxl.styles.PatternFill(
                            start_color='1E88E5',
                            end_color='1E88E5',
                            fill_type='solid'
                        )
                        for cell in worksheet[1]:
                            cell.fill = header
                            cell.font = openpyxl.styles.Font(color='FFFFFF', bold=True)

                elif file_path.endswith('.csv'):
                    df.to_csv(file_path, index=False)
                elif file_path.endswith('.txt'):
                    df.to_csv(file_path, index=False, sep='\t')

                messagebox.showinfo("Export Successful", 
                    f"Report exported successfully to:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Export Failed", 
                f"Error exporting data: {str(e)}\n\nPlease make sure you have permission to write to the selected location.")

    # Add export button
    export_button = ctk.CTkButton(
        status_frame,
        text="Export Report",
        command=export_to_excel,
        width=120,
        height=30,
        fg_color="#27AE60",
        hover_color="#219A52"
    )
    export_button.pack(side="left", padx=20)

    # Add keyboard shortcuts for navigation
    def handle_keypress(event):
        if event.keysym == "Left" and page > 1:
            show_profit_report(
                page=page-1,
                selected_period=selected_period,
                selected_date=selected_date
            )
        elif event.keysym == "Right" and page < total_pages:
            show_profit_report(
                page=page+1,
                selected_period=selected_period,
                selected_date=selected_date
            )

    root.bind("<Left>", handle_keypress)
    root.bind("<Right>", handle_keypress)
def calculate_and_insert_consolidated_daily_actual_profit():
    connection = mysql.connector.connect(
        host="localhost",
        user="root",
        password="123@123",
        database="pharmacy"  
    )
    
    cursor = connection.cursor()
    
    try:
        # Get the current date
        current_date = datetime.now().date()
        
        # First, check if an entry for today already exists
        check_query = "SELECT profit_id FROM profit WHERE DATE(profit_date) = %s"
        cursor.execute(check_query, (current_date,))
        existing_entry = cursor.fetchone()
        
        if existing_entry:
            # Update existing entry
            update_query = """
            UPDATE profit SET
                total_sales = (
                    SELECT COALESCE(SUM(sales.total_price), 0)
                    FROM sales
                    WHERE DATE(sales.sale_date) = %s
                ),
                total_cost = (
                    SELECT COALESCE(SUM(medicines.buying_price * sales.quantity_sold), 0)
                    FROM sales
                    JOIN medicines ON sales.medicine_id = medicines.medicine_id
                    WHERE DATE(sales.sale_date) = %s
                ),
                total_refunds = (
                    SELECT COALESCE(SUM(returns.refund_amount), 0)
                    FROM returns
                    WHERE DATE(returns.return_date) = %s
                ),
                profit_amount = total_sales - total_cost - total_refunds
            WHERE DATE(profit_date) = %s
            """
            cursor.execute(update_query, (current_date, current_date, current_date, current_date))
            
        else:
            # Insert new entry
            insert_query = """
            INSERT INTO profit (total_sales, total_cost, total_refunds, profit_amount, profit_date)
            SELECT 
                COALESCE(SUM(sales.total_price), 0) AS total_sales,
                COALESCE(SUM(medicines.buying_price * sales.quantity_sold), 0) AS total_cost,
                COALESCE((SELECT SUM(refund_amount) FROM returns WHERE DATE(return_date) = %s), 0) AS total_refunds,
                COALESCE(SUM(sales.total_price) - SUM(medicines.buying_price * sales.quantity_sold), 0) 
                    - COALESCE((SELECT SUM(refund_amount) FROM returns WHERE DATE(return_date) = %s), 0) AS profit_amount,
                %s AS profit_date
            FROM sales
            JOIN medicines ON sales.medicine_id = medicines.medicine_id
            WHERE DATE(sales.sale_date) = %s
            """
            cursor.execute(insert_query, (current_date, current_date, current_date, current_date))
        
        connection.commit()
    
    except mysql.connector.Error as err:
        print(f"Error: {err}")
        connection.rollback()
    
    finally:
        cursor.close()
        connection.close()

def fetch_filtered_profit(selected_period="Lifetime", start_date=None, end_date=None):
    try:
        conn = mysql.connector.connect(
            host="localhost",
            user="root",
            password="123@123",
            database="pharmacy"
        )
        cursor = conn.cursor(prepared=True)

        query = """
        SELECT profit_id, total_sales, total_cost, total_refunds, profit_amount, DATE(profit_date) as profit_date
        FROM profit
        WHERE 1=1
        """
        params = []

        # Add date filtering based on selected period
        if selected_period != "Lifetime":
            end_date = datetime.now().date()
            if selected_period == "Today":
                start_date = end_date
            elif selected_period == "Last Week":
                start_date = end_date - timedelta(days=7)
            elif selected_period == "Last Month":
                start_date = end_date - timedelta(days=30)
            elif selected_period == "6 Months":
                start_date = end_date - timedelta(days=180)
            elif selected_period == "1 Year":
                start_date = end_date - timedelta(days=365)
            elif selected_period == "Other" and start_date and end_date:
                if isinstance(start_date, str):
                    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                if isinstance(end_date, str):
                    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            
            query += " AND DATE(profit_date) BETWEEN %s AND %s"
            params.extend([start_date, end_date])

        # Order by profit date, most recent first
        query += " ORDER BY profit_date DESC"

        cursor.execute(query, params)
        profits = cursor.fetchall()
        
        return profits

    except mysql.connector.Error as err:
        print(f"Error: {err}")
        return []

    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

def show_actual_profit_report(page=1, selected_period="Lifetime", search_term="", selected_date=None):
    clear_widgets(root)
    reset_grid_configuration(root)
    calculate_and_insert_consolidated_daily_actual_profit()
    
    ITEMS_PER_PAGE = 15  # Number of items per page
    
    # Configure column weights for proper alignment
    for col in range(13):  # Increased to 13 columns to accommodate the new "Refunds" column
        root.grid_columnconfigure(col, weight=1)

    # Title Section
    title_label = ctk.CTkLabel(
        root, 
        text="Profit Report", 
        font=("Arial", 24, "bold"),
        text_color="#2E86C1"
    )
    title_label.grid(row=0, column=0, columnspan=13, pady=20, sticky="nsew")

    # Back button section
    back_icon_image = ctk.CTkImage(Image.open("back.png"), size=(30, 30))
    
    back_frame = ctk.CTkFrame(root, fg_color="transparent")
    back_frame.grid(row=0, column=0, padx=20, pady=10, sticky="nw")
    
    back_button = ctk.CTkButton(
        back_frame,
        image=back_icon_image,
        text="",
        width=40,
        command=show_profit_management,
        fg_color="transparent",
        hover_color="#E5E7E9"
    )
    back_button.grid(row=0, column=0, padx=(0, 5), pady=0, sticky="nw")

    def show_calendar_popup():
        calendar_window = tk.Toplevel(root)
        calendar_window.title("Select Date Range")
        calendar_window.geometry("300x300")

        start_date_label = ctk.CTkLabel(calendar_window, text="From:", font=("Arial", 12, "bold"))
        start_date_label.pack(pady=(10, 0))

        start_cal = DateEntry(calendar_window, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        start_cal.pack(pady=(5, 10))

        end_date_label = ctk.CTkLabel(calendar_window, text="End Date:", font=("Arial", 12, "bold"))
        end_date_label.pack(pady=(10, 0))

        end_cal = DateEntry(calendar_window, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        end_cal.pack(pady=(5, 10))

        def on_select_dates():
            start_date = start_cal.get_date()
            end_date = end_cal.get_date()
            calendar_window.destroy()
            show_actual_profit_report(page=1, selected_period="Other", selected_date=(start_date, end_date))

        select_button = ctk.CTkButton(calendar_window, text="Select Dates", command=on_select_dates)
        select_button.pack(pady=20)

    def toggle_calendar(choice):
        if choice == "Other":
            show_calendar_popup()
        else:
            show_actual_profit_report(page=1, selected_period=choice)

    # Period combobox
    period_options = ["Today", "Last Week", "Last Month", "6 Months", "1 Year", "Lifetime", "Other"]
    period_combobox = ctk.CTkComboBox(
        root,
        values=period_options,
        command=toggle_calendar,
        width=150,
        height=32,
        border_color="#2E86C1",
        button_color="#2E86C1",
        dropdown_hover_color="#3498DB"
    )
    period_combobox.set(selected_period)
    period_combobox.grid(row=0, column=12, padx=20, pady=10, sticky="ne")

    # Header Section
    headers_profit = [
        ("Profit ID", 10, "center"),
        ("Total Sales", 15, "center"),
        ("Total Cost", 15, "center"),
        ("Refunds", 15, "center"),
        ("Profit Amount", 15, "center"),
        ("Date", 20, "center")
    ]

    # Create header frame
    header_frame = ctk.CTkFrame(root, fg_color="transparent")
    header_frame.grid(row=1, column=0, columnspan=13, sticky="nsew", padx=10)
    
    # Configure header columns
    total_weight = sum(weight for _, weight, _ in headers_profit)
    for col, (_, weight, _) in enumerate(headers_profit):
        header_frame.grid_columnconfigure(col, weight=weight)

    # Create headers
    for col, (header, _, _) in enumerate(headers_profit):
        header_label = ctk.CTkLabel(
            header_frame,
            text=header,
            font=("Arial", 14, "bold"),
            text_color="#FFFFFF",
            fg_color="#2E86C1",
            corner_radius=8,
            height=35,
            anchor="center"
        )
        header_label.grid(row=0, column=col, padx=2, pady=10, sticky="nsew")

    # Data frame
    data_frame = ctk.CTkFrame(
        root,
        fg_color="#F8F9F9",
        corner_radius=10
    )
    data_frame.grid(row=2, column=0, columnspan=13, padx=10, pady=10, sticky="nsew")

    # Configure data frame columns
    for col, (_, weight, _) in enumerate(headers_profit):
        data_frame.grid_columnconfigure(col, weight=weight)

    try:
        # Calculate offset for pagination
        offset = (page - 1) * ITEMS_PER_PAGE

        if selected_period == "Other" and selected_date:
            start_date, end_date = selected_date
            profit_data = fetch_filtered_profit("Other", start_date, end_date)
        else:
            profit_data = fetch_filtered_profit(selected_period)

        # Implement pagination
        total_records = len(profit_data)
        total_pages = ceil(total_records / ITEMS_PER_PAGE)
        paginated_data = profit_data[offset:offset + ITEMS_PER_PAGE]

        # Update total entries label
        total_entries_label = ctk.CTkLabel(
            root,
            text=f"Total Entries: {total_records}",
            font=("Arial", 14, "bold"),
            text_color="#2E86C1"
        )
        total_entries_label.grid(row=0, column=11, padx=(0, 10), pady=10, sticky="ne")

        if paginated_data:
            for i, profit_item in enumerate(paginated_data):
                if isinstance(profit_item, tuple) and len(profit_item) == 6:
                    profit_id, total_sales, total_cost, total_refunds, profit_amount, profit_date = profit_item

                    # Format profit date
                    profit_date = profit_date.strftime('%Y-%m-%d') if profit_date else "N/A"

                    # Format monetary values
                    sales_fmt = f"Rs{float(total_sales):,.2f}"
                    cost_fmt = f"Rs{float(total_cost):,.2f}"
                    refunds_fmt = f"Rs{float(total_refunds):,.2f}"
                    profit_fmt = f"Rs{float(profit_amount):,.2f}"

                    fields = [
                        (str(profit_id), "center"),
                        (sales_fmt, "center"),
                        (cost_fmt, "center"),
                        (refunds_fmt, "center"),
                        (profit_fmt, "center"),
                        (profit_date, "center")
                    ]

                    for col, (field, align) in enumerate(fields):
                        profit_label = ctk.CTkLabel(
                            data_frame,
                            text=field,
                            font=("Arial", 13),
                            text_color="#34495E",
                            fg_color="#FFFFFF" if i % 2 == 0 else "#F7F9F9",
                            corner_radius=8,
                            anchor=align,
                            height=30
                        )
                        padx = (10, 5) if align == "w" else (5, 10) if align == "e" else (5, 5)
                        profit_label.grid(
                            row=i,
                            column=col,
                            padx=padx,
                            pady=2,
                            sticky="nsew"
                        )
        else:
            ctk.CTkLabel(
                data_frame,
                text="No profit entries found for the selected period.",
                font=("Arial", 13, "bold"),
                text_color="#E74C3C",
                anchor="center",
                height=40
            ).grid(row=0, column=0, columnspan=len(headers_profit), pady=20, sticky="nsew")

        # Pagination frame
        pagination_frame = ctk.CTkFrame(root, fg_color="transparent")
        pagination_frame.grid(row=3, column=0, columnspan=13, pady=10, sticky="nsew")

        def change_page(new_page):
            show_actual_profit_report(
                page=new_page,
                selected_period=selected_period,
                selected_date=selected_date
            )

        # Previous page button
        prev_button = ctk.CTkButton(
            pagination_frame,
            text="←",
            width=40,
            command=lambda: change_page(page - 1) if page > 1 else None,
            fg_color="#2E86C1" if page > 1 else "#B3B6B7",
            state="normal" if page > 1 else "disabled"
        )
        prev_button.grid(row=0, column=0, padx=5)

        # Page numbers
        page_label = ctk.CTkLabel(
            pagination_frame,
            text=f"Page {page} of {total_pages}",
            font=("Arial", 12, "bold")
        )
        page_label.grid(row=0, column=1, padx=10)

        # Next page button
        next_button = ctk.CTkButton(
            pagination_frame,
            text="→",
            width=40,
            command=lambda: change_page(page + 1) if page < total_pages else None,
            fg_color="#2E86C1" if page < total_pages else "#B3B6B7",
            state="normal" if page < total_pages else "disabled"
        )
        next_button.grid(row=0, column=2, padx=5)

    except Exception as e:
        error_label = ctk.CTkLabel(
            data_frame,
            text=f"Error fetching profit data: {e}",
            font=("Arial", 13, "bold"),
            text_color="#E74C3C",
            anchor="center",
            height=40
        )
        error_label.grid(row=0, column=0, columnspan=len(headers_profit), pady=20, sticky="nsew")

        # Empty pagination frame in case of error
        pagination_frame = ctk.CTkFrame(root, fg_color="transparent")
        pagination_frame.grid(row=3, column=0, columnspan=13, pady=10, sticky="nsew")

    # Add a status bar showing records per page
    status_frame = ctk.CTkFrame(root, fg_color="transparent")
    status_frame.grid(row=4, column=0, columnspan=13, pady=(5, 10), sticky="nsew")

    records_label = ctk.CTkLabel(
        status_frame,
        text=f"Showing {ITEMS_PER_PAGE} records per page",
        font=("Arial", 12),
        text_color="#7F8C8D"
    )
    records_label.pack(side="right", padx=20)

    # Add export button
    def export_to_excel():
        try:
            # Check if required modules are installed
            try:
                import pandas as pd
                import openpyxl
            except ImportError as e:
                if 'openpyxl' in str(e):
                    if messagebox.askyesno("Missing Dependency", 
                        "The openpyxl module is required for Excel export. Would you like to install it now?"):
                        import subprocess
                        try:
                            subprocess.check_call(['pip', 'install', 'openpyxl'])
                            messagebox.showinfo("Success", "openpyxl installed successfully. Please try exporting again.")
                        except subprocess.CalledProcessError:
                            messagebox.showerror("Installation Failed", 
                                "Could not install openpyxl. Please install it manually using 'pip install openpyxl'")
                    return
                else:
                    messagebox.showerror("Missing Dependency", f"Required module not found: {str(e)}")
                    return

            # Fetch all records for export
            if selected_period == "Other" and selected_date:
                profit_data = fetch_filtered_profit("Other", selected_date[0], selected_date[1])
            else:
                profit_data = fetch_filtered_profit(selected_period)

            if not profit_data:
                messagebox.showwarning("Export Failed", "No records to export.")
                return

            # Convert records to list of dicts
            formatted_records = []
            for record in profit_data:
                profit_id, total_sales, total_cost, total_refunds, profit_amount, profit_date = record
                formatted_records.append({
                    'Profit ID': profit_id,
                    'Total Sales': f"Rs{float(total_sales):,.2f}",
                    'Total Cost': f"Rs{float(total_cost):,.2f}",
                    'Refunds': f"Rs{float(total_refunds):,.2f}",
                    'Profit Amount': f"Rs{float(profit_amount):,.2f}",
                    'Date': profit_date.strftime('%Y-%m-%d') if profit_date else "N/A"
                })

            # Create DataFrame
            df = pd.DataFrame(formatted_records)

            # Ask user for save location and format
            file_types = [
                ("Excel files", "*.xlsx"),
                ("CSV files", "*.csv"),
                ("Text files", "*.txt")
            ]
            file_path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=file_types,
                title="Export Profit Report"
            )

            if file_path:
                if file_path.endswith('.xlsx'):
                    # Excel export with formatting
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='Profit Report')
                        workbook = writer.book
                        worksheet = writer.sheets['Profit Report']
                        
                        # Auto-adjust column widths
                        for column in worksheet.columns:
                            max_length = 0
                            column = [cell for cell in column]
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(cell.value)
                                except:
                                    pass
                            adjusted_width = (max_length + 2)
                            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

                        # Add formatting
                        header = openpyxl.styles.PatternFill(
                            start_color='1E88E5',
                            end_color='1E88E5',
                            fill_type='solid'
                        )
                        for cell in worksheet[1]:
                            cell.fill = header
                            cell.font = openpyxl.styles.Font(color='FFFFFF', bold=True)

                elif file_path.endswith('.csv'):
                    df.to_csv(file_path, index=False)
                elif file_path.endswith('.txt'):
                    df.to_csv(file_path, index=False, sep='\t')

                messagebox.showinfo("Export Successful", 
                    f"Report exported successfully to:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Export Failed", 
                f"Error exporting data: {str(e)}\n\nPlease make sure you have permission to write to the selected location.")

    # Add export button
    export_button = ctk.CTkButton(
        status_frame,
        text="Export Report",
        command=export_to_excel,
        width=120,
        height=30,
        fg_color="#27AE60",
        hover_color="#219A52"
    )
    export_button.pack(side="left", padx=20)

    # Add keyboard shortcuts for navigation
    def handle_keypress(event):
        if event.keysym == "Left" and page > 1:
            show_actual_profit_report(
                page=page-1,
                selected_period=selected_period,
                selected_date=selected_date
            )
        elif event.keysym == "Right" and page < total_pages:
            show_actual_profit_report(
                page=page+1,
                selected_period=selected_period,
                selected_date=selected_date
            )

    root.bind("<Left>", handle_keypress)
    root.bind("<Right>", handle_keypress)
def fetch_filtered_profit_with_medicines(selected_period="Lifetime", start_date=None, end_date=None, search_term=""):
    try:
        conn = mysql.connector.connect(
            host="localhost",
            user="root",
            password="123@123",
            database="pharmacy"
        )
        cursor = conn.cursor(prepared=True)

        query = """
        SELECT 
            medicines.name AS medicine_name,
            SUM(sales.quantity_sold) as total_quantity,
            SUM(sales.total_price) as total_sales,
            SUM(medicines.buying_price * sales.quantity_sold) as total_cost,
            SUM(sales.total_price - (medicines.buying_price * sales.quantity_sold)) as profit_amount,
            DATE(sales.sale_date) as sale_date
        FROM 
            sales
        JOIN 
            medicines ON sales.medicine_id = medicines.medicine_id
        WHERE 1=1
        """
        params = []

        # Add date filtering based on selected period
        if selected_period != "Lifetime":
            end_date = datetime.now().date()
            if selected_period == "Today":
                start_date = end_date
            elif selected_period == "Last Week":
                start_date = end_date - timedelta(days=7)
            elif selected_period == "Last Month":
                start_date = end_date - timedelta(days=30)
            elif selected_period == "6 Months":
                start_date = end_date - timedelta(days=180)
            elif selected_period == "1 Year":
                start_date = end_date - timedelta(days=365)
            elif selected_period == "Other" and start_date and end_date:
                if isinstance(start_date, str):
                    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                if isinstance(end_date, str):
                    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            
            query += " AND DATE(sales.sale_date) BETWEEN %s AND %s"
            params.extend([start_date, end_date])

        # Add search term filtering
        if search_term:
            query += " AND medicines.name LIKE %s"
            params.append(f"%{search_term}%")

        query += """
        GROUP BY medicines.medicine_id, DATE(sales.sale_date)
        ORDER BY profit_amount DESC, sale_date DESC
        """

        cursor.execute(query, params)
        profits = cursor.fetchall()
        
        return profits

    except mysql.connector.Error as err:
        print(f"Error: {err}")
        return []

    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()
def show_daily_medicine_profit_report(page=1, selected_period="Lifetime", search_term="", selected_date=None):
    clear_widgets(root)
    reset_grid_configuration(root)
    
    ITEMS_PER_PAGE = 15  # Number of items per page
    
    # Configure column weights for proper alignment
    for col in range(12):
        root.grid_columnconfigure(col, weight=1)

    # Title Section
    title_label = ctk.CTkLabel(
        root, 
        text="Profit Report by Medicine", 
        font=("Arial", 24, "bold"),
        text_color="#2E86C1"
    )
    title_label.grid(row=0, column=0, columnspan=12, pady=20, sticky="nsew")

    # Back button section
    back_icon_image = ctk.CTkImage(Image.open("back.png"), size=(30, 30))
    
    back_frame = ctk.CTkFrame(root, fg_color="transparent")
    back_frame.grid(row=0, column=0, padx=20, pady=10, sticky="nw")
    
    back_button = ctk.CTkButton(
        back_frame,
        image=back_icon_image,
        text="",
        width=40,
        command=show_profit_management,
        fg_color="transparent",
        hover_color="#E5E7E9"
    )
    back_button.grid(row=0, column=0, padx=(0, 5), pady=0, sticky="nw")

    # Search functionality
    search_var = tk.StringVar()
    search_entry = ctk.CTkEntry(
        back_frame,
        placeholder_text="🔍 Search Medicine",
        width=200,
        height=32,
        border_color="#2E86C1",
        corner_radius=8,
        textvariable=search_var
    )
    search_entry.grid(row=0, column=1, padx=(0, 5), pady=6, sticky="nw")
    search_entry.insert(0, search_term)
    search_entry.focus_set()

    def update_report(*args):
        search_term = search_var.get().strip()
        show_daily_medicine_profit_report(page=1, selected_period=selected_period, search_term=search_term, selected_date=selected_date)

    search_var.trace_add("write", update_report)

    def show_calendar_popup():
        calendar_window = tk.Toplevel(root)
        calendar_window.title("Select Date Range")
        calendar_window.geometry("300x300")

        start_date_label = ctk.CTkLabel(calendar_window, text="From:", font=("Arial", 12, "bold"))
        start_date_label.pack(pady=(10, 0))

        start_cal = DateEntry(calendar_window, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        start_cal.pack(pady=(5, 10))

        end_date_label = ctk.CTkLabel(calendar_window, text="End Date:", font=("Arial", 12, "bold"))
        end_date_label.pack(pady=(10, 0))

        end_cal = DateEntry(calendar_window, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        end_cal.pack(pady=(5, 10))

        def on_select_dates():
            start_date = start_cal.get_date()
            end_date = end_cal.get_date()
            calendar_window.destroy()
            show_daily_medicine_profit_report(page=1, selected_period="Other", search_term=search_var.get(), selected_date=(start_date, end_date))

        select_button = ctk.CTkButton(calendar_window, text="Select Dates", command=on_select_dates)
        select_button.pack(pady=20)

    def toggle_calendar(choice):
        if choice == "Other":
            show_calendar_popup()
        else:
            show_daily_medicine_profit_report(page=1, selected_period=choice, search_term=search_var.get())

    # Period combobox
    period_options = ["Today", "Last Week", "Last Month", "6 Months", "1 Year", "Lifetime", "Other"]
    period_combobox = ctk.CTkComboBox(
        root,
        values=period_options,
        command=toggle_calendar,
        width=150,
        height=32,
        border_color="#2E86C1",
        button_color="#2E86C1",
        dropdown_hover_color="#3498DB"
    )
    period_combobox.set(selected_period)
    period_combobox.grid(row=0, column=11, padx=20, pady=10, sticky="ne")

    # Header Section
    headers_profit = [
        ("Medicine Name", 10, "center"),
        ("Quantity Sold", 10, "center"),
        ("Total Sales", 15, "center"),
        ("Total Cost", 15, "center"),
        ("Profit Amount", 15, "center"),
        ("Date", 15, "center")
    ]

    # Create header frame
    header_frame = ctk.CTkFrame(root, fg_color="transparent")
    header_frame.grid(row=1, column=0, columnspan=12, sticky="nsew", padx=10)
    
    # Configure header columns
    total_weight = sum(weight for _, weight, _ in headers_profit)
    for col, (_, weight, _) in enumerate(headers_profit):
        header_frame.grid_columnconfigure(col, weight=weight)

    # Create headers
    for col, (header, _, _) in enumerate(headers_profit):
        header_label = ctk.CTkLabel(
            header_frame,
            text=header,
            font=("Arial", 14, "bold"),
            text_color="#FFFFFF",
            fg_color="#2E86C1",
            corner_radius=8,
            height=35,
            anchor="center"
        )
        header_label.grid(row=0, column=col, padx=2, pady=10, sticky="nsew")

    # Data frame
    data_frame = ctk.CTkFrame(
        root,
        fg_color="#F8F9F9",
        corner_radius=10
    )
    data_frame.grid(row=2, column=0, columnspan=12, padx=10, pady=10, sticky="nsew")

    # Configure data frame columns
    for col, (_, weight, _) in enumerate(headers_profit):
        data_frame.grid_columnconfigure(col, weight=weight)

    try:
        # Calculate offset for pagination
        offset = (page - 1) * ITEMS_PER_PAGE

        if selected_period == "Other" and selected_date:
            start_date, end_date = selected_date
            profit_data = fetch_filtered_profit_with_medicines("Other", start_date, end_date, search_term)
        else:
            profit_data = fetch_filtered_profit_with_medicines(selected_period, search_term=search_term)

        # Implement pagination
        total_records = len(profit_data)
        total_pages = ceil(total_records / ITEMS_PER_PAGE)
        paginated_data = profit_data[offset:offset + ITEMS_PER_PAGE]

        # Update total entries label
        total_entries_label = ctk.CTkLabel(
            root,
            text=f"Total Entries: {total_records}",
            font=("Arial", 14, "bold"),
            text_color="#2E86C1"
        )
        total_entries_label.grid(row=0, column=10, padx=(0, 10), pady=10, sticky="ne")

        if paginated_data:
            for i, profit_item in enumerate(paginated_data):
                medicine_name, quantity_sold, total_sales, total_cost, profit_amount, sale_date = profit_item

                # Format sale date
                sale_date = sale_date.strftime('%Y-%m-%d') if sale_date else "N/A"

                # Format monetary values
                sales_fmt = f"Rs{float(total_sales):,.2f}"
                cost_fmt = f"Rs{float(total_cost):,.2f}"
                profit_fmt = f"Rs{float(profit_amount):,.2f}"

                fields = [
                    (str(medicine_name), "center"),
                    (str(quantity_sold), "center"),
                    (sales_fmt, "center"),
                    (cost_fmt, "center"),
                    (profit_fmt, "center"),
                    (sale_date, "center")
                ]

                for col, (field, align) in enumerate(fields):
                    profit_label = ctk.CTkLabel(
                        data_frame,
                        text=field,
                        font=("Arial", 13),
                        text_color="#34495E",
                        fg_color="#FFFFFF" if i % 2 == 0 else "#F7F9F9",
                        corner_radius=8,
                        anchor=align,
                        height=30
                    )
                    padx = (10, 5) if align == "w" else (5, 10) if align == "e" else (5, 5)
                    profit_label.grid(
                        row=i,
                        column=col,
                        padx=padx,
                        pady=2,
                        sticky="nsew"
                    )
        else:
            ctk.CTkLabel(
                data_frame,
                text="No medicine profit entries found for the selected period.",
                font=("Arial", 13, "bold"),
                text_color="#E74C3C",
                anchor="center",
                height=40
            ).grid(row=0, column=0, columnspan=len(headers_profit), pady=20, sticky="nsew")

        # Pagination frame
        pagination_frame = ctk.CTkFrame(root, fg_color="transparent")
        pagination_frame.grid(row=3, column=0, columnspan=12, pady=10, sticky="nsew")

        def change_page(new_page):
            show_daily_medicine_profit_report(
                page=new_page,
                selected_period=selected_period,
                search_term=search_term,
                selected_date=selected_date
            )

        # Previous page button
        prev_button = ctk.CTkButton(
            pagination_frame,
            text="←",
            width=40,
            command=lambda: change_page(page - 1) if page > 1 else None,
            fg_color="#2E86C1" if page > 1 else "#B3B6B7",
            state="normal" if page > 1 else "disabled"
        )
        prev_button.grid(row=0, column=0, padx=5)

        # Page numbers
        page_label = ctk.CTkLabel(
            pagination_frame,
            text=f"Page {page} of {total_pages}",
            font=("Arial", 12, "bold")
        )
        page_label.grid(row=0, column=1, padx=10)

        # Next page button
        next_button = ctk.CTkButton(
            pagination_frame,
            text="→",
            width=40,
            command=lambda: change_page(page + 1) if page < total_pages else None,
            fg_color="#2E86C1" if page < total_pages else "#B3B6B7",
            state="normal" if page < total_pages else "disabled"
        )
        next_button.grid(row=0, column=2, padx=5)

    except Exception as e:
        error_label = ctk.CTkLabel(
            data_frame,
            text=f"Error fetching medicine profit data: {e}",
            font=("Arial", 13, "bold"),
            text_color="#E74C3C",
            anchor="center",
            height=40
        )
        error_label.grid(row=0, column=0, columnspan=len(headers_profit), pady=20, sticky="nsew")

        # Empty pagination frame in case of error
        pagination_frame = ctk.CTkFrame(root, fg_color="transparent")
        pagination_frame.grid(row=3, column=0, columnspan=12, pady=10, sticky="nsew")

    # Add a status bar showing records per page
    status_frame = ctk.CTkFrame(root, fg_color="transparent")
    status_frame.grid(row=4, column=0, columnspan=12, pady=(5, 10), sticky="nsew")

    records_label = ctk.CTkLabel(
        status_frame,
        text=f"Showing {ITEMS_PER_PAGE} records per page",
        font=("Arial", 12),
        text_color="#7F8C8D"
    )
    records_label.pack(side="right", padx=20)

    # Add export button with the same functionality as the first code
    def export_to_excel():
        try:
            try:
                import pandas as pd
                import openpyxl
            except ImportError as e:
                if 'openpyxl' in str(e):
                    if messagebox.askyesno("Missing Dependency", 
                        "The openpyxl module is required for Excel export. Would you like to install it now?"):
                        import subprocess
                        try:
                            subprocess.check_call(['pip', 'install', 'openpyxl'])
                            messagebox.showinfo("Success", "openpyxl installed successfully. Please try exporting again.")
                        except subprocess.CalledProcessError:
                            messagebox.showerror("Installation Failed", 
                                "Could not install openpyxl. Please install it manually using 'pip install openpyxl'")
                    return
                else:
                    messagebox.showerror("Missing Dependency", f"Required module not found: {str(e)}")
                    return

            if selected_period == "Other" and selected_date:
                profit_data = fetch_filtered_profit_with_medicines("Other", selected_date[0], selected_date[1], search_term)
            else:
                profit_data = fetch_filtered_profit_with_medicines(selected_period, search_term=search_term)

            if not profit_data:
                messagebox.showwarning("Export Failed", "No records to export.")
                return

            # Convert records to list of dicts
            formatted_records = []
            for record in profit_data:
                medicine_name, quantity_sold, total_sales, total_cost, profit_amount, sale_date = record
                formatted_records.append({
                    'Medicine Name': medicine_name,
                    'Quantity Sold': quantity_sold,
                    'Total Sales': f"Rs{float(total_sales):,.2f}",
                    'Total Cost': f"Rs{float(total_cost):,.2f}",
                    'Profit Amount': f"Rs{float(profit_amount):,.2f}",
                    'Date': sale_date.strftime('%Y-%m-%d') if sale_date else "N/A"
                })

            # Create DataFrame
            df = pd.DataFrame(formatted_records)

            # Ask user for save location and format
            file_types = [
                ("Excel files", "*.xlsx"),
                ("CSV files", "*.csv"),
                ("Text files", "*.txt")
            ]
            file_path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=file_types,
                title="Export Medicine Profit Report"
            )

            if file_path:
                if file_path.endswith('.xlsx'):
                    # Excel export with formatting
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='Medicine Profit Report')
                        workbook = writer.book
                        worksheet = writer.sheets['Medicine Profit Report']
                        
                        # Auto-adjust column widths
                        for column in worksheet.columns:
                            max_length = 0
                            column = [cell for cell in column]
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(cell.value)
                                except:
                                    pass
                            adjusted_width = (max_length + 2)
                            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

                        # Add formatting
                        header = openpyxl.styles.PatternFill(
                            start_color='1E88E5',
                            end_color='1E88E5',
                            fill_type='solid'
                        )
                        for cell in worksheet[1]:
                            cell.fill = header
                            cell.font = openpyxl.styles.Font(color='FFFFFF', bold=True)

                elif file_path.endswith('.csv'):
                    df.to_csv(file_path, index=False)
                elif file_path.endswith('.txt'):
                    df.to_csv(file_path, index=False, sep='\t')

                messagebox.showinfo("Export Successful", 
                    f"Report exported successfully to:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Export Failed", 
                f"Error exporting data: {str(e)}\n\nPlease make sure you have permission to write to the selected location.")

    # Add export button
    export_button = ctk.CTkButton(
        status_frame,
        text="Export Report",
        command=export_to_excel,
        width=120,
        height=30,
        fg_color="#27AE60",
        hover_color="#219A52"
    )
    export_button.pack(side="left", padx=20)

    # Add keyboard shortcuts for navigation
    def handle_keypress(event):
        if event.keysym == "Left" and page > 1:
            show_daily_medicine_profit_report(
                page=page-1,
                selected_period=selected_period,
                search_term=search_term,
                selected_date=selected_date
            )
        elif event.keysym == "Right" and page < total_pages:
            show_daily_medicine_profit_report(
                page=page+1,
                selected_period=selected_period,
                search_term=search_term,
                selected_date=selected_date
            )

    root.bind("<Left>", handle_keypress)
    root.bind("<Right>", handle_keypress)
def load_image(image_path, size):
    if os.path.exists(image_path):
        return ImageTk.PhotoImage(Image.open(image_path).resize(size))
    else:
        print(f"Image not found: {image_path}")
        return None# Create the main window
app = ctk.CTk()
app.title("Pharmacy Management System")
screen1 = app.winfo_screenwidth()
screen2 = app.winfo_screenheight()
app.geometry(f"{screen1}x{screen2}+0+0")
root = ctk.CTkFrame(app)
root.place(relx=0, rely=0, relwidth=1.0, relheight=1.0)

def security():
    def get_mac_address():
        try:
            mac = ':'.join(['{:02x}'.format((uuid.getnode() >> elements) & 0xff) for elements in range(0, 2 * 6, 2)][::-1])
            print(f"Retrieved MAC address: {mac}")
            return mac
        except Exception as e:
            print(f"Error retrieving MAC address: {e}")
            return None

    def connect_to_db():
        try:
            conn = mysql.connector.connect(
                host="localhost",
                user="root",
                password="123@123",
                database="pharmacy"
            )
            print("Database connection established.")
            return conn
        except mysql.connector.Error as e:
            print(f"Database connection error: {e}")
            return None

    def insert_license_key(username, machine_id, license_key):
        try:
            conn = connect_to_db()
            if conn is None:
                return
            
            cursor = conn.cursor()
            cursor.execute("INSERT INTO licenses (username, machine_id, license_key) VALUES (%s,%s,%s)",
                        (username, machine_id, license_key))
            conn.commit()
            print("License key inserted successfully.")
            conn.close()
        except mysql.connector.Error as e:
            print(f"Database error: {e}")

    def is_activated():
        try:
            conn = connect_to_db()
            if conn is None:
                return False
            
            cursor = conn.cursor()
            cursor.execute("SELECT activated FROM licenses WHERE activated = 1")
            result = cursor.fetchone()
            conn.close()
            
            return result is not None
        except mysql.connector.Error as e:
            print(f"Database error: {e}")
            return False

    def validate_license_key(username, entered_key):
        try:
            current_mac_address = get_mac_address()
            if not current_mac_address:
                messagebox.showerror("Error", "Failed to retrieve MAC address.")
                return

            conn = connect_to_db()
            if conn is None:
                messagebox.showerror("Error", "Failed to connect to database.")
                return

            cursor = conn.cursor()
            cursor.execute("SELECT machine_id, license_key FROM licenses WHERE username = %s", (username,))
            result = cursor.fetchone()

            if result:
                stored_machine_id, stored_license_key = result
                print(f"Stored Machine ID: {stored_machine_id}")
                print(f"Entered License Key: {entered_key}")
                print(f"Stored License Key: {stored_license_key}")
                
                if entered_key == stored_license_key:
                    if stored_machine_id.strip() == current_mac_address.strip():
                        cursor.execute("UPDATE licenses SET activated = 1 WHERE username = %s", (username,))
                        conn.commit()
                        messagebox.showinfo("Success", "Software Activated!")
                        show_main_menu()
                    else:
                        print("Machine ID mismatch.")
                        messagebox.showerror("Error", "Invalid Machine ID.")
                else:
                    print("Invalid license key.")
                    messagebox.showerror("Error", "Invalid license key.")
            else:
                print("Username not found.")
                messagebox.showerror("Error", "Username not found.")

            conn.close()
        except mysql.connector.Error as e:
            messagebox.showerror("Database Error", f"Database error: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Error: {e}")

    def on_startup():
        if is_activated():
            show_main_menu()
        else:
            clear_window()
            show_license_window()


    def show_license_window():
        username_label = ctk.CTkLabel(root, text="Enter Username:")
        username_label.place(relx=0.5, rely=0.3, anchor="center")

        username_entry = ctk.CTkEntry(root, width=200, placeholder_text="Enter your username...")
        username_entry.place(relx=0.5, rely=0.35, anchor="center")

        label = ctk.CTkLabel(root, text="Enter License Key:")
        label.place(relx=0.5, rely=0.4, anchor="center")

        license_key_entry = ctk.CTkEntry(root, width=200, placeholder_text="Enter the license key...")
        license_key_entry.place(relx=0.5, rely=0.45, anchor="center")

        submit_button = ctk.CTkButton(root, text="Submit", command=lambda: validate_license_key(username_entry.get(), license_key_entry.get()))
        submit_button.place(relx=0.5, rely=0.5, anchor="center")
    
    on_startup()


# Load icons for navigation buttons
Medicine  = load_image(r"pics\medicine.png", (30, 30))
Inventory = load_image(r"pics\inventory.png", (30, 30))
Billing = load_image(r"pics\bill.png", (30, 30))
Reporting = load_image(r"pics\reporting.png", (30, 30))
Reporting = load_image(r"pics\reporting.png", (30, 30))
Supplier = load_image(r"pics\supplier.png", (30, 30))
financelogo = load_image(r"pics\hand.png", (30, 30))

# Load vector graphics for main frame buttons
dashboard_bord = load_image(r"pics\OIP.jpg", (190, 170))
medicine_managment = load_image(r"pics\d2.jpg", (190, 170))
inventory_managment = load_image(r"pics\invent.jpg", (190, 170))
salesandbillings = load_image(r"pics\d4.jpg", (190, 170))
reportings = load_image(r"pics\d5.jpg", (190, 170))
shiffting= load_image(r"pics\alalal.jpg", (190, 170))
Finance = load_image(r"pics\finance.jpg", (190, 170))
suppliers = load_image(r"pics\Supplier.jpg", (190, 170))
search_pic = load_image(r"pics\search (1).png", (20, 20))

button_style = {
    'fg_color': '#a3b18a',
    'text_color': 'black',
    'width': 250,
    'height': 50,
    'hover_color': '#869974'
    
}

# Function to clear the window
def clear_window():
    if root.winfo_exists():
        for widget in root.winfo_children():
            widget.destroy()
    else:
        print("Frame does not exist")
        
# Navigation frame on the left
def nav_frame():
    # Create a frame for the navigation menu
    nav_frame = ctk.CTkFrame(root, width=250, bg_color="midnight blue", fg_color="#a3b18a", corner_radius=0)
    nav_frame.place(x=0, y=0, relheight=1)

    nav_top_photo = load_image(r"pics\image.png", (280, 280))
    if nav_top_photo:
        nav_photo_label = ctk.CTkLabel(nav_frame, image=nav_top_photo, text="")
        nav_photo_label.place(x=0, y=0)

    # Add navigation buttons with separate commands
    dashboard_button = ctk.CTkButton(nav_frame, text="Medicine Management", image=Medicine, compound="left", anchor="w", **button_style, command=show_medicine_management)
    dashboard_button.place(x=0, y=180)

    inventory_button = ctk.CTkButton(nav_frame, text="Inventory Management", image=Inventory, compound="left", anchor="w", **button_style, command=show_inventory)
    inventory_button.place(x=0, y=220)

    finance_button = ctk.CTkButton(nav_frame, text="Sales & Billing", image=Billing, compound="left", anchor="w", **button_style, command=sales_and_billing)
    finance_button.place(x=0, y=260)

    report_button = ctk.CTkButton(nav_frame, text="Reporting", image=Reporting, compound="left", anchor="w", **button_style, command=show_reporting)
    report_button.place(x=0, y=300)

    supplier_button = ctk.CTkButton(nav_frame, text="Supplier", image=Supplier, compound="left", anchor="w", **button_style, command=show_supplier)
    supplier_button.place(x=0, y=340)

    supplier_button = ctk.CTkButton(nav_frame, text="Finance", image=financelogo, compound="left", anchor="w", **button_style, command=finance)
    supplier_button.place(x=0, y=380)
from tkinter import Listbox  # Import Listbox here
def search_inventory(query):
    # Establish a connection to the MySQL database
    conn = mysql.connector.connect(
        host="localhost",        # Your MySQL host
        user="root",             # Your MySQL username
        password="123@123", # Your MySQL password
        database="pharmacy"          # The name of your database
    )
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM medicines WHERE name LIKE %s", ('%' + query + '%',))
    results = cursor.fetchall()
    conn.close()
    return [result[0] for result in results]  # Return a list of medicine names

# Function to update the search results in real-time
def update_search_results(event, search_entry, result_listbox):
    query = search_entry.get()
    if query:  # If the user has entered something
        results = search_inventory(query)  # Fetch matching medicines from the database
        result_listbox.delete(0, tk.END)  # Clear previous results
        for result in results:
            result_listbox.insert(tk.END, result)  # Insert new results
        result_listbox.place(x=20, y=112)  # Show the dropdown list
    else:
        result_listbox.place_forget()  # Hide the listbox when the query is empty

# Function to handle selection of an item from the dropdown list
def on_select(event, search_entry, result_listbox):
    selected = result_listbox.get(result_listbox.curselection())
    search_entry.delete(0, tk.END)  # Clear the search bar
    search_entry.insert(0, selected)  # Insert the selected item into the search bar
    result_listbox.place_forget()  # Hide the dropdown after selection

def main_frame():
    main = ctk.CTkFrame(root, fg_color="#dad7cd")
    main.place(x=250, y=0, relwidth=0.9, relheight=1)

    # Search bar and search button under the title label
    search_entry = ctk.CTkEntry(main, width=1050, height=35, corner_radius=50, border_color="#a3b18a", font=("Arial", 13),
                                placeholder_text_color="#869974", placeholder_text="Search medicine.....")
    search_entry.place(x=20, y=80)

    # Listbox to display the search results (Dropdown list)
    result_listbox =Listbox(main, width=162, height=5 ,font=("Arial", 12), fg='black', bg='#f1faee', bd=0, highlightthickness=0)

    # Bind events: Key release to update search and selection from the list
    search_entry.bind("<KeyRelease>", lambda event: update_search_results(event, search_entry, result_listbox))
    result_listbox.bind("<Double-Button-1>", lambda event: on_select(event, search_entry, result_listbox))

    # Add buttons to the main frame with separate commands
    dashboard_button = ctk.CTkButton(main,hover_color="#f1faee",text="Dashboard",text_color="black", fg_color="white",corner_radius=10, image=dashboard_bord, compound="top",
                                    font=("Arial", 13), command=show_dashboard)
    dashboard_button.place(x=40, y=200)

    pharmacist_button = ctk.CTkButton(main,hover_color="#f1faee", text="Inventory",text_color="black", fg_color="white", image=inventory_managment, compound="top",
                                    font=("Arial", 13), corner_radius=10, command=show_inventory)
    pharmacist_button.place(x=320, y=200)

    manager_button = ctk.CTkButton(main,hover_color="#f1faee", text="Medicine",text_color="black", fg_color="white", image=medicine_managment, compound="top",
                                font=("Arial", 13), corner_radius=10, command=show_medicine)
    manager_button.place(x=600, y=200)

    cashier_button = ctk.CTkButton(main,hover_color="#f1faee", text="Sale & Billing",text_color="black", fg_color="white", image=salesandbillings, compound="top",
                                font=("Arial", 13), corner_radius=10, command=show_sales)
    cashier_button.place(x=880, y=200)

    local_data_button = ctk.CTkButton(main,hover_color="#f1faee", text="Report",text_color="black", fg_color="white", image=reportings, compound="top",
                                font=("Arial", 13), command=show_report, corner_radius=10)
    local_data_button.place(x=40, y=490)

    search_button_main = ctk.CTkButton(main,hover_color="#f1faee", text="Supplier",text_color="black", fg_color="white", image=suppliers, compound="top",
                                    font=("Arial", 13), corner_radius=10, command=show_supplier)
    search_button_main.place(x=320, y=490)
    
    search_button_main1 = ctk.CTkButton(main,hover_color="#f1faee", text="Shifting",text_color="black", fg_color="white", image=shiffting, compound="top",
                                    font=("Arial", 13), corner_radius=10, command=show_shift)
    search_button_main1.place(x=600, y=490)
    
    search_button_main2 = ctk.CTkButton(main,hover_color="#f1faee", text="Finance",text_color="black", fg_color="white", image=Finance, compound="top",
                                    font=("Arial", 13), corner_radius=10, command=show_finance)
    search_button_main2.place(x=880, y=490)


    """Creates the header for the main content."""
    header = ctk.CTkFrame(main, height=70,width=1200, fg_color="#588157", corner_radius=0)
    header.place(x=0,y=0)

    title_label = ctk.CTkLabel(header, text="Tahir Pharmacy",
                               font=("montserrat", 34, "italic"), text_color="white")
    title_label.place(x=350,y=19)
# Define separate functions for each button
def show_dashboard():
    clear_window()
    show_combined_medicines_report()    

def show_inventory():
    clear_window()
    nav_frame()
    show_inventory_management()


def show_medicine():
    clear_window()
    nav_frame()
    show_medicine_management()


def show_sales():
    clear_window()
    nav_frame()
    sales_and_billing()
    

def show_report():
    clear_window()
    nav_frame()
    show_reporting()
    
def show_shift():
    clear_window()
    nav_frame()
    Shifting()
    
def show_supplier():
    nav_frame()
    show_supplier_management()
    
def show_finance():
    nav_frame()
    finance()
    
def back_button():
    back_btn = ctk.CTkButton(root, text="Back", command=show_main_menu)
    back_btn.place(relx=0.5, rely=0.9, anchor="center")

def show_main_menu():
    clear_window()
    nav_frame()
    main_frame()
# Start the application
security()
app.mainloop()
